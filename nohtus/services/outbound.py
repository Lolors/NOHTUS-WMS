"""Outbound service helpers for NOHTUS WMS.

출고지시 화면/저장된 출고지시 화면에서 공통으로 쓰는 조회, 추천,
다운로드, 취소 관련 함수를 app.py에서 단계적으로 분리한다.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd

from nohtus.db import connect, q
from nohtus.dates import display_date_only
from nohtus.services.inventory import insert_transaction_log
from nohtus.services.inbound import product_mapping_name_for


def _location_sort_key(value):
    text = str(value or "")
    parts = text.split("-")
    key = []
    for part in parts:
        if part.isdigit():
            key.append((0, int(part)))
        else:
            key.append((1, part))
    return key


def sort_outbound_rows_for_picking(rows):
    return sorted(
        list(rows or []),
        key=lambda r: (
            _location_sort_key((r or {}).get("로케이션") or (r or {}).get("location")),
            str((r or {}).get("제품명") or (r or {}).get("product_name") or ""),
            str((r or {}).get("LOT") or (r or {}).get("lot") or ""),
            str((r or {}).get("유통기한") or (r or {}).get("exp_date") or ""),
        ),
    )


def outbound_erp_note_for_row(row, cache=None):
    cache = cache if cache is not None else {}
    company = str(row.get("사업장") or row.get("company") or row.get("사업체") or "").strip()
    product = str(row.get("제품명") or row.get("product_name") or "").strip()
    if not company or not product:
        return ""
    key = (company, product)
    if key in cache:
        return cache[key]
    try:
        val = product_mapping_name_for(company, product) or ""
    except Exception:
        val = ""
    cache[key] = val
    return val


def outbound_excel_bytes(rows, title="출고지시서"):
    rows = sort_outbound_rows_for_picking(rows)
    note_cache = {}
    rows = [{**(r or {}), "비고": outbound_erp_note_for_row(r or {}, note_cache)} for r in rows]
    df = pd.DataFrame(rows)
    cols = ["로케이션", "제품명", "LOT", "유통기한", "요청수량", "비고"]
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    out = df[cols].copy()
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="출고지시서", startrow=2)
        ws = writer.book["출고지시서"]
        ws["A1"] = title or "출고지시서"
        ws.merge_cells("A1:F1")
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")
        widths = {"A":18,"B":30,"C":18,"D":18,"E":12,"F":32}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows(min_row=3, max_row=3 + len(out), min_col=1, max_col=6):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", horizontal="center" if cell.column in [1,3,4,5] else "left", wrap_text=True)
                if cell.row == 3:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
        ws.auto_filter.ref = f"A3:F{3+len(out)}"
    bio.seek(0)
    return bio.getvalue()


def _find_korean_font():
    candidates = [
        r"C:\Windows\Fonts\malgun.ttf",
        r"C:\Windows\Fonts\malgunbd.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return str(path)
    return None


def outbound_pdf_bytes(rows, title="출고지시서"):
    rows = sort_outbound_rows_for_picking(rows)
    note_cache = {}
    rows = [{**(r or {}), "비고": outbound_erp_note_for_row(r or {}, note_cache)} for r in rows]
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    bio = BytesIO()
    font_name = "Helvetica"
    font_path = _find_korean_font()
    if font_path:
        try:
            pdfmetrics.registerFont(TTFont("NOHTUS_KR", font_path))
            font_name = "NOHTUS_KR"
        except Exception:
            font_name = "Helvetica"
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4), leftMargin=22, rightMargin=22, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = font_name
    styles["Normal"].fontName = font_name
    story = [Paragraph(title or "출고지시서", styles["Title"]), Spacer(1, 12)]
    cols = ["로케이션", "제품명", "LOT", "유통기한", "요청수량", "비고"]
    data = [cols]
    for r in rows:
        data.append([str(r.get(c, "")) for c in cols])
    table = Table(data, colWidths=[80, 220, 105, 105, 70, 160], repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), font_name),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E5E7EB")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#111827")),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
        ("ALIGN", (0,1), (0,-1), "CENTER"),
        ("ALIGN", (2,1), (4,-1), "CENTER"),
        ("ALIGN", (1,1), (1,-1), "LEFT"),
        ("ALIGN", (5,1), (5,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(table)
    doc.build(story)
    bio.seek(0)
    return bio.getvalue()


def recommend_picks(pick_df, request_qty, expiry_short_first=True):
    rows = []
    try:
        need = int(request_qty or 0)
    except Exception:
        need = 0
    if need <= 0 or pick_df is None or pick_df.empty:
        return rows, max(0, need)

    df = pick_df.copy()
    if expiry_short_first:
        df["_exp_sort"] = pd.to_datetime(df.get("exp_date"), errors="coerce")
        df["_exp_sort"] = df["_exp_sort"].fillna(pd.Timestamp.max)
        df = df.sort_values(["_exp_sort", "location", "lot", "company"], na_position="last")
    else:
        df = df.sort_values(["location", "lot", "company", "exp_date"], na_position="last")

    for r in df.itertuples():
        if need <= 0:
            break
        available = int(getattr(r, "qty", 0) or 0)
        if available <= 0:
            continue
        take = min(available, need)
        rows.append({
            "id": int(getattr(r, "id")),
            "로케이션": getattr(r, "location", ""),
            "사업장": getattr(r, "company", ""),
            "제품명": getattr(r, "product_name", ""),
            "LOT": getattr(r, "lot", "-") or "-",
            "유통기한": display_date_only(getattr(r, "exp_date", "-")),
            "요청수량": int(take),
        })
        need -= take
    return rows, max(0, need)


def build_outbound_order_title(customer_name, cart_items, fallback_title=""):
    customer = str(customer_name or "").strip()
    items = list(cart_items or [])
    if not items:
        return str(fallback_title or "").strip()
    first = items[0] or {}
    first_name = str(first.get("제품명") or first.get("product_name") or "").strip()
    if customer and first_name:
        title = f"{customer} - {first_name}"
    elif customer:
        title = customer
    else:
        title = first_name
    rest_count = max(0, len(items) - 1)
    if rest_count:
        title = f"{title} 외 {rest_count}품목"
    return title or str(fallback_title or "").strip()


def load_outbound_order(order_id):
    df = q(
        """
        SELECT inventory_id AS id, location AS 로케이션, product_name AS 제품명,
               lot AS LOT, exp_date AS 유통기한, qty AS 요청수량,
               company AS 사업장, warehouse_name AS '전산상 명칭'
        FROM outbound_order_items
        WHERE order_id=?
        ORDER BY id
        """,
        (int(order_id),),
    )
    return df.to_dict("records")


def _fallback_restore_outbound_from_items(cur, order_id, now):
    item_rows = cur.execute(
        """
        SELECT inventory_id, location, product_name, lot, exp_date, qty, company, warehouse_name
        FROM outbound_order_items
        WHERE order_id=?
        ORDER BY id
        """,
        (int(order_id),),
    ).fetchall()
    restored_count = 0
    for inv_id, location, product_name, lot, exp_date, qty, company, warehouse_name in item_rows:
        qty = int(qty or 0)
        if qty <= 0:
            continue
        inv = None
        if inv_id:
            inv = cur.execute("SELECT id, qty FROM inventory WHERE id=?", (int(inv_id),)).fetchone()
        if not inv:
            inv = cur.execute(
                """
                SELECT id, qty FROM inventory
                WHERE company=? AND product_name=? AND IFNULL(warehouse_name,'')=?
                  AND IFNULL(lot,'')=? AND IFNULL(exp_date,'')=? AND location=?
                """,
                (company or "", product_name or "", warehouse_name or "", lot or "", exp_date or "", location or ""),
            ).fetchone()
        if inv:
            row_qty_after = int(inv[1] or 0) + qty
            cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (row_qty_after, now, int(inv[0])))
        else:
            cur.execute(
                """
                INSERT INTO inventory(company, product_name, warehouse_name, lot, exp_date, location, qty, updated_at)
                VALUES(?,?,?,?,?,?,?,?)
                """,
                (company or "", product_name or "", warehouse_name or "", lot or "-", exp_date or "-", location or "", qty, now),
            )
        restored_count += 1
        insert_transaction_log(
            cur,
            created_at=now,
            tx_type="출고지시취소",
            product_name=product_name or "",
            warehouse_name=warehouse_name or "",
            lot=lot or "-",
            exp_date=exp_date or "-",
            from_company=company or "",
            from_location=location or "",
            to_company=company or "",
            to_location=location or "",
            qty=qty,
            memo=f"출고지시서 #{order_id} 취소 / 원복",
        )
    return len(item_rows), restored_count


def cancel_outbound_order(order_id):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    order_id = int(order_id)
    with connect() as con:
        cur = con.cursor()
        order = cur.execute("SELECT id, status FROM outbound_orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            raise ValueError("취소할 출고지시서를 찾을 수 없습니다.")
        if str(order[1] or "") == "취소됨":
            raise ValueError("이미 취소된 출고지시서입니다.")

        already_cancelled = cur.execute(
            """
            SELECT COUNT(*) FROM transactions
            WHERE tx_type='출고지시취소' AND memo LIKE ?
            """,
            (f"%출고지시서 #{order_id}%",),
        ).fetchone()[0]
        if int(already_cancelled or 0) > 0:
            cur.execute(
                "UPDATE outbound_orders SET status='취소됨', memo=IFNULL(memo,'') || ? WHERE id=?",
                ("\n" + now + " 출고지시 취소 상태 보정", order_id),
            )
            con.commit()
            raise ValueError("이미 이 출고지시서의 재고 원복 이력이 있습니다.")

        item_count, restored_count = _fallback_restore_outbound_from_items(cur, order_id, now)
        cur.execute(
            "UPDATE outbound_orders SET status='취소됨', memo=IFNULL(memo,'') || ? WHERE id=?",
            ("\n" + now + " 출고지시 취소", order_id),
        )
        con.commit()
        return item_count, restored_count


def restore_inventory_from_log(order_id):
    return cancel_outbound_order(int(order_id))


def cancel_saved_order(order_id):
    return restore_inventory_from_log(int(order_id))
