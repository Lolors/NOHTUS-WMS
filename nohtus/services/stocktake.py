"""Stocktake service helpers."""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

import pandas as pd

from nohtus.config import COMPANIES
from nohtus.db import connect, q
from nohtus.dates import display_date_only, normalize_exp_date
from nohtus.services.inventory import insert_transaction_log


def _normalize_stocktake_location(value):
    """로케이션 비교용 문자열을 정규화한다."""
    return str(value or "").strip().upper().replace(" ", "").replace("-", "").replace("_", "")


def _normalized_stocktake_location_sql():
    return "REPLACE(REPLACE(REPLACE(UPPER(TRIM(COALESCE(location,''))), ' ', ''), '-', ''), '_', '')"


def _is_zero_qty_exception_location(location):
    """G1, G2는 수량이 0이어도 재고 행을 유지한다."""
    normalized = _normalize_stocktake_location(location)
    return normalized.startswith("G1") or normalized.startswith("G2")


def _zero_qty_exception_sql():
    """0이어도 실사/기준재고 양식에 남겨야 하는 로케이션 조건."""
    location = _normalized_stocktake_location_sql()
    return f"{location} LIKE 'G1%' OR {location} LIKE 'G2%'"


def current_baseline_stock_excel_bytes(exclude_zero=False):
    """현재 WMS 재고를 기준재고 업로드 양식에 채워서 내려받는다.

    0 수량 제외 옵션을 사용해도 G1, G2 재고는 포함한다.
    """
    where_sql = f"WHERE qty>0 OR {_zero_qty_exception_sql()}" if exclude_zero else ""
    inv = q(
        f"""
        SELECT company, product_name, warehouse_name, lot, exp_date, location, qty
        FROM inventory
        {where_sql}
        ORDER BY company, product_name, lot, exp_date, location
        """
    )
    cols = ["사업장", "ERP제품코드", "ERP제품명", "표준제품명", "LOT/제조번호", "유통기한", "로케이션", "수량"]
    if inv.empty:
        return _baseline_stock_excel_bytes_from_dataframe(pd.DataFrame(columns=cols))

    product_df = q(
        """
        SELECT standard_name, product_code, erp_noh_code,
               erp_nohtuspharm_name, erp_noh_name, erp_nohtus_name, bidata_name
        FROM products
        """
    )
    product_map = {}
    if not product_df.empty:
        for row in product_df.itertuples(index=False):
            product_map[str(getattr(row, "standard_name") or "").strip()] = {
                "product_code": str(getattr(row, "product_code") or "").strip(),
                "erp_noh_code": str(getattr(row, "erp_noh_code") or "").strip(),
                "erp_nohtuspharm_name": str(getattr(row, "erp_nohtuspharm_name") or "").strip(),
                "erp_noh_name": str(getattr(row, "erp_noh_name") or "").strip(),
                "erp_nohtus_name": str(getattr(row, "erp_nohtus_name") or "").strip(),
                "bidata_name": str(getattr(row, "bidata_name") or "").strip(),
            }

    rows = []
    for row in inv.itertuples(index=False):
        company = str(getattr(row, "company") or "").strip()
        standard = str(getattr(row, "product_name") or "").strip()
        warehouse = str(getattr(row, "warehouse_name") or "").strip()
        info = product_map.get(standard, {})
        code = ""
        erp_name = warehouse or standard

        if company == "노투스팜":
            code = info.get("product_code", "")
            erp_name = info.get("erp_nohtuspharm_name", "") or warehouse or standard
        elif company == "NOH":
            code = info.get("erp_noh_code", "")
            erp_name = info.get("erp_noh_name", "") or warehouse or standard
        elif company == "노투스":
            erp_name = info.get("erp_nohtus_name", "") or warehouse or standard
        elif company == "비자료":
            erp_name = info.get("bidata_name", "") or warehouse or standard

        rows.append(
            {
                "사업장": company,
                "ERP제품코드": code,
                "ERP제품명": erp_name,
                "표준제품명": standard,
                "LOT/제조번호": str(getattr(row, "lot") or "-").strip() or "-",
                "유통기한": display_date_only(getattr(row, "exp_date") or "-"),
                "로케이션": str(getattr(row, "location") or "").strip(),
                "수량": int(getattr(row, "qty") or 0),
            }
        )

    return _baseline_stock_excel_bytes_from_dataframe(pd.DataFrame(rows, columns=cols))


def location_series(location):
    """로케이션의 계열(알파벳 구역군)을 반환한다: G/X/T/N, 해당 없으면 빈 문자열."""
    normalized = _normalize_stocktake_location(location)
    if normalized.startswith("G"):
        return "G"
    if normalized.startswith("X"):
        return "X"
    if normalized.startswith("T"):
        return "T"
    from nohtus.config import SPECIAL_LOCATIONS
    if str(location or "").strip() in SPECIAL_LOCATIONS:
        return "N"
    return ""


_TRUE_MATERIAL_VALUES = ("1", "true", "yes", "y", "o", "v", "체크", "부자재")


def _inventory_survey_excel_bytes(df, sheet_name):
    out = pd.DataFrame()
    out["로케이션"] = df["location"] if not df.empty else []
    out["제품명(표준제품명)"] = df["product_name"] if not df.empty else []
    out["제조번호"] = df["lot"] if not df.empty else []
    out["유통기한"] = df["exp_date"].apply(display_date_only) if not df.empty else []
    out["전산수량"] = df["qty"] if not df.empty else []
    out["실물수량"] = ""

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]
        widths = {"A": 16, "B": 30, "C": 18, "D": 16, "E": 12, "F": 12}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill

    bio.seek(0)
    return bio.getvalue()


def full_inventory_excel_bytes(exclude_zero=True, exclude_series=None):
    exclude_series = set(exclude_series or ())
    where_sql = f"WHERE qty>0 OR {_zero_qty_exception_sql()}" if exclude_zero else ""
    df = q(
        f"""
        SELECT location, product_name, warehouse_name, lot, exp_date, qty
        FROM inventory
        {where_sql}
        ORDER BY location, product_name, lot, exp_date
        """
    )
    if not df.empty and exclude_series:
        df = df[~df["location"].apply(location_series).isin(exclude_series)]

    return _inventory_survey_excel_bytes(df, "전체재고실사")


def material_inventory_excel_bytes(exclude_zero=True):
    """부자재로 분류된 제품만 모은 재고실사용 엑셀."""
    placeholders = ",".join("?" for _ in _TRUE_MATERIAL_VALUES)
    zero_sql = "AND qty>0" if exclude_zero else ""
    df = q(
        f"""
        SELECT location, product_name, warehouse_name, lot, exp_date, qty
        FROM inventory
        WHERE TRIM(COALESCE(product_name,'')) IN (
            SELECT TRIM(standard_name) FROM products
            WHERE LOWER(TRIM(CAST(COALESCE(is_material, 0) AS TEXT))) IN ({placeholders})
        )
        {zero_sql}
        ORDER BY location, product_name, lot, exp_date
        """,
        _TRUE_MATERIAL_VALUES,
    )
    return _inventory_survey_excel_bytes(df, "부자재재고실사")


def import_stock_survey_excel(uploaded_file, replace_current=True):
    """기준재고 엑셀을 현재 WMS 재고로 불러온다.

    일반 로케이션의 0 수량 행은 제외하지만 G1, G2의 0 수량 행은
    DB에 저장하여 이후 실사 및 기준재고 양식에도 계속 나타나게 한다.
    """
    normal_df, issue_df = prepare_baseline_stock_dataframe(uploaded_file)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    inserted = 0
    skipped = int(len(issue_df)) if issue_df is not None else 0
    product_inserted = 0

    with connect() as con:
        cur = con.cursor()
        if replace_current:
            cur.execute("DELETE FROM inventory")
            cur.execute("DELETE FROM transactions WHERE tx_type='재고조사불러오기'")

        for _, row in normal_df.iterrows():
            company = str(row.get("사업장") or "").strip()
            code = str(row.get("ERP제품코드") or "").strip()
            product_raw = str(row.get("ERP제품명") or "").strip()
            product = str(row.get("표준제품명") or "").strip()
            lot = str(row.get("LOT/제조번호") or "").strip() or "-"
            exp = _excel_date_to_iso(row.get("유통기한"))
            location = str(row.get("로케이션") or "").strip()
            qty = int(float(row.get("수량") or 0))

            keep_zero = qty == 0 and _is_zero_qty_exception_location(location)
            if not company or not product or not location or qty < 0 or (qty == 0 and not keep_zero):
                skipped += 1
                continue

            exists = cur.execute("SELECT id FROM products WHERE standard_name=?", (product,)).fetchone()
            if not exists:
                cur.execute(
                    """INSERT INTO products(
                           product_code, standard_name, warehouse_name, aliases,
                           erp_nohtuspharm_name, erp_noh_name, erp_noh_code,
                           erp_nohtus_name, bidata_name
                       ) VALUES(?,?,?,?,?,?,?,?,?)""",
                    (
                        code if company == "노투스팜" else "",
                        product,
                        product_raw,
                        "",
                        product_raw if company == "노투스팜" else "",
                        product_raw if company == "NOH" else "",
                        code if company == "NOH" else "",
                        product_raw if company == "노투스" else "",
                        product_raw if company == "비자료" else "",
                    ),
                )
                product_inserted += 1
            else:
                product_id = int(exists[0])
                if company == "노투스팜":
                    cur.execute(
                        "UPDATE products SET erp_nohtuspharm_name=COALESCE(NULLIF(erp_nohtuspharm_name,''), ?), product_code=COALESCE(NULLIF(product_code,''), ?) WHERE id=?",
                        (product_raw, code, product_id),
                    )
                elif company == "NOH":
                    cur.execute(
                        "UPDATE products SET erp_noh_name=COALESCE(NULLIF(erp_noh_name,''), ?), erp_noh_code=COALESCE(NULLIF(erp_noh_code,''), ?) WHERE id=?",
                        (product_raw, code, product_id),
                    )
                elif company == "노투스":
                    cur.execute(
                        "UPDATE products SET erp_nohtus_name=COALESCE(NULLIF(erp_nohtus_name,''), ?) WHERE id=?",
                        (product_raw, product_id),
                    )
                elif company == "비자료":
                    cur.execute(
                        "UPDATE products SET bidata_name=COALESCE(NULLIF(bidata_name,''), ?) WHERE id=?",
                        (product_raw, product_id),
                    )

            cur.execute(
                "INSERT INTO inventory(company, product_name, warehouse_name, lot, exp_date, location, qty, updated_at) VALUES(?,?,?,?,?,?,?,?)",
                (company, product, product_raw, lot, exp, location, qty, now),
            )
            insert_transaction_log(
                cur,
                created_at=now,
                tx_type="재고조사불러오기",
                product_name=product,
                warehouse_name=product_raw,
                lot=lot,
                exp_date=exp,
                from_company=None,
                from_location=None,
                to_company=company,
                to_location=location,
                qty=qty,
                memo=f"기준재고 엑셀 업로드 / 원본명: {product_raw}",
            )
            inserted += 1

        con.commit()

    return inserted, skipped, product_inserted, skipped


def _baseline_stock_excel_bytes_from_dataframe(df):
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="기준재고업로드")
        ws = writer.book["기준재고업로드"]

        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        optional_fill = PatternFill("solid", fgColor="EEF2FF")
        header_widths = {
            "사업장": 14,
            "로케이션": 18,
            "ERP제품코드": 18,
            "ERP제품명": 34,
            "표준제품명": 30,
            "LOT/제조번호": 18,
            "유통기한": 16,
            "전산수량": 12,
            "실제수량": 12,
            "차이": 10,
            "수량": 10,
        }
        headers = {
            cell.value: cell.column
            for cell in ws[1]
            if cell.value is not None
        }
        for header, column_index in headers.items():
            ws.column_dimensions[get_column_letter(column_index)].width = header_widths.get(header, 18)

        ws.freeze_panes = "A2"
        max_row = max(1, len(df) + 1)
        max_column = max(1, len(df.columns))
        ws.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max_row}"
        product_code_column = headers.get("ERP제품코드")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if product_code_column and cell.column == product_code_column and cell.row > 1:
                    cell.number_format = "@"
                    if cell.value is not None:
                        cell.value = str(cell.value)
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = optional_fill if cell.value == "표준제품명" else header_fill

    bio.seek(0)
    return bio.getvalue()


def prepare_baseline_stock_dataframe(uploaded_file):
    df = pd.read_excel(uploaded_file, dtype=str).fillna("")
    col_alias = {
        "구분": "사업장",
        "ERP제품코드": "ERP제품코드",
        "ERP 제품코드": "ERP제품코드",
        "전산제품코드": "ERP제품코드",
        "제품코드": "ERP제품코드",
        "노투스팜 ERP 제품코드": "ERP제품코드",
        "NOH ERP 제품코드": "ERP제품코드",
        "ERP상제품명": "ERP제품명",
        "ERP제품명": "ERP제품명",
        "제품명": "ERP제품명",
        "전산상명칭": "ERP제품명",
        "전산상 명칭": "ERP제품명",
        "전산상제품명": "ERP제품명",
        "비자료명": "비자료명",
        "LOT": "LOT/제조번호",
        "제조번호": "LOT/제조번호",
        "수량": "수량",
        "기준수량": "수량",
        "현재재고": "수량",
        "실재고": "수량",
    }
    df = df.rename(columns={column: col_alias.get(column, column) for column in df.columns})

    required_columns = [
        "사업장", "ERP제품코드", "ERP제품명", "비자료명", "표준제품명",
        "LOT/제조번호", "유통기한", "로케이션", "수량",
    ]
    for column in required_columns:
        if column not in df.columns:
            df[column] = ""

    rows = []
    for _, row in df.iterrows():
        rows.append(
            {
                "사업장": first_nonblank(row.get("사업장")),
                "ERP제품코드": first_nonblank(row.get("ERP제품코드"), row.get("노투스팜 ERP 제품코드"), row.get("NOH ERP 제품코드")),
                "ERP제품명": _baseline_get_product_raw(row),
                "표준제품명": first_nonblank(row.get("표준제품명")),
                "LOT/제조번호": first_nonblank(row.get("LOT/제조번호")) or "-",
                "유통기한": first_nonblank(row.get("유통기한")) or "-",
                "로케이션": first_nonblank(row.get("로케이션")),
                "수량": first_nonblank(row.get("수량")),
            }
        )

    return pd.DataFrame(rows), pd.DataFrame()


def first_nonblank(*values):
    for value in values:
        if value is None or pd.isna(value):
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _baseline_get_product_raw(row):
    return first_nonblank(row.get("ERP제품명"), row.get("비자료명"))


def _excel_date_to_iso(value):
    if value is None or pd.isna(value):
        return "-"
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return value.strftime("%Y-%m-%d")
    return normalize_exp_date(value)
