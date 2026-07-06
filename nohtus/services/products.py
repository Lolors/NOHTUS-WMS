"""Product and product-matching service functions for NOHTUS WMS.

This module is migrated gradually from app.py. Keep functions independent from
Streamlit whenever possible.
"""

from __future__ import annotations

from io import BytesIO

import pandas as pd

from nohtus.db import connect, q


def product_master_excel_bytes(highlight_missing=False):
    """제품 마스터를 사용자가 수정하기 쉬운 엑셀 양식으로 내보낸다.
    v3.7부터 제품코드는 노투스팜 ERP 전용 코드로 취급하고, 전산상 명칭은 제품마스터에서 제외한다.
    """
    df = q("SELECT standard_name, erp_nohtuspharm_name, product_code, erp_noh_name, erp_noh_code, erp_nohtus_name, bidata_name, aliases FROM products ORDER BY standard_name, id")
    out = df.rename(columns={
        "standard_name": "표준제품명",
        "erp_nohtuspharm_name": "노투스팜 ERP명",
        "product_code": "노투스팜 ERP 제품코드",
        "erp_noh_name": "NOH ERP명",
        "erp_noh_code": "NOH ERP 제품코드",
        "erp_nohtus_name": "노투스 ERP명",
        "bidata_name": "비자료명",
        "aliases": "별칭",
    })
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="제품마스터")
        ws = writer.book["제품마스터"]
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        need_fill = PatternFill("solid", fgColor="FFF2CC")
        widths = {"A":24,"B":34,"C":28,"D":34,"E":28,"F":34,"G":34,"H":34}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:H{max(1, len(out)+1)}"
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                # ERP 제품코드는 계산값이 아니라 텍스트다. 003 같은 앞자리 0을 보존한다.
                if cell.column_letter in ["C", "E"]:
                    cell.number_format = "@"
                    if cell.value is not None:
                        cell.value = str(cell.value)
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                elif highlight_missing:
                    row_values = [ws.cell(row=cell.row, column=i).value for i in [2,4,6,7]]
                    if all(str(v or "").strip() == "" for v in row_values):
                        cell.fill = need_fill
    bio.seek(0)
    return bio.getvalue()


def _clean_text(value):
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def import_product_master_excel(uploaded_file):
    """업로드된 제품매칭표 엑셀을 products 테이블에 반영한다.

    중요 원칙:
    - 같은 표준제품명에 ERP명이 여러 개 존재할 수 있다.
    - 표준제품명만 기준으로 중복 제거하지 않는다.
    - 완전히 같은 행만 중복으로 보며, 서로 다른 ERP명/코드/비자료명은 별도 매칭으로 보존한다.
    """
    df = pd.read_excel(uploaded_file, dtype=str).fillna("")
    rename = {
        "노투스팜 ERP 제품코드": "product_code",
        "제품코드": "product_code",
        "표준제품명": "standard_name",
        "제품명": "standard_name",
        "별칭": "aliases",
        "노투스팜 ERP명": "erp_nohtuspharm_name",
        "NOH ERP명": "erp_noh_name",
        "NOH ERP 제품코드": "erp_noh_code",
        "노투스 ERP명": "erp_nohtus_name",
        "비자료명": "bidata_name",
    }
    df = df.rename(columns={c: rename.get(c, c) for c in df.columns})
    if "standard_name" not in df.columns:
        raise ValueError("엑셀에 '표준제품명' 컬럼이 필요합니다.")
    for c in ["product_code", "aliases", "erp_nohtuspharm_name", "erp_nohtus_name", "erp_noh_name", "erp_noh_code", "bidata_name"]:
        if c not in df.columns:
            df[c] = ""
    inserted = 0
    skipped = 0
    seen = set()
    rows_to_insert = []
    for _, r in df.iterrows():
        code = _clean_text(r.get("product_code"))
        name = _clean_text(r.get("standard_name"))
        aliases = _clean_text(r.get("aliases"))
        erp_np = _clean_text(r.get("erp_nohtuspharm_name"))
        erp_nt = _clean_text(r.get("erp_nohtus_name"))
        erp_noh = _clean_text(r.get("erp_noh_name"))
        erp_noh_code = _clean_text(r.get("erp_noh_code"))
        bidata_name = _clean_text(r.get("bidata_name"))
        if not name:
            skipped += 1
            continue
        key = (
            name,
            erp_np,
            code,
            erp_noh,
            erp_noh_code,
            erp_nt,
            bidata_name,
            aliases,
        )
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        rows_to_insert.append((code, name, name, aliases, erp_np, erp_nt, erp_noh, erp_noh_code, bidata_name))
    with connect() as con:
        cur = con.cursor()
        cur.execute("DELETE FROM products")
        for row in rows_to_insert:
            cur.execute("""INSERT INTO products(product_code,standard_name,warehouse_name,aliases,erp_nohtuspharm_name,erp_nohtus_name,erp_noh_name,erp_noh_code,bidata_name)
                           VALUES(?,?,?,?,?,?,?,?,?)""", row)
            inserted += 1
        con.commit()
    return 0, inserted, skipped


def product_options(term=""):
    term = (term or "").strip().lower()
    df = q("""SELECT standard_name, warehouse_name, aliases,
                    erp_nohtuspharm_name, erp_nohtus_name, erp_noh_name, bidata_name
             FROM products ORDER BY standard_name, id""")
    if term:
        search_cols = ["standard_name", "warehouse_name", "aliases", "erp_nohtuspharm_name", "erp_nohtus_name", "erp_noh_name", "bidata_name"]
        mask = df.apply(lambda r: any(term in str(r.get(c, "")).lower() for c in search_cols), axis=1)
        df = df[mask]
    return df
