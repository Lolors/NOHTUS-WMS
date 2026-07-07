"""Service helpers."""
from __future__ import annotations
from io import BytesIO
from datetime import date
import pandas as pd
from nohtus.db import q
import streamlit as st
import re

def _infer_customer_from_title(title, customers_df=None):
    """출고지시서 제목에서 거래처명을 추정한다.
    제목 규칙: [출고처] [첫 제품명] 외 x품목.
    거래처 관리에 등록된 이름 중 title 시작과 일치하는 가장 긴 이름을 우선 사용한다.
    """
    title = str(title or '').strip()
    if not title:
        return ('', '')
    if customers_df is None:
        customers_df = q('SELECT customer_name, manager FROM customers ORDER BY LENGTH(customer_name) DESC')
    if not customers_df.empty:
        for r in customers_df.itertuples():
            name = str(getattr(r, 'customer_name', '') or '').strip()
            if name and title.startswith(name):
                return (name, str(getattr(r, 'manager', '') or ''))
    return (title.split()[0] if title.split() else title, '')

def _extract_inbound_source_from_memo(memo):
    """입고 이력 memo에서 입고처만 추출한다.
    저장 형식 예: '매입처: 거래처명 / 기타메모'
    """
    text = str(memo or '').strip()
    if not text or text == '입고 등록':
        return ''
    prefixes = ['매입처:', '입고처:']
    for prefix in prefixes:
        if text.startswith(prefix):
            text = text[len(prefix):].strip()
            if ' / ' in text:
                text = text.split(' / ', 1)[0].strip()
            break
    return text

def dataframe_to_excel_bytes(df, sheet_name='Sheet1'):
    """DataFrame을 엑셀 bytes로 변환한다.
    openpyxl이 허용하지 않는 제어문자/특수 공백은 저장 전에 전부 제거한다.
    """
    bio = BytesIO()
    safe_df = df.copy() if df is not None else pd.DataFrame()
    safe_sheet = clean_excel_text(sheet_name)[:31] or 'Sheet1'
    safe_df.columns = [clean_excel_text(c) for c in safe_df.columns]
    for col in safe_df.columns:
        if safe_df[col].dtype == object:
            safe_df[col] = safe_df[col].apply(lambda v: clean_excel_text(v) if v is not None else '')
    with pd.ExcelWriter(bio, engine='openpyxl') as writer:
        safe_df.to_excel(writer, index=False, sheet_name=safe_sheet)
        ws = writer.book[safe_sheet]
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill('solid', fgColor='E5E7EB')
        for col in ws.columns:
            max_len = 10
            letter = col[0].column_letter
            for cell in col:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                max_len = max(max_len, len(str(cell.value or '')) + 2)
            ws.column_dimensions[letter].width = min(max_len, 42)
        if safe_sheet == '마감체크':
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 50
            from openpyxl.styles import PatternFill
            current_fill = PatternFill('solid', fgColor='DDEBF7')
            header_map = {str(ws.cell(row=1, column=i).value or '').strip(): i for i in range(1, ws.max_column + 1)}
            cur_col = header_map.get('현재수량')
            if cur_col:
                for rr in range(1, ws.max_row + 1):
                    ws.cell(row=rr, column=cur_col).fill = current_fill
        ws.freeze_panes = 'A2'
        if safe_df.shape[1] > 0:
            ws.auto_filter.ref = ws.dimensions
    bio.seek(0)
    return bio.getvalue()

def page_erp_stock_compare():
    st.title('ERP 재고 비교')
    st.caption('WMS 재고와 ERP 현재고를 사업장별 ERP 제품명 기준으로 비교합니다. 제조번호와 유통기한은 무시하고 총수량만 합산하며, 비자료는 제외합니다.')

    parsed = []
    cols = st.columns(3, gap='large')
    company_specs = [
        ('노투스팜', 'SIMS', '제품명', '현재고수량'),
        ('NOH', 'SIMS', '제품명', '현재고수량'),
        ('노투스', 'IBK우리은행 전산', '품목명/규격', '현재재고'),
    ]
    for col, (company, system_name, name_col, qty_col) in zip(cols, company_specs):
        with col:
            st.markdown(f'### {company}')
            st.caption(f'{system_name} · 제품명: {name_col} · 수량: {qty_col}')
            up = st.file_uploader(f'{company} 현재고 엑셀', type=['xlsx', 'xls'], key=f'erp_compare_upload_{company}')
            if up is not None:
                parsed_file = _read_erp_current_file(up, company)
                parsed.append(parsed_file)
                if parsed_file is not None:
                    st.success('파일 선택됨')
            else:
                parsed.append(None)

    if not any(parsed):
        st.info('노투스팜 / NOH / 노투스 ERP 현재고 엑셀을 업로드하세요.')
        return

    _, run_col, _ = st.columns([3, 2, 3], gap='large')
    with run_col:
        run_compare = st.button('ERP 재고 비교 실행', type='primary', use_container_width=True)

    if run_compare:
        erp_sum, erp_row_count = _normalize_erp_current_rows(parsed)
        wms_sum, unmapped = _wms_stock_for_erp_compare()
        comp = _build_erp_wms_compare_result(erp_sum, wms_sum)
        diff_count = int((comp['차이'] != 0).sum()) if not comp.empty else 0

        st.session_state['erp_compare_rows'] = comp.to_dict('records')
        st.session_state['erp_compare_unmapped'] = unmapped.to_dict('records') if not unmapped.empty else []
        st.session_state['erp_compare_summary'] = {
            'erp_rows': erp_row_count,
            'erp_products': len(erp_sum),
            'wms_products': len(wms_sum),
            'diff': diff_count,
            'unmapped': len(unmapped),
        }

    if 'erp_compare_summary' in st.session_state:
        sm = st.session_state['erp_compare_summary']
        m1, m2, m3, m4 = st.columns(4)
        m1.metric('ERP 원본 행', f"{sm['erp_rows']}건")
        m2.metric('ERP 제품 합산', f"{sm['erp_products']}건")
        m3.metric('WMS 제품 합산', f"{sm['wms_products']}건")
        m4.metric('차이 항목', f"{sm['diff']}건")

        comp = pd.DataFrame(st.session_state.get('erp_compare_rows', []))
        unmapped = pd.DataFrame(st.session_state.get('erp_compare_unmapped', []))
        if not comp.empty:
            st.markdown('### 재고 비교 결과')
            only_diff = st.checkbox('차이 있는 항목만 보기', value=True, key='erp_compare_only_diff')
            shown = comp[comp['차이'] != 0] if only_diff else comp
            st.dataframe(shown, use_container_width=True, hide_index=True)
            st.download_button(
                '비교 결과 엑셀 다운로드',
                data=dataframe_to_excel_bytes(comp, 'ERP_WMS_비교'),
                file_name=f"NOHTUS_ERP_WMS_비교_{date.today().strftime('%Y%m%d')}.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
            )
        else:
            st.info('비교할 재고가 없습니다.')

        if not unmapped.empty:
            st.warning('제품매칭표에 해당 사업장 ERP명이 없어 표준제품명으로 비교한 WMS 항목입니다. ERP명과 표준제품명이 다르면 차이로 표시될 수 있습니다.')
            st.dataframe(unmapped, use_container_width=True, hide_index=True)

def _normalize_erp_current_rows(parsed_files):
    """ERP 파일을 사업장 + 제품명 기준으로 합산한다."""
    rows = []
    source_row_count = 0
    for info in parsed_files:
        if not info:
            continue
        company = info['company']
        raw = info['raw'].copy()
        name_col = info['name_col']
        qty_col = info['qty_col']
        raw[name_col] = raw[name_col].apply(clean_excel_text)
        raw[qty_col] = pd.to_numeric(raw[qty_col], errors='coerce').fillna(0).astype(int)
        raw = raw[~raw[name_col].apply(is_ignored_erp_product_name)]
        raw = raw[raw[qty_col] != 0]
        source_row_count += len(raw)
        for _, rr in raw.iterrows():
            rows.append({
                '사업장': company,
                'ERP제품명': clean_excel_text(rr[name_col]),
                'ERP수량': int(rr[qty_col]),
            })
    if not rows:
        return (pd.DataFrame(columns=['사업장', 'ERP제품명', 'ERP수량']), 0)
    erp = pd.DataFrame(rows)
    erp = erp.groupby(['사업장', 'ERP제품명'], as_index=False)['ERP수량'].sum()
    return (erp, source_row_count)

def _wms_stock_for_erp_compare():
    """WMS 재고를 사업장별 ERP 제품명으로 변환한 뒤 합산한다."""
    wms_raw = q("""
        SELECT company AS 사업장, product_name AS 표준제품명, SUM(qty) AS 수량
        FROM inventory
        WHERE qty<>0 AND company IN ('노투스팜','NOH','노투스')
        GROUP BY company, product_name
    """)
    if wms_raw.empty:
        empty = pd.DataFrame(columns=['사업장', 'ERP제품명', 'WMS수량'])
        return (empty, pd.DataFrame(columns=['사업장', '표준제품명', '비교제품명', 'WMS수량']))

    rows = []
    unmapped = []
    for rr in wms_raw.itertuples(index=False):
        company = str(getattr(rr, '사업장') or '').strip()
        standard_name = str(getattr(rr, '표준제품명') or '').strip()
        qty = int(getattr(rr, '수량') or 0)
        erp_name = product_mapping_name_for(company, standard_name)
        used_fallback = not bool(erp_name)
        compare_name = clean_excel_text(erp_name or standard_name)
        if is_ignored_erp_product_name(compare_name):
            continue
        rows.append({'사업장': company, 'ERP제품명': compare_name, 'WMS수량': qty})
        if used_fallback:
            unmapped.append({'사업장': company, '표준제품명': standard_name, '비교제품명': compare_name, 'WMS수량': qty})

    if rows:
        wms = pd.DataFrame(rows).groupby(['사업장', 'ERP제품명'], as_index=False)['WMS수량'].sum()
    else:
        wms = pd.DataFrame(columns=['사업장', 'ERP제품명', 'WMS수량'])
    unmapped_df = pd.DataFrame(unmapped, columns=['사업장', '표준제품명', '비교제품명', 'WMS수량'])
    return (wms, unmapped_df)

def _build_erp_wms_compare_result(erp_sum, wms_sum):
    comp = wms_sum.merge(erp_sum, how='outer', on=['사업장', 'ERP제품명'])
    if comp.empty:
        return pd.DataFrame(columns=['사업장', 'ERP제품명', 'ERP수량', 'WMS수량', '차이'])
    comp['ERP수량'] = comp['ERP수량'].fillna(0).astype(int)
    comp['WMS수량'] = comp['WMS수량'].fillna(0).astype(int)
    comp['차이'] = comp['WMS수량'] - comp['ERP수량']
    return comp[['사업장', 'ERP제품명', 'ERP수량', 'WMS수량', '차이']].sort_values(['사업장', 'ERP제품명'])

def _read_erp_current_file(uploaded, company):
    """ERP 현재고 파일을 고정 규칙으로 읽는다.
    - 노투스팜/NOH: 제품명, 현재고수량
    - 노투스: 8행 헤더, 품목명/규격, 현재재고
    """
    if uploaded is None:
        return None
    try:
        if company == '노투스':
            raw = pd.read_excel(uploaded, header=7)
            name_col = _find_required_column(raw.columns, ['품목명/규격'])
            qty_col = _find_required_column(raw.columns, ['현재재고'])
        else:
            raw = pd.read_excel(uploaded)
            name_col = _find_required_column(raw.columns, ['제품명'])
            qty_col = _find_required_column(raw.columns, ['현재고수량'])
    except Exception as e:
        st.error(f'{company} 엑셀 읽기 실패: {e}')
        return None
    if raw.empty:
        st.warning(f'{company} 엑셀에 데이터가 없습니다.')
        return None
    raw.columns = [str(c).strip() for c in raw.columns]
    if not name_col or not qty_col or name_col not in raw.columns or qty_col not in raw.columns:
        st.error(f'{company} ERP 파일에서 필요한 컬럼을 찾을 수 없습니다.')
        st.caption('노투스팜/NOH: 제품명, 현재고수량 · 노투스: 8행 헤더의 품목명/규격, 현재재고')
        return None
    return {'company': company, 'raw': raw, 'name_col': name_col, 'qty_col': qty_col}

def clean_excel_text(value):
    """openpyxl이 저장하지 못하는 제어문자와 특수 공백을 제거한다."""
    if value is None:
        return ''
    text = str(value)
    try:
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        text = ILLEGAL_CHARACTERS_RE.sub('', text)
    except Exception:
        text = re.sub('[\\x00-\\x08\\x0B-\\x0C\\x0E-\\x1F]', '', text)
    text = re.sub('[\\x00-\\x1F\\x7F-\\x9F\\u200b\\u200c\\u200d\\ufeff]', '', text)
    text = text.replace('\xa0', ' ')
    return text.strip()

def is_ignored_erp_product_name(value):
    key = _erp_name_key(value)
    return not key or key in {'합계', '배송비'} or '합계' in key or ('배송비' in key)

def product_compare_name_for(company, standard_name):
    """ERP 비교용 제품명.
    제품매칭표에 해당 사업장의 ERP명이 있으면 ERP명 기준으로 비교하고,
    없으면 표준제품명 기준으로 비교한다.
    """
    return product_mapping_name_for(company, standard_name) or (standard_name or '')

def _erp_name_key(value):
    text = clean_excel_text(value)
    return re.sub('\\s+', '', text).replace('[', '').replace(']', '')

def _find_required_column(columns, candidates):
    cols = [str(c).strip() for c in columns]
    for cand in candidates:
        for col in cols:
            if col == cand:
                return col
    return None

def product_mapping_name_for(company, standard_name):
    if not standard_name:
        return ''
    col = {'노투스팜': 'erp_nohtuspharm_name', 'NOH': 'erp_noh_name', '노투스': 'erp_nohtus_name', '비자료': 'bidata_name'}.get(company)
    if not col:
        return ''
    df = q(f'SELECT {col} AS nm FROM products WHERE standard_name=? ORDER BY id', (standard_name,))
    if df.empty:
        return ''
    return first_nonblank(*df['nm'].tolist())

def first_nonblank(*values):
    for v in values:
        if v is None:
            continue
        text = str(v).strip()
        if text and text.lower() != 'nan' and (text != '-'):
            return text
    return ''