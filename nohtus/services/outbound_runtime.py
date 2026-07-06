from pathlib import Path
from datetime import datetime
from io import BytesIO

import pandas as pd

from nohtus.db import connect, q
from nohtus.dates import display_date_only
from nohtus.locations import location_picking_key


def outbound_erp_note_for_row(row, cache=None):
    """출고지시서 비고용 ERP명/비자료명.
    매출등록은 ERP명 기준으로 하므로, 사업장+표준제품명에 맞는 제품매칭표 값을 표시한다.
    매칭값이 없으면 빈칸으로 둔다.
    """
    cache = cache if cache is not None else {}
    company = str(row.get("사업장") or row.get("company") or row.get("사업체") or "").strip()
    product = str(row.get("제품명") or row.get("product_name") or "").strip()
    if not company or not product:
        return ""
    key = (company, product)
    if key in cache:
        return cache[key]
    try:
        val = product_mapping_name_for(company, product) or ""
    except Exception:
        val = ""
    cache[key] = val
    return val

def outbound_excel_bytes(rows, title="출고지시서"):
    rows = sort_outbound_rows_for_picking(rows)
    note_cache = {}
    rows = [{**(r or {}), "비고": outbound_erp_note_for_row(r or {}, note_cache)} for r in rows]
    df = pd.DataFrame(rows)
    cols = ["로케이션", "제품명", "LOT", "유통기한", "요청수량", "비고"]
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    out = df[cols].copy()
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="출고지시서", startrow=2)
        ws = writer.book["출고지시서"]
        ws["A1"] = title or "출고지시서"
        ws.merge_cells("A1:F1")
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")
        widths = {"A":18,"B":30,"C":18,"D":18,"E":12,"F":32}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows(min_row=3, max_row=3 + len(out), min_col=1, max_col=6):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", horizontal="center" if cell.column in [1,3,4,5] else "left", wrap_text=True)
                if cell.row == 3:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
        ws.auto_filter.ref = f"A3:F{3+len(out)}"
    bio.seek(0)
    return bio.getvalue()

def _find_korean_font():
    candidates = [
        r"C:\Windows\Fonts\malgun.ttf",
        r"C:\Windows\Fonts\malgunbd.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return str(path)
    return None

def outbound_pdf_bytes(rows, title="출고지시서"):
    rows = sort_outbound_rows_for_picking(rows)
    note_cache = {}
    rows = [{**(r or {}), "비고": outbound_erp_note_for_row(r or {}, note_cache)} for r in rows]
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    bio = BytesIO()
    font_name = "Helvetica"
    font_path = _find_korean_font()
    if font_path:
        try:
            pdfmetrics.registerFont(TTFont("NOHTUS_KR", font_path))
            font_name = "NOHTUS_KR"
        except Exception:
            font_name = "Helvetica"
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4), leftMargin=22, rightMargin=22, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = font_name
    styles["Normal"].fontName = font_name
    story = [Paragraph(title or "출고지시서", styles["Title"]), Spacer(1, 12)]
    cols = ["로케이션", "제품명", "LOT", "유통기한", "요청수량", "비고"]
    data = [cols]
    for r in rows:
        data.append([str(r.get(c, "")) for c in cols])
    table = Table(data, colWidths=[80, 220, 105, 105, 70, 160], repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), font_name),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E5E7EB")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#111827")),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
        ("ALIGN", (0,1), (0,-1), "CENTER"),
        ("ALIGN", (2,1), (4,-1), "CENTER"),
        ("ALIGN", (1,1), (1,-1), "LEFT"),
        ("ALIGN", (5,1), (5,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(table)
    doc.build(story)
    bio.seek(0)
    return bio.getvalue()

def first_nonblank(*values):
    for v in values:
        if v is None:
            continue
        text = str(v).strip()
        if text and text.lower() != "nan" and text != "-":
            return text
    return ""

def product_mapping_name_for(company, standard_name):
    if not standard_name:
        return ""
    col = {
        "노투스팜": "erp_nohtuspharm_name",
        "NOH": "erp_noh_name",
        "노투스": "erp_nohtus_name",
        "비자료": "bidata_name",
    }.get(company)
    if not col:
        return ""
    df = q(f"SELECT {col} AS nm FROM products WHERE standard_name=?", (standard_name,))
    if df.empty:
        return ""
    return first_nonblank(df.iloc[0].get("nm"))

def product_compare_name_for(company, standard_name):
    """ERP 비교용 제품명.
    제품매칭표에 해당 사업장의 ERP명이 있으면 ERP명 기준으로 비교하고,
    없으면 표준제품명 기준으로 비교한다.
    """
    return product_mapping_name_for(company, standard_name) or (standard_name or "")

def sort_outbound_rows_for_picking(rows):
    """출고지시서 출력용 피킹 순서 정렬. 화면 장바구니 순서는 건드리지 않는다."""
    if not rows:
        return rows
    copied = [dict(r) for r in rows]
    return sorted(copied, key=lambda r: location_picking_key(r.get("로케이션") or r.get("location") or ""))

def product_total_stock(cur, product_name):
    """현재 inventory 기준 표준제품명 전체 총재고.
    사업장/LOT/유통기한/로케이션을 모두 무시하고 product_name만 기준으로 합산한다.
    거래 이력의 final_stock은 이 함수의 값을 작업 직후 스냅샷으로 저장한다.
    """
    product_name = str(product_name or "").strip()
    if not product_name:
        return 0
    row = cur.execute("SELECT COALESCE(SUM(qty), 0) FROM inventory WHERE product_name=?", (product_name,)).fetchone()
    return int((row[0] if row else 0) or 0)

def insert_transaction_log(cur, *, created_at, tx_type, product_name, warehouse_name=None,
                           lot="-", exp_date="-", from_company=None, from_location=None,
                           to_company=None, to_location=None, qty=0, memo="", final_stock=None):
    """거래 이력을 한 곳에서 기록한다.
    final_stock을 넘기지 않으면 현재 inventory 기준 표준제품명 총재고를 저장한다.
    """
    if final_stock is None:
        final_stock = product_total_stock(cur, product_name)
    cur.execute("""INSERT INTO transactions(created_at,tx_type,product_name,warehouse_name,lot,exp_date,
                   from_company,from_location,to_company,to_location,qty,memo,final_stock)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (created_at, tx_type, product_name, warehouse_name, lot or "-", exp_date or "-",
                 from_company, from_location, to_company, to_location, int(qty or 0), memo, int(final_stock or 0)))

def create_outbound_instruction(src_id, qty, memo="출고지시"):
    """출고지시는 피킹 지시서만 남기고 실제 inventory 수량은 차감하지 않는다."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with connect() as con:
        cur = con.cursor()
        src = cur.execute("SELECT * FROM inventory WHERE id=?", (src_id,)).fetchone()
        cols = [d[0] for d in cur.description]
        src = dict(zip(cols, src)) if src else None
        if not src:
            raise ValueError("출고 지시할 재고를 찾을 수 없습니다.")
        if qty <= 0 or qty > src["qty"]:
            raise ValueError("지시 수량이 현재 재고보다 많거나 올바르지 않습니다.")
        cur.execute("""INSERT INTO transactions(created_at,tx_type,product_name,warehouse_name,lot,exp_date,from_company,from_location,to_company,to_location,qty,memo)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (now,"출고지시",src["product_name"],src.get("warehouse_name"),src["lot"],src["exp_date"],src["company"],src["location"],None,None,qty,memo))
        con.commit()

def load_outbound_order(order_id):
    df = q("""SELECT inventory_id AS id, location AS 로케이션, product_name AS 제품명, lot AS LOT, exp_date AS 유통기한, qty AS 요청수량, company AS 사업장, warehouse_name AS '전산상 명칭'
              FROM outbound_order_items WHERE order_id=? ORDER BY id""", (order_id,))
    return df.to_dict("records")

def build_outbound_order_title(customer_name, cart_items, fallback_title=""):
    """출고지시서 제목 기본값 생성.
    형식: [매출처] [1번째 제품명] 외 x품목
    """
    customer = str(customer_name or "").strip()
    items = list(cart_items or [])
    if not items:
        return str(fallback_title or "").strip()
    first = items[0] or {}
    first_name = str(first.get("제품명") or first.get("product_name") or "").strip()
    if customer and first_name:
        title = f"{customer} - {first_name}"
    elif customer:
        title = customer
    else:
        title = first_name
    rest_count = max(0, len(items) - 1)
    if rest_count:
        title = f"{title} 외 {rest_count}품목"
    return title or str(fallback_title or "").strip()

def _fallback_restore_outbound_from_items(cur, order_id, now):
    """출고지시 품목 테이블 기준으로 취소 수량을 현재 재고에 더한다.
    transactions.final_stock은 재고행 수량이 아니라 표준제품명 전체 총재고 스냅샷으로 저장한다.
    """
    item_rows = cur.execute("""SELECT inventory_id, location, product_name, lot, exp_date, qty, company, warehouse_name
                                   FROM outbound_order_items WHERE order_id=? ORDER BY id""", (order_id,)).fetchall()
    restored_count = 0
    for inv_id, location, product_name, lot, exp_date, qty, company, warehouse_name in item_rows:
        qty = int(qty or 0)
        if qty <= 0:
            continue
        inv = None
        if inv_id:
            inv = cur.execute("SELECT id, qty FROM inventory WHERE id=?", (int(inv_id),)).fetchone()
        if not inv:
            inv = cur.execute("""SELECT id, qty FROM inventory
                                     WHERE company=? AND product_name=? AND IFNULL(warehouse_name,'')=? AND IFNULL(lot,'')=? AND IFNULL(exp_date,'')=? AND location=?""",
                                  (company or "", product_name or "", warehouse_name or "", lot or "", exp_date or "", location or "")).fetchone()
        if inv:
            row_qty_after = int(inv[1] or 0) + qty
            cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (row_qty_after, now, int(inv[0])))
        else:
            row_qty_after = qty
            cur.execute("""INSERT INTO inventory(company, product_name, warehouse_name, lot, exp_date, location, qty, updated_at)
                               VALUES(?,?,?,?,?,?,?,?)""", (company or "", product_name or "", warehouse_name or "", lot or "-", exp_date or "-", location or "", qty, now))
        restored_count += 1
        insert_transaction_log(cur, created_at=now, tx_type="출고지시취소", product_name=product_name or "", warehouse_name=warehouse_name or "",
                               lot=lot or "-", exp_date=exp_date or "-", from_company=company or "", from_location=location or "",
                               to_company=company or "", to_location=location or "", qty=qty, memo=f"출고지시서 #{order_id} 취소 / 원복")
    return len(item_rows), restored_count

def cancel_outbound_order(order_id):
    """저장된 출고지시를 취소 처리하고 출고 품목 수량을 현재 재고에 되돌린다.

    주의: transactions.final_stock은 표준제품명 전체 총재고 스냅샷이므로
    재고행 원복 계산에 사용하지 않는다. 원복은 outbound_order_items의 품목/수량 기준으로 처리한다.
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    order_id = int(order_id)
    with connect() as con:
        cur = con.cursor()
        order = cur.execute("SELECT id, status FROM outbound_orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            raise ValueError("취소할 출고지시서를 찾을 수 없습니다.")
        if str(order[1] or "") == "취소됨":
            raise ValueError("이미 취소된 출고지시서입니다.")

        already_cancelled = cur.execute("""SELECT COUNT(*) FROM transactions
                                         WHERE tx_type='출고지시취소' AND memo LIKE ?""", (f"%출고지시서 #{order_id}%",)).fetchone()[0]
        if int(already_cancelled or 0) > 0:
            cur.execute("UPDATE outbound_orders SET status='취소됨', memo=IFNULL(memo,'') || ? WHERE id=?", ("\n" + now + " 출고지시 취소 상태 보정", order_id))
            con.commit()
            raise ValueError("이미 이 출고지시서의 재고 원복 이력이 있습니다.")

        item_count, restored_count = _fallback_restore_outbound_from_items(cur, order_id, now)
        cur.execute("UPDATE outbound_orders SET status='취소됨', memo=IFNULL(memo,'') || ? WHERE id=?", ("\n" + now + " 출고지시 취소", order_id))
        con.commit()
        return item_count, restored_count

def restore_inventory_from_log(order_id):
    """출고지시 거래이력(final_stock + qty)을 기준으로 전체 취소/원복한다."""
    return cancel_outbound_order(int(order_id))

def cancel_saved_order(order_id):
    """저장된 출고지시 취소 버튼에서 호출하는 명시적 래퍼."""
    return restore_inventory_from_log(int(order_id))

def partial_cancel_outbound_order(order_id, cancel_qty_by_item_id):
    """저장된 출고지시 품목 일부를 취소하고 해당 수량만 재고에 되돌린다.

    cancel_qty_by_item_id: {outbound_order_items.id: cancel_qty}
    - 전체 취소는 cancel_saved_order/restore_inventory_from_log를 사용한다.
    - 부분취소는 현재 지시서 품목 기준으로 수량을 줄이고, inventory에는 취소 수량만 더한다.
    """
    order_id = int(order_id)
    clean = {}
    for k, v in (cancel_qty_by_item_id or {}).items():
        try:
            iid = int(k)
            qty = int(float(v or 0))
        except Exception:
            continue
        if qty > 0:
            clean[iid] = qty
    if not clean:
        raise ValueError("부분취소할 수량이 없습니다.")

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with connect() as con:
        con.row_factory = sqlite3.Row
        cur = con.cursor()
        order = cur.execute("SELECT id, status FROM outbound_orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            raise ValueError("출고지시서를 찾을 수 없습니다.")
        if str(order["status"] or "") == "취소됨":
            raise ValueError("이미 취소된 출고지시서는 부분취소할 수 없습니다.")

        item_ids = list(clean.keys())
        placeholders = ",".join(["?"] * len(item_ids))
        rows = cur.execute(f"""
            SELECT id, order_id, inventory_id, location, product_name, lot, exp_date, qty, company, warehouse_name
            FROM outbound_order_items
            WHERE order_id=? AND id IN ({placeholders})
            ORDER BY id
        """, [order_id] + item_ids).fetchall()
        row_map = {int(r["id"]): r for r in rows}
        missing = [iid for iid in item_ids if iid not in row_map]
        if missing:
            raise ValueError(f"출고지시 품목을 찾을 수 없습니다: {missing}")

        restored_lines = 0
        restored_qty_total = 0
        for iid, cancel_qty in clean.items():
            r = row_map[iid]
            original_qty = int(r["qty"] or 0)
            if cancel_qty > original_qty:
                raise ValueError(f"{r['product_name']} 취소수량이 지시수량보다 큽니다. 지시 {original_qty}, 취소 {cancel_qty}")

        for iid, cancel_qty in clean.items():
            r = row_map[iid]
            original_qty = int(r["qty"] or 0)
            remain_qty = original_qty - cancel_qty
            inv_id = r["inventory_id"]
            inv = None
            if inv_id:
                inv = cur.execute("SELECT id, qty FROM inventory WHERE id=?", (int(inv_id),)).fetchone()
            if not inv:
                inv = cur.execute("""
                    SELECT id, qty FROM inventory
                    WHERE company=? AND product_name=? AND IFNULL(warehouse_name,'')=?
                      AND IFNULL(lot,'')=? AND IFNULL(exp_date,'')=? AND location=?
                """, (r["company"] or "", r["product_name"] or "", r["warehouse_name"] or "", r["lot"] or "-", r["exp_date"] or "-", r["location"] or "")).fetchone()
            if inv:
                row_qty_after = int(inv["qty"] or 0) + int(cancel_qty)
                cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (row_qty_after, now, int(inv["id"])))
            else:
                row_qty_after = int(cancel_qty)
                cur.execute("""
                    INSERT INTO inventory(company, product_name, warehouse_name, lot, exp_date, location, qty, updated_at)
                    VALUES(?,?,?,?,?,?,?,?)
                """, (r["company"] or "", r["product_name"] or "", r["warehouse_name"] or "", r["lot"] or "-", r["exp_date"] or "-", r["location"] or "", row_qty_after, now))

            if remain_qty > 0:
                cur.execute("UPDATE outbound_order_items SET qty=? WHERE id=?", (remain_qty, iid))
            else:
                cur.execute("DELETE FROM outbound_order_items WHERE id=?", (iid,))

            insert_transaction_log(cur, created_at=now, tx_type="출고지시부분취소", product_name=r["product_name"] or "", warehouse_name=r["warehouse_name"] or "",
                                   lot=r["lot"] or "-", exp_date=r["exp_date"] or "-", from_company=r["company"] or "", from_location=r["location"] or "",
                                   to_company=r["company"] or "", to_location=r["location"] or "", qty=int(cancel_qty), memo=f"출고지시서 #{order_id} 부분취소")
            restored_lines += 1
            restored_qty_total += int(cancel_qty)

        remaining_count = cur.execute("SELECT COUNT(*) FROM outbound_order_items WHERE order_id=?", (order_id,)).fetchone()[0]
        if int(remaining_count or 0) == 0:
            new_status = "취소됨"
            memo_add = f"\n{now} 출고지시 전체 부분취소 완료"
        else:
            new_status = "수정됨"
            memo_add = f"\n{now} 출고지시 부분취소: {restored_qty_total}EA 원복"
        cur.execute("UPDATE outbound_orders SET status=?, memo=IFNULL(memo,'') || ? WHERE id=?", (new_status, memo_add, order_id))
        con.commit()
        return restored_lines, restored_qty_total, int(remaining_count or 0)

def outbound_inventory(src_id, qty, memo="출고지시 완료"):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with connect() as con:
        cur = con.cursor()
        src = cur.execute("SELECT * FROM inventory WHERE id=?", (src_id,)).fetchone()
        cols = [d[0] for d in cur.description]
        src = dict(zip(cols, src)) if src else None
        if not src:
            raise ValueError("출고 재고를 찾을 수 없습니다.")
        if qty <= 0 or qty > src["qty"]:
            raise ValueError("출고 수량이 현재 재고보다 많거나 올바르지 않습니다.")
        cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (src["qty"]-qty, now, src_id))
        insert_transaction_log(cur, created_at=now, tx_type="출고", product_name=src["product_name"], warehouse_name=src.get("warehouse_name"),
                               lot=src["lot"], exp_date=src["exp_date"], from_company=src["company"], from_location=src["location"],
                               to_company=None, to_location=None, qty=qty, memo=memo)
        con.commit()

def recommend_picks(pick_df, request_qty, expiry_short_first=True):
    """출고 요청 수량만큼 재고 행을 추천한다.
    기본은 유통기한이 빠른 것부터 선택한다.
    expiry_short_first=False이면 해당 범위 안에서 로케이션/LOT 기준으로만 안정 정렬한다.
    반환: (추천행 list, 부족수량)
    """
    rows = []
    try:
        need = int(request_qty or 0)
    except Exception:
        need = 0
    if need <= 0 or pick_df is None or pick_df.empty:
        return rows, max(0, need)

    df = pick_df.copy()
    if expiry_short_first:
        df["_exp_sort"] = pd.to_datetime(df.get("exp_date"), errors="coerce")
        df["_exp_sort"] = df["_exp_sort"].fillna(pd.Timestamp.max)
        df = df.sort_values(["_exp_sort", "location", "lot", "company"], na_position="last")
    else:
        df = df.sort_values(["location", "lot", "company", "exp_date"], na_position="last")

    for r in df.itertuples():
        if need <= 0:
            break
        available = int(getattr(r, "qty", 0) or 0)
        if available <= 0:
            continue
        take = min(available, need)
        rows.append({
            "id": int(getattr(r, "id")),
            "로케이션": getattr(r, "location", ""),
            "사업장": getattr(r, "company", ""),
            "제품명": getattr(r, "product_name", ""),
            "LOT": getattr(r, "lot", "-") or "-",
            "유통기한": display_date_only(getattr(r, "exp_date", "-")),
            "요청수량": int(take),
        })
        need -= take
    return rows, max(0, need)
