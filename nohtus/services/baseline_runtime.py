from io import BytesIO

import pandas as pd

from nohtus.db import q
from nohtus.services.master import match_erp_name
from nohtus.services.stocktake import _baseline_stock_excel_bytes_from_dataframe
from nohtus.services.outbound_runtime import first_nonblank

def baseline_stock_template_excel_bytes():
    """기준재고 업로드용 빈 샘플 양식."""
    sample = pd.DataFrame([
        {
            "사업장": "노투스팜",
            "ERP제품코드": "003",
            "ERP제품명": "JS Tox 100U",
            "표준제품명": "",
            "LOT/제조번호": "NF20CL0901",
            "유통기한": "2027-09-30",
            "로케이션": "A1-01-01",
            "수량": 100,
        },
        {
            "사업장": "비자료",
            "ERP제품코드": "",
            "ERP제품명": "홍보 브로슈어",
            "표준제품명": "",
            "LOT/제조번호": "-",
            "유통기한": "-",
            "로케이션": "홍보물랙",
            "수량": 20,
        },
    ])
    return _baseline_stock_excel_bytes_from_dataframe(sample)

def _baseline_get_product_raw(row):
    return first_nonblank(
        row.get("ERP제품명"), row.get("제품명"), row.get("비자료명"),
        row.get("노투스팜 ERP명"), row.get("NOH ERP명"), row.get("노투스 ERP명")
    )

def _baseline_match_standard(company, product_raw):
    company = (company or "").strip()
    product_raw = (product_raw or "").strip()
    if not company or not product_raw:
        return ""
    if company in ["노투스팜", "NOH", "노투스"]:
        m = match_erp_name(company, product_raw)
        if m.get("status") == "auto" and m.get("candidates"):
            return m["candidates"][0]
        return ""
    if company == "비자료":
        df = q("SELECT standard_name FROM products WHERE TRIM(COALESCE(bidata_name, '')) = ?", (product_raw,))
        if len(df) == 1:
            return str(df.iloc[0]["standard_name"] or "")
        if df.empty:
            same = q("SELECT standard_name FROM products WHERE TRIM(standard_name)=?", (product_raw,))
            if len(same) == 1:
                return str(same.iloc[0]["standard_name"] or "")
    return ""

def _baseline_mapping_payload(company, code, product_raw, standard_name):
    payload = {
        "표준제품명": standard_name or "",
        "노투스팜 ERP명": "",
        "노투스팜 ERP 제품코드": "",
        "NOH ERP명": "",
        "NOH ERP 제품코드": "",
        "노투스 ERP명": "",
        "비자료명": "",
    }
    if company == "노투스팜":
        payload["노투스팜 ERP명"] = product_raw or ""
        payload["노투스팜 ERP 제품코드"] = code or ""
    elif company == "NOH":
        payload["NOH ERP명"] = product_raw or ""
        payload["NOH ERP 제품코드"] = code or ""
    elif company == "노투스":
        payload["노투스 ERP명"] = product_raw or ""
    elif company == "비자료":
        payload["비자료명"] = product_raw or ""
    return payload

def baseline_stock_supplement_excel_bytes(issue_df):
    """보완이 필요한 기준재고 행만 내려받기 위한 엑셀 파일."""
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        out = issue_df.copy() if issue_df is not None else pd.DataFrame()
        out.to_excel(writer, index=False, sheet_name="보완필요")
        ws = writer.book["보완필요"]
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        need_fill = PatternFill("solid", fgColor="FFF2CC")
        widths = {
            "A":34,"B":14,"C":18,"D":34,"E":30,"F":34,"G":24,"H":34,"I":22,
            "J":34,"K":34,"L":18,"M":16,"N":18,"O":10
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        max_row = max(1, len(out) + 1)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:O{max_row}"
        required_headers = {"사업장", "ERP제품명", "표준제품명", "LOT/제조번호", "유통기한", "로케이션", "수량"}
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                elif ws.cell(row=cell.row, column=1).value and (cell.value is None or str(cell.value).strip() == ""):
                    header = ws.cell(row=1, column=cell.column).value
                    if header in required_headers or header in ["노투스팜 ERP명", "NOH ERP명", "노투스 ERP명", "비자료명"]:
                        cell.fill = need_fill
                if cell.column_letter in ["C", "G", "I"]:
                    cell.number_format = "@"
    bio.seek(0)
    return bio.getvalue()

def _row_override_standard_name(row):
    """재고 업로드 파일에 표준제품명/실제제품명 컬럼이 있으면 그 값을 우선 사용한다.
    모든 품목을 매핑할 필요는 없고, JS Tox 100U처럼 ERP명만으로 실제 제품을 알 수 없는 행에만 넣으면 된다.
    """
    for col in ["표준제품명", "WMS표준제품명", "실제제품명", "실제품명"]:
        if col in row.index and not pd.isna(row.get(col)):
            v = str(row.get(col)).strip()
            if v and v.lower() != "nan":
                return v
    return ""
