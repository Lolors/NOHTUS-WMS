import sqlite3
from datetime import datetime
from io import BytesIO

import pandas as pd

import nohtus.pages.stocktake as stocktake_page
import nohtus.services.stocktake as stocktake_service
from nohtus.db import connect, q
from nohtus.dates import display_date_only, normalize_exp_date
from nohtus.services.inventory import insert_transaction_log
from nohtus.services.export_waiting import STAGING_LOCATIONS, ensure_export_waiting_tables


def _inventory_survey_excel_bytes_business(df, sheet_name):
    out = pd.DataFrame()
    out["사업장"] = df["company"] if not df.empty else []
    out["로케이션"] = df["location"] if not df.empty else []
    out["제품명"] = df["product_name"] if not df.empty else []
    out["유통기한"] = df["exp_date"].apply(display_date_only) if not df.empty else []
    out["전산수량"] = df["qty"] if not df.empty else []
    out["실물수량"] = ""

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]
        widths = {"A": 14, "B": 16, "C": 34, "D": 16, "E": 12, "F": 12}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
    bio.seek(0)
    return bio.getvalue()


def full_inventory_excel_bytes_business(exclude_zero=True, exclude_series=None):
    exclude_series = set(exclude_series or ())
    where_sql = "WHERE qty<>0" if exclude_zero else ""
    df = q(
        f"""
        SELECT company, location, product_name, exp_date, qty
        FROM inventory
        {where_sql}
        ORDER BY company, location, product_name, exp_date
        """
    )
    if not df.empty and exclude_series:
        df = df[~df["location"].apply(stocktake_service.location_series).isin(exclude_series)]

    return _inventory_survey_excel_bytes_business(df, "전체재고실사")


def material_inventory_excel_bytes_business(exclude_zero=True):
    """부자재로 분류된 제품만 모은 재고실사용 엑셀(사업장 포함)."""
    placeholders = ",".join("?" for _ in stocktake_service._TRUE_MATERIAL_VALUES)
    zero_sql = "AND qty<>0" if exclude_zero else ""
    df = q(
        f"""
        SELECT company, location, product_name, exp_date, qty
        FROM inventory
        WHERE TRIM(COALESCE(product_name,'')) IN (
            SELECT TRIM(standard_name) FROM products
            WHERE LOWER(TRIM(CAST(COALESCE(is_material, 0) AS TEXT))) IN ({placeholders})
        )
        {zero_sql}
        ORDER BY company, location, product_name, exp_date
        """,
        stocktake_service._TRUE_MATERIAL_VALUES,
    )
    return _inventory_survey_excel_bytes_business(df, "부자재재고실사")


def _merge_inventory_rows(con, keep_id, duplicate_ids):
    """같은 실제 재고키로 합쳐지는 행의 수량과 수출대기 연결 ID를 보존한다."""
    keep_id = int(keep_id)
    for duplicate_id in sorted({int(x) for x in duplicate_ids if int(x) != keep_id}):
        row = con.execute("SELECT qty FROM inventory WHERE id=?", (duplicate_id,)).fetchone()
        if not row:
            continue
        con.execute(
            "UPDATE export_waiting_items SET source_inventory_id=? WHERE source_inventory_id=?",
            (keep_id, duplicate_id),
        )
        con.execute(
            "UPDATE export_waiting_items SET waiting_inventory_id=? WHERE waiting_inventory_id=?",
            (keep_id, duplicate_id),
        )
        con.execute(
            "UPDATE inventory SET qty=COALESCE(qty,0)+? WHERE id=?",
            (int(row[0] or 0), keep_id),
        )
        con.execute("DELETE FROM inventory WHERE id=?", (duplicate_id,))


def _update_inventory_identity(con, inventory_id, product_name, lot, exp_date, now):
    """재고 정보를 수정하고 같은 사업장·제품·LOT·유통기한·로케이션 행은 합친다.

    ERP 원본명(warehouse_name)은 제품매칭 정보이므로 재고 행을 나누는 기준으로
    사용하지 않는다. 로케이션이 다르면 물리적으로 다른 재고이므로 별도 행으로
    유지한다.
    """
    inventory_id = int(inventory_id)
    row = con.execute(
        """SELECT company,COALESCE(warehouse_name,''),location
           FROM inventory WHERE id=?""",
        (inventory_id,),
    ).fetchone()
    if not row:
        return False
    company, _warehouse_name, location = row
    duplicates = con.execute(
        """SELECT id FROM inventory
           WHERE id<>? AND company=? AND product_name=?
             AND COALESCE(lot,'-')=? AND COALESCE(exp_date,'-')=?
             AND location=?
           ORDER BY id""",
        (inventory_id, company, product_name, lot, exp_date, location),
    ).fetchall()
    _merge_inventory_rows(con, inventory_id, [row[0] for row in duplicates])
    con.execute(
        """UPDATE inventory
           SET product_name=?,lot=?,exp_date=?,updated_at=?
           WHERE id=?""",
        (product_name, lot, exp_date, now, inventory_id),
    )
    return bool(duplicates)


def _linked_waiting_items(con, inventory_id, current):
    """선택 재고와 고유 ID 또는 과거의 완전한 재고키로 연결되는 미확정 수출대기를 찾는다.

    고유 ID 연결이 없는 과거 데이터는 전체 재고키가 일치하는 P 행만 묶던
    시절의 로직이 남아 있었는데, 수출대기 보관 위치가 P 외에 T1~T5까지
    선택 가능해진 뒤에도 이 부분은 'P'만 확인해서 T1~T5에 있는 품목은
    LOT/유통기한을 나중에 채워도 수출대기 쪽에 반영되지 않았다."""
    company, warehouse_name, location, _qty, old_lot, old_exp = current[1:]
    old_product_name = str(current[0] or "").strip()
    staging_placeholders = ",".join("?" for _ in STAGING_LOCATIONS)
    return con.execute(
        f"""
        SELECT id,source_inventory_id,waiting_inventory_id,company,product_name,
               COALESCE(warehouse_name,''),COALESCE(lot,'-'),COALESCE(exp_date,'-')
        FROM export_waiting_items
        WHERE COALESCE(confirmed,0)=0
          AND (
              source_inventory_id=?
              OR waiting_inventory_id=?
              OR (
                  ? IN ({staging_placeholders})
                  AND company=?
                  AND product_name=?
                  AND COALESCE(warehouse_name,'')=?
                  AND COALESCE(lot,'-')=?
                  AND COALESCE(exp_date,'-')=?
              )
          )
        ORDER BY id
        """,
        (
            int(inventory_id),
            int(inventory_id),
            location,
            *STAGING_LOCATIONS,
            company,
            old_product_name,
            warehouse_name,
            old_lot,
            old_exp,
        ),
    ).fetchall()


def _synchronize_export_waiting_master(con, inventory_id, current, product_name, lot, exp_date, now):
    """원본 재고·P 재고·미확정 수출대기행을 같은 제품마스터 값으로 원자적으로 갱신한다."""
    ensure_export_waiting_tables(con.cursor())
    rows = _linked_waiting_items(con, inventory_id, current)
    if not rows:
        return _update_inventory_identity(con, inventory_id, product_name, lot, exp_date, now)

    selected_location = str(current[3] or "")
    related_item_ids = {int(row[0]) for row in rows}
    waiting_ids = {int(row[2]) for row in rows if int(row[2] or 0) > 0}
    merged = False

    # 연결 ID가 없던 과거 엑셀 자료는 전체 재고키가 완전히 같은 수출대기
    # 보관 위치(P, T1~T5) 행만 묶는다. 동일 키의 행이 여러 개면 먼저 한
    # 행으로 합쳐 참조 대상을 하나로 만든다.
    staging_placeholders = ",".join("?" for _ in STAGING_LOCATIONS)
    for row in rows:
        item_id, _source_id, waiting_id, company, item_product, warehouse_name, item_lot, item_exp = row
        if int(waiting_id or 0) > 0:
            continue
        candidates = con.execute(
            f"""SELECT id FROM inventory
               WHERE location IN ({staging_placeholders}) AND company=? AND product_name=?
                 AND COALESCE(warehouse_name,'')=?
                 AND COALESCE(lot,'-')=? AND COALESCE(exp_date,'-')=?
               ORDER BY id""",
            (*STAGING_LOCATIONS, company, item_product, warehouse_name, item_lot, item_exp),
        ).fetchall()
        candidate_ids = [int(candidate[0]) for candidate in candidates]
        if not candidate_ids:
            continue
        keep_id = (
            int(inventory_id)
            if selected_location in STAGING_LOCATIONS and int(inventory_id) in candidate_ids
            else candidate_ids[0]
        )
        if any(candidate_id != keep_id for candidate_id in candidate_ids):
            merged = True
        _merge_inventory_rows(con, keep_id, candidate_ids)
        con.execute(
            """UPDATE export_waiting_items
               SET waiting_inventory_id=?
               WHERE COALESCE(confirmed,0)=0
                 AND waiting_inventory_id IS NULL
                 AND company=? AND product_name=?
                 AND COALESCE(warehouse_name,'')=?
                 AND COALESCE(lot,'-')=? AND COALESCE(exp_date,'-')=?""",
            (keep_id, company, item_product, warehouse_name, item_lot, item_exp),
        )
        waiting_ids.add(keep_id)

    if waiting_ids:
        placeholders = ",".join("?" for _ in waiting_ids)
        expanded = con.execute(
            f"""SELECT id FROM export_waiting_items
                WHERE COALESCE(confirmed,0)=0
                  AND waiting_inventory_id IN ({placeholders})""",
            tuple(sorted(waiting_ids)),
        ).fetchall()
        related_item_ids.update(int(row[0]) for row in expanded)

    placeholders = ",".join("?" for _ in related_item_ids)
    final_links = con.execute(
        f"""SELECT source_inventory_id,waiting_inventory_id
            FROM export_waiting_items
            WHERE id IN ({placeholders})""",
        tuple(sorted(related_item_ids)),
    ).fetchall()
    linked_inventory_ids = {int(inventory_id)}
    for source_id, waiting_id in final_links:
        if int(source_id or 0) > 0:
            linked_inventory_ids.add(int(source_id))
        if int(waiting_id or 0) > 0:
            linked_inventory_ids.add(int(waiting_id))

    con.execute(
        f"""UPDATE export_waiting_items
            SET product_name=?,lot=?,exp_date=?
            WHERE id IN ({placeholders})""",
        (product_name, lot, exp_date, *sorted(related_item_ids)),
    )
    for linked_id in sorted(linked_inventory_ids):
        if _update_inventory_identity(con, linked_id, product_name, lot, exp_date, now):
            merged = True
    return merged


def _update_inventory_and_product_mappings_business(inv_id, product_name, lot, exp_date, edited_mappings):
    product_name = str(product_name or "").strip()
    if not product_name:
        raise ValueError("제품명을 입력하세요.")

    lot = stocktake_page._normalize_master_text(lot)
    exp_date = normalize_exp_date(exp_date)
    mapping_rows = edited_mappings.copy()
    if "id" not in mapping_rows.columns:
        mapping_rows.insert(0, "id", None)

    with connect() as con:
        current = con.execute(
            """
            SELECT product_name, company, COALESCE(warehouse_name,''),
                   COALESCE(location,''), qty, COALESCE(lot,'-'), COALESCE(exp_date,'-')
            FROM inventory
            WHERE id=?
            """,
            (int(inv_id),),
        ).fetchone()
        if not current:
            raise ValueError("수정할 재고를 찾을 수 없습니다.")

        old_product_name = str(current[0] or "").strip()
        company = str(current[1] or "").strip()
        warehouse_name = str(current[2] or "").strip()
        location = str(current[3] or "").strip()
        current_qty = int(current[4] or 0)
        old_lot = str(current[5] or "-")
        old_exp = str(current[6] or "-")
        current = (
            old_product_name,
            company,
            warehouse_name,
            location,
            current_qty,
            old_lot,
            old_exp,
        )

        existing_rows = con.execute(
            """
            SELECT id, warehouse_name, image_path
            FROM products
            WHERE TRIM(standard_name)=?
            ORDER BY id
            """,
            (old_product_name,),
        ).fetchall()
        existing_by_id = {
            int(row[0]): {
                "warehouse_name": str(row[1] or ""),
                "image_path": str(row[2] or ""),
            }
            for row in existing_rows
        }

        try:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            merged = _synchronize_export_waiting_master(
                con,
                int(inv_id),
                current,
                product_name,
                lot,
                exp_date,
                now,
            )

            kept_ids = set()
            for _, mapping in mapping_rows.iterrows():
                values = {
                    db_column: stocktake_page._clean_mapping_text(mapping.get(label))
                    for label, db_column in stocktake_page.PRODUCT_MAPPING_DB_COLUMNS.items()
                }
                raw_id = mapping.get("id")
                mapping_id = None
                if not pd.isna(raw_id) and str(raw_id).strip():
                    try:
                        mapping_id = int(raw_id)
                    except (TypeError, ValueError):
                        mapping_id = None

                if mapping_id in existing_by_id:
                    kept_ids.add(mapping_id)
                    preserved = existing_by_id[mapping_id]
                    con.execute(
                        """
                        UPDATE products
                        SET standard_name=?, warehouse_name=?, aliases=?,
                            erp_nohtuspharm_name=?, product_code=?,
                            erp_noh_name=?, erp_noh_code=?,
                            erp_nohtus_name=?, bidata_name=?
                        WHERE id=?
                        """,
                        (
                            product_name,
                            preserved["warehouse_name"] or product_name,
                            values["aliases"],
                            values["erp_nohtuspharm_name"],
                            values["product_code"],
                            values["erp_noh_name"],
                            values["erp_noh_code"],
                            values["erp_nohtus_name"],
                            values["bidata_name"],
                            mapping_id,
                        ),
                    )
                else:
                    con.execute(
                        """
                        INSERT INTO products(
                            product_code, standard_name, warehouse_name, aliases,
                            erp_nohtuspharm_name, erp_nohtus_name, erp_noh_name,
                            erp_noh_code, bidata_name, image_path
                        ) VALUES(?,?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            values["product_code"],
                            product_name,
                            product_name,
                            values["aliases"],
                            values["erp_nohtuspharm_name"],
                            values["erp_nohtus_name"],
                            values["erp_noh_name"],
                            values["erp_noh_code"],
                            values["bidata_name"],
                            "",
                        ),
                    )

            delete_ids = [mapping_id for mapping_id in existing_by_id if mapping_id not in kept_ids]
            if delete_ids:
                placeholders = ",".join("?" for _ in delete_ids)
                con.execute(f"DELETE FROM products WHERE id IN ({placeholders})", delete_ids)

            con.commit()
        except sqlite3.IntegrityError as exc:
            con.rollback()
            raise ValueError("수정 결과와 동일한 재고 또는 제품매칭 정보가 이미 존재합니다.") from exc
        except Exception:
            con.rollback()
            raise

    propagation_error = stocktake_page._propagate_master_edit_to_shipment_items(int(inv_id), product_name, lot, exp_date)

    return product_name, lot, exp_date, len(mapping_rows), propagation_error, merged


def import_stock_survey_excel_business(uploaded_file, replace_current=True):
    normal_df, issue_df = stocktake_service.prepare_baseline_stock_dataframe(uploaded_file)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    inserted = 0
    skipped = int(len(issue_df)) if issue_df is not None else 0
    product_inserted = 0

    with connect() as con:
        cur = con.cursor()
        try:
            if replace_current:
                cur.execute("DELETE FROM inventory")
                cur.execute("DELETE FROM transactions WHERE tx_type='재고조사불러오기'")

            for _, r in normal_df.iterrows():
                company = str(r.get("사업장") or "").strip()
                code = str(r.get("ERP제품코드") or "").strip()
                product_raw = str(r.get("ERP제품명") or "").strip()
                product = str(r.get("표준제품명") or "").strip()
                lot = str(r.get("LOT/제조번호") or "").strip() or "-"
                exp = stocktake_service._excel_date_to_iso(r.get("유통기한"))
                loc = str(r.get("로케이션") or "").strip()
                qty = int(float(r.get("수량") or 0))
                if not company or not product or not loc or qty <= 0:
                    skipped += 1
                    continue

                exists = cur.execute(
                    "SELECT id FROM products WHERE standard_name=? ORDER BY id LIMIT 1",
                    (product,),
                ).fetchone()
                if not exists:
                    cur.execute(
                        """
                        INSERT INTO products(
                            product_code, standard_name, warehouse_name, aliases,
                            erp_nohtuspharm_name, erp_noh_name, erp_noh_code,
                            erp_nohtus_name, bidata_name
                        ) VALUES(?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            code if company == "노투스팜" else "",
                            product,
                            product_raw,
                            "",
                            product_raw if company == "노투스팜" else "",
                            product_raw if company == "NOH" else "",
                            code if company == "NOH" else "",
                            product_raw if company == "노투스" else "",
                            product_raw if company == "비자료" else "",
                        ),
                    )
                    product_inserted += 1
                else:
                    pid = int(exists[0])
                    if company == "노투스팜":
                        cur.execute(
                            "UPDATE products SET erp_nohtuspharm_name=COALESCE(NULLIF(erp_nohtuspharm_name,''), ?), product_code=COALESCE(NULLIF(product_code,''), ?) WHERE id=?",
                            (product_raw, code, pid),
                        )
                    elif company == "NOH":
                        cur.execute(
                            "UPDATE products SET erp_noh_name=COALESCE(NULLIF(erp_noh_name,''), ?), erp_noh_code=COALESCE(NULLIF(erp_noh_code,''), ?) WHERE id=?",
                            (product_raw, code, pid),
                        )
                    elif company == "노투스":
                        cur.execute(
                            "UPDATE products SET erp_nohtus_name=COALESCE(NULLIF(erp_nohtus_name,''), ?) WHERE id=?",
                            (product_raw, pid),
                        )
                    elif company == "비자료":
                        cur.execute(
                            "UPDATE products SET bidata_name=COALESCE(NULLIF(bidata_name,''), ?) WHERE id=?",
                            (product_raw, pid),
                        )

                existing_inventory = cur.execute(
                    """
                    SELECT id, qty
                    FROM inventory
                    WHERE company=?
                      AND product_name=?
                      AND COALESCE(warehouse_name,'')=?
                      AND COALESCE(lot,'-')=?
                      AND COALESCE(exp_date,'-')=?
                      AND location=?
                    ORDER BY id
                    LIMIT 1
                    """,
                    (company, product, product_raw, lot, exp, loc),
                ).fetchone()

                if existing_inventory:
                    inventory_id = int(existing_inventory[0])
                    merged_qty = int(existing_inventory[1] or 0) + qty
                    cur.execute(
                        "UPDATE inventory SET qty=?, updated_at=? WHERE id=?",
                        (merged_qty, now, inventory_id),
                    )
                else:
                    cur.execute(
                        """
                        INSERT INTO inventory(
                            company, product_name, warehouse_name, lot,
                            exp_date, location, qty, updated_at
                        ) VALUES(?,?,?,?,?,?,?,?)
                        """,
                        (company, product, product_raw, lot, exp, loc, qty, now),
                    )

                insert_transaction_log(
                    cur,
                    created_at=now,
                    tx_type="재고조사불러오기",
                    product_name=product,
                    warehouse_name=product_raw,
                    lot=lot,
                    exp_date=exp,
                    from_company=None,
                    from_location=None,
                    to_company=company,
                    to_location=loc,
                    qty=qty,
                    memo=f"기준재고 엑셀 업로드 / 원본명: {product_raw}",
                )
                inserted += 1

            con.commit()
        except Exception:
            con.rollback()
            raise

    return inserted, skipped, product_inserted, skipped


def page_stocktake():
    original_full_inventory_excel_bytes = stocktake_service.full_inventory_excel_bytes
    original_material_inventory_excel_bytes = stocktake_service.material_inventory_excel_bytes
    original_import_stock_survey_excel = stocktake_service.import_stock_survey_excel
    original_update_master = stocktake_page._update_inventory_and_product_mappings

    stocktake_service.full_inventory_excel_bytes = full_inventory_excel_bytes_business
    stocktake_service.material_inventory_excel_bytes = material_inventory_excel_bytes_business
    stocktake_service.import_stock_survey_excel = import_stock_survey_excel_business
    stocktake_page._update_inventory_and_product_mappings = _update_inventory_and_product_mappings_business
    try:
        return stocktake_page.page_stocktake()
    finally:
        stocktake_service.full_inventory_excel_bytes = original_full_inventory_excel_bytes
        stocktake_service.material_inventory_excel_bytes = original_material_inventory_excel_bytes
        stocktake_service.import_stock_survey_excel = original_import_stock_survey_excel
        stocktake_page._update_inventory_and_product_mappings = original_update_master
