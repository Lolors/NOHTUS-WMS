from pathlib import Path
from datetime import datetime
from io import BytesIO
import re

import pandas as pd
import streamlit as st

from nohtus.db import q

PROJECT_ROOT = Path(__file__).resolve().parents[2]

def customer_master_template_bytes():
    df = pd.DataFrame(columns=["거래처코드", "거래처명", "사업장", "유형", "담당자", "연락처", "주소", "메모"])
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="거래처관리")
        ws = writer.book["거래처관리"]
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill = PatternFill("solid", fgColor="E5E7EB")
        widths = {"A":18,"B":30,"C":16,"D":16,"E":18,"F":18,"G":44,"H":30}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=8):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = fill
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:H2"
    bio.seek(0)
    return bio.getvalue()

def detect_column(columns, candidates):
    cols = [str(c).strip() for c in columns]
    for cand in candidates:
        for c in cols:
            if cand.lower() == c.lower():
                return c
    for cand in candidates:
        for c in cols:
            if cand.lower() in c.lower():
                return c
    return cols[0] if cols else None

def clean_excel_text(value):
    """openpyxl이 저장하지 못하는 제어문자와 특수 공백을 제거한다."""
    if value is None:
        return ""
    text = str(value)
    try:
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        text = ILLEGAL_CHARACTERS_RE.sub("", text)
    except Exception:
        text = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", text)
    text = re.sub(r"[\x00-\x1F\x7F-\x9F\u200b\u200c\u200d\ufeff]", "", text)
    text = text.replace(" ", " ")
    return text.strip()

def _erp_name_key(value):
    text = clean_excel_text(value)
    return re.sub(r"\s+", "", text).replace("[", "").replace("]", "")

def is_ignored_erp_product_name(value):
    key = _erp_name_key(value)
    return (not key) or key in {"합계", "배송비"} or "합계" in key or "배송비" in key

def _find_required_column(columns, candidates):
    cols = [str(c).strip() for c in columns]
    for cand in candidates:
        for col in cols:
            if col == cand:
                return col
    for cand in candidates:
        for col in cols:
            if cand in col:
                return col
    return None

def _read_erp_current_file(uploaded, company, key_prefix):
    """ERP 현재고 파일을 고정 규칙으로 읽는다.
    - 노투스팜/NOH: 제품명, 현재고수량
    - 노투스: 8행 header, 품목명/규격, 현재재고
    제조번호/유통기한은 ERP 비교에서 사용하지 않는다.
    """
    if uploaded is None:
        return None
    try:
        if company == "노투스":
            raw = pd.read_excel(uploaded, header=7)
            name_col = _find_required_column(raw.columns, ["품목명/규격", "품목명", "제품명"])
            qty_col = _find_required_column(raw.columns, ["현재재고", "현재고", "현재고수량"])
        else:
            raw = pd.read_excel(uploaded)
            name_col = _find_required_column(raw.columns, ["제품명"])
            qty_col = _find_required_column(raw.columns, ["현재고수량", "현재고"])
    except Exception as e:
        st.error(f"{company} 엑셀 읽기 실패: {e}")
        return None
    if raw.empty:
        st.warning(f"{company} 엑셀에 데이터가 없습니다.")
        return None
    raw.columns = [str(c).strip() for c in raw.columns]
    if not name_col or not qty_col or name_col not in raw.columns or qty_col not in raw.columns:
        st.error(f"{company} ERP 파일에서 필요한 컬럼을 찾을 수 없습니다.")
        st.caption("노투스팜/NOH: 제품명, 현재고수량 · 노투스: 8행 헤더의 품목명/규격, 현재재고")
        return None
    return {"company": company, "raw": raw, "name_col": name_col, "qty_col": qty_col}

def _normalize_erp_current_rows(parsed_files):
    """ERP 파일을 사업장 + ERP 제품명 기준으로 합산한다.
    ERP 파일은 이미 ERP명으로 되어 있으므로 제품매칭표 역변환은 WMS 쪽에만 적용한다.
    """
    rows = []
    fail = []
    amb = []
    auto_count = 0
    for info in parsed_files:
        if not info:
            continue
        company = info["company"]
        raw = info["raw"].copy()
        name_col = info["name_col"]
        qty_col = info["qty_col"]
        raw[name_col] = raw[name_col].apply(clean_excel_text)
        raw[qty_col] = pd.to_numeric(raw[qty_col], errors="coerce").fillna(0).astype(int)
        raw = raw[~raw[name_col].apply(is_ignored_erp_product_name)]
        raw = raw[raw[qty_col] != 0]
        if raw.empty:
            continue
        for _, rr in raw.iterrows():
            erp_name = clean_excel_text(rr[name_col])
            erp_qty = int(rr[qty_col])
            if is_ignored_erp_product_name(erp_name):
                continue
            rows.append({"사업장": company, "제품명": erp_name, "ERP수량": erp_qty})
            auto_count += 1
    if rows:
        erp = pd.DataFrame(rows)
        erp = erp.groupby(["사업장", "제품명"], as_index=False)["ERP수량"].sum()
    else:
        erp = pd.DataFrame(columns=["사업장", "제품명", "ERP수량"])
    return erp, pd.DataFrame(amb), pd.DataFrame(fail), auto_count

def mapping_source_column_for_company(company):
    """사업장별 원본명 컬럼을 반환한다. 비자료는 ERP명이 아니라 비자료명 컬럼만 사용한다."""
    if company == "노투스팜":
        return "erp_nohtuspharm_name", "노투스팜 ERP명"
    if company == "NOH":
        return "erp_noh_name", "NOH ERP명"
    if company == "노투스":
        return "erp_nohtus_name", "노투스 ERP명"
    if company == "비자료":
        return "bidata_name", "비자료명"
    return None, "원본명"

def is_mapping_conflict_approved(company, source_name):
    company = (company or "").strip()
    source_name = (source_name or "").strip()
    if not company or not source_name:
        return False
    df = q("""SELECT 1 FROM product_match_conflict_approvals
             WHERE company=? AND source_name=? LIMIT 1""", (company, source_name))
    return not df.empty

def page_erp_data_upload():
    st.title("ERP 데이터 업로드")
    st.caption("매일 아침 노투스팜/NOH/노투스 ERP 데이터를 각각 업로드합니다.")
    upload_dir = PROJECT_ROOT / "data" / "erp_uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    upload_area = st.container()
    with upload_area:
        cols = st.columns(3, gap="large")
        info = [("노투스팜", "SIMS"), ("NOH", "SIMS"), ("노투스", "IBK우리은행 전산")]
        for col, (company, system_name) in zip(cols, info):
            with col:
                st.markdown(f"### {company}")
                st.caption(system_name)
                up = st.file_uploader(f"{company} ERP 엑셀", type=["xlsx","xls"], key=f"erp_upload_{company}")
                if up is not None:
                    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    safe = f"{company}_{stamp}_{up.name}"
                    path = upload_dir / safe
                    path.write_bytes(up.getvalue())
                    st.success("업로드 파일 저장 완료")
                    try:
                        df = pd.read_excel(BytesIO(up.getvalue()))
                        st.dataframe(df.head(10), hide_index=True, use_container_width=True)
                    except Exception as e:
                        st.warning(f"미리보기 실패: {e}")
    st.markdown("---")
    files = sorted(upload_dir.glob("*"), reverse=True)
    st.markdown("### 최근 업로드 파일")
    if not files:
        st.info("아직 업로드된 ERP 파일이 없습니다.")
    else:
        recent = pd.DataFrame([{"파일명": f.name, "크기KB": round(f.stat().st_size/1024,1)} for f in files[:30]])
        st.dataframe(recent, hide_index=True, use_container_width=True)
