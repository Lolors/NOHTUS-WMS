from __future__ import annotations

from pathlib import Path
import shutil
from zipfile import BadZipFile, ZipFile

from nohtus.export_app import db
from nohtus.export_app.services import folder_service


SECTION_TITLES = {'기본 정보', '주문 목록', '출고 진행 상황', '미패킹 제품', '국내배송 정보'}


def _style_sheet(ws, widths: dict[str, float], header_rows: set[int] | None = None) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_rows = header_rows or set()
    header_fill = PatternFill('solid', fgColor='D9EAF7')
    section_fill = PatternFill('solid', fgColor='EAF2F8')
    thin = Side(style='thin', color='B8C2CC')
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, size=14)
            if cell.value is not None:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_idx in range(1, ws.max_row + 1):
        first = ws.cell(row_idx, 1)
        if first.value in SECTION_TITLES:
            for cell in ws[row_idx]:
                cell.fill = section_fill
                cell.font = Font(bold=True)
        elif row_idx in header_rows:
            for cell in ws[row_idx]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = 'A2'


def _box_size_text(row) -> str:
    values = [row['length_cm'], row['width_cm'], row['height_cm']]
    if not all(value not in (None, '') for value in values):
        return ''
    return ' × '.join(f'{float(value):g}' for value in values)


def _validate_xlsx(path: Path) -> None:
    try:
        with ZipFile(path, 'r') as archive:
            bad_member = archive.testzip()
            names = set(archive.namelist())
    except (BadZipFile, OSError) as exc:
        raise ValueError(f'생성된 엑셀 파일이 손상되었습니다: {path}') from exc
    if bad_member:
        raise ValueError(f'생성된 엑셀 내부 파일이 손상되었습니다: {bad_member}')
    required = {'[Content_Types].xml', 'xl/workbook.xml'}
    if not required.issubset(names):
        raise ValueError(f'생성된 엑셀 파일 구성이 올바르지 않습니다: {path}')


def is_valid_xlsx(path: Path) -> bool:
    try:
        _validate_xlsx(path)
    except ValueError:
        return False
    return True


def write_case_workbook(case_id: int, folder: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    case = db.row('SELECT * FROM export_cases WHERE id=?', (case_id,))
    if not case:
        raise ValueError(f'수출 건을 찾을 수 없습니다: {case_id}')

    folder.mkdir(parents=True, exist_ok=True)
    workbook_path = folder / '수출진행내역.xlsx'

    orders = db.rows(
        'SELECT id, product_name, quantity, unit, created_at FROM order_items WHERE case_id=? ORDER BY id',
        (case_id,),
    )
    shipments = db.rows(
        '''SELECT s.id AS shipment_id, s.order_item_id, s.business_unit,
                  s.product_name AS actual_product_name, s.lot_no, s.expiry_date,
                  s.requested_qty, s.box_no, s.updated_at,
                  o.product_name AS order_product_name, o.quantity AS order_quantity, o.unit,
                  b.length_cm, b.width_cm, b.height_cm, b.weight_kg
           FROM shipment_items s
           JOIN order_items o ON o.id=s.order_item_id AND o.case_id=s.case_id
           LEFT JOIN boxes b ON b.case_id=s.case_id AND b.box_no=s.box_no
           WHERE s.case_id=?
           ORDER BY CASE WHEN s.box_no IS NULL THEN 1 ELSE 0 END,
                    s.box_no, s.id''',
        (case_id,),
    )

    wb = Workbook()

    ws1 = wb.active
    ws1.title = '주문 접수 내역'
    ws1.append(['주문 접수 내역'])
    ws1.append(['기본 정보'])
    ws1.append(['수출번호', '국가', '바이어', '운송방식', '진행단계', '상태', '비고', '생성일', '수정일'])
    ws1.append([
        case['export_no'], case['country'], case['buyer'], case['transport_mode'],
        case['stage'], case['status'], case['note'], case['created_at'], case['updated_at'],
    ])
    ws1.append([])
    ws1.append(['주문 목록'])
    ws1.append(['제품명', '수량', '단위', '등록일'])
    for item in orders:
        ws1.append([item['product_name'], item['quantity'], item['unit'], item['created_at']])
    _style_sheet(
        ws1,
        {'A': 28, 'B': 14, 'C': 18, 'D': 14, 'E': 16, 'F': 12, 'G': 30, 'H': 20, 'I': 20},
        {3, 7},
    )

    ws2 = wb.create_sheet('출고 진행 상황')
    ws2.append(['출고 진행 상황'])
    ws2.append([
        'CTN No.', '제품명', '사업장', '제조번호', '유통기한',
        '수량', '단위', 'GW (kg)', 'CTN Size (cm)', '수정일',
    ])

    packed_by_box: dict[int, list] = {}
    unpacked: list = []
    for shipment in shipments:
        if shipment['box_no'] is None:
            unpacked.append(shipment)
        else:
            packed_by_box.setdefault(int(shipment['box_no']), []).append(shipment)

    box_fill = PatternFill('solid', fgColor='F5F8FB')
    total_quantity = 0.0
    total_gw = 0.0

    for box_no in sorted(packed_by_box):
        box_items = packed_by_box[box_no]
        first_row = ws2.max_row + 1
        box_weight = float(box_items[0]['weight_kg'] or 0)
        box_size = _box_size_text(box_items[0])
        total_gw += box_weight

        for index, shipment in enumerate(box_items):
            quantity = float(shipment['requested_qty'] or 0)
            total_quantity += quantity
            ws2.append([
                box_no if index == 0 else '',
                shipment['actual_product_name'] or shipment['order_product_name'] or '',
                shipment['business_unit'] or '',
                shipment['lot_no'] or '',
                shipment['expiry_date'] or '',
                quantity,
                shipment['unit'] or '',
                box_weight if index == 0 else '',
                box_size if index == 0 else '',
                shipment['updated_at'] or '',
            ])
            row_no = ws2.max_row
            ws2.cell(row_no, 6).number_format = '#,##0'
            if index == 0:
                ws2.cell(row_no, 8).number_format = '0.00" kg"'
                for cell in ws2[row_no]:
                    cell.fill = box_fill

        last_row = ws2.max_row
        if last_row > first_row:
            for column in (1, 8, 9):
                ws2.merge_cells(
                    start_row=first_row,
                    start_column=column,
                    end_row=last_row,
                    end_column=column,
                )
        for column in (1, 8, 9):
            ws2.cell(first_row, column).alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True
            )

    if unpacked:
        ws2.append([])
        unpacked_title_row = ws2.max_row + 1
        ws2.append(['미패킹 제품'])
        for cell in ws2[unpacked_title_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='FFF2CC')
        for shipment in unpacked:
            quantity = float(shipment['requested_qty'] or 0)
            total_quantity += quantity
            ws2.append([
                '미패킹',
                shipment['actual_product_name'] or shipment['order_product_name'] or '',
                shipment['business_unit'] or '',
                shipment['lot_no'] or '',
                shipment['expiry_date'] or '',
                quantity,
                shipment['unit'] or '',
                '',
                '',
                shipment['updated_at'] or '',
            ])
            ws2.cell(ws2.max_row, 6).number_format = '#,##0'

    total_row = ws2.max_row + 1
    ws2.append(['합계', '', '', '', '', total_quantity, '', total_gw, '', ''])
    ws2.cell(total_row, 6).number_format = '#,##0'
    ws2.cell(total_row, 8).number_format = '0.00" kg"'
    for cell in ws2[total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9EAF7')
    ws2.cell(total_row, 1).alignment = Alignment(horizontal='center', vertical='center')

    _style_sheet(
        ws2,
        {
            'A': 11, 'B': 34, 'C': 14, 'D': 18, 'E': 16,
            'F': 12, 'G': 10, 'H': 12, 'I': 22, 'J': 20,
        },
        {2},
    )

    ws3 = wb.create_sheet('국내배송 정보')
    ws3.append(['국내배송 정보'])
    ws3.append(['항목', '내용'])
    for label, value in [
        ('국내배송 방식', case['domestic_method']), ('국내배송 일자', case['actual_ship_date']),
        ('수하인명', case['consignee_name']), ('수하인주소', case['consignee_address']),
        ('송장번호', case['tracking_no']), ('배송기사 이름', case['driver_name']),
        ('배송기사 연락처', case['driver_phone']), ('현재 단계', case['stage']),
        ('상태', case['status']), ('비고', case['note']), ('취소 사유', case['cancel_reason']),
        ('취소 일시', case['cancelled_at']), ('최종 수정일', case['updated_at']),
    ]:
        ws3.append([label, value or ''])
    _style_sheet(ws3, {'A': 24, 'B': 64}, {2})

    temporary_path = workbook_path.with_name(f'{workbook_path.stem}.tmp.xlsx')
    backup_folder = folder / folder_service.BACKUP_FOLDER_NAME
    previous_path = backup_folder / f'{workbook_path.stem}.previous.xlsx'
    previous_temporary_path = backup_folder / f'{workbook_path.stem}.previous.tmp.xlsx'
    legacy_previous_path = workbook_path.with_name(f'{workbook_path.stem}.previous.xlsx')

    if legacy_previous_path.exists():
        backup_folder.mkdir(parents=True, exist_ok=True)
        folder_service.set_hidden(backup_folder)
        legacy_destination = previous_path
        if legacy_destination.exists():
            legacy_destination = backup_folder / f'{workbook_path.stem}.legacy.previous.xlsx'
        if not legacy_destination.exists():
            legacy_previous_path.replace(legacy_destination)

    if temporary_path.exists():
        temporary_path.unlink()
    wb.save(temporary_path)
    _validate_xlsx(temporary_path)

    if workbook_path.exists() and is_valid_xlsx(workbook_path):
        backup_folder.mkdir(parents=True, exist_ok=True)
        folder_service.set_hidden(backup_folder)
        if previous_temporary_path.exists():
            previous_temporary_path.unlink()
        shutil.copy2(workbook_path, previous_temporary_path)
        _validate_xlsx(previous_temporary_path)
        previous_temporary_path.replace(previous_path)

    temporary_path.replace(workbook_path)
    _validate_xlsx(workbook_path)
    return workbook_path
