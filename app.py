import sqlite3
from pathlib import Path
from datetime import datetime, date
import calendar
import json
import re
from io import BytesIO
from html import escape
from urllib.parse import quote
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

APP_TITLE = "NOHTUS WMS"

############################################################
# RC3.0 STABLE BASE 개발 원칙
#
# [CORE FREEZE / 절대 수정 금지]
# - 입고도면 클릭 및 입고 위치 연동
# - 로케이션맵 상세보기
# - 로케이션맵 제품명 클릭 -> 제품검색 자동 실행
# - 도면 SVG / 클릭 JS / query parameter 연동
#
# RC2.82는 위 기능이 정상 작동하던 안정 기준입니다.
# 이후 기능 추가는 이 코어를 직접 수정하지 않고,
# CSS/UI/서비스 함수 레이어에서만 확장하는 방식으로 진행합니다.
############################################################
VERSION = "v4.9 RC3.0 Stable Base"
DB_PATH = Path(__file__).parent / "data" / "nohtus.db"
COMPANIES = ["노투스팜", "노투스", "NOH", "비자료"]
INBOUND_COMPANIES = COMPANIES + ["등록대기"]
SPECIAL_LOCATIONS = ["홍보물랙", "회색 카트", "오른쪽 창고", "사무실(4층)"]

AREA_CONFIG = {
    "A1": {"lines": ["01","02","03","04","05","06"], "levels": ["01","02","03"]},
    "A2": {"lines": ["01","02","03","04","05","06"], "levels": ["01","02","03"]},
    "B1": {"lines": ["01","02","03","04","05","06"], "levels": ["01","02","03"]},
    "B2": {"lines": ["01","02","03","04","05","06"], "levels": ["01","02","03"]},
    "C1": {"lines": ["01","02","03","04","05","06"], "levels": ["01","02","03"]},
    "C2": {"lines": ["01","02","03","04","05","06"], "levels": ["01","02","03"]},
    "D1": {"lines": ["01","02","03","04","05","06"], "levels": ["01","02","03"]},
    "E1": {"lines": ["01","02","03","04","05","06"], "levels": ["01","02","03"]},
    "F1": {"lines": ["01","02","03"], "levels": ["01","02","03"]},
    "G1": {"lines": ["01","02","03"], "levels": ["01","02","03"]},
    "G2": {"lines": [], "levels": []},
    "T1": {"lines": [], "levels": []},
    "T2": {"lines": [], "levels": []},
    "X1": {"lines": ["01","02","03"], "levels": ["01","02","03","04"]},
    "X2": {"lines": [], "levels": []},
    "REC": {"lines": [], "levels": []},
    "Q": {"lines": ["Q1","Q2"], "levels": []},
    "P": {"lines": [], "levels": []},
    "R1": {"lines": [], "levels": []},
    "R2": {"lines": [], "levels": []},
    "N": {"lines": SPECIAL_LOCATIONS, "levels": []},
}

AREA_COLOR = {
    "A1":"yellow", "A2":"yellow", "B1":"yellow", "B2":"yellow", "C1":"yellow",
    "C2":"blue", "D1":"blue",
    "E1":"pink", "Q":"pink",
    "F1":"bidata", "G1":"gray", "G2":"gray", "X1":"gray", "X2":"gray", "N":"gray",
    "REC":"white", "P":"white", "R1":"white", "R2":"white", "T1":"white", "T2":"white"
}

# ---------------- DB ----------------
def connect():
    DB_PATH.parent.mkdir(exist_ok=True)
    return sqlite3.connect(DB_PATH)

def init_db():
    con = connect(); cur = con.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS products(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_code TEXT,
        standard_name TEXT NOT NULL,
        warehouse_name TEXT,
        aliases TEXT
    )
    """)
    # v3.6: ERP별 제품명/대체 매칭 메모 컬럼 추가. 기존 DB도 자동 보강.
    product_cols = {r[1] for r in cur.execute("PRAGMA table_info(products)").fetchall()}
    for col in ["erp_nohtuspharm_name", "erp_nohtus_name", "erp_noh_name", "erp_noh_code", "bidata_name", "substitute_note", "image_path"]:
        if col not in product_cols:
            cur.execute(f"ALTER TABLE products ADD COLUMN {col} TEXT")
    cur.execute("""
    CREATE TABLE IF NOT EXISTS inventory(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company TEXT NOT NULL,
        product_name TEXT NOT NULL,
        warehouse_name TEXT,
        lot TEXT,
        exp_date TEXT,
        location TEXT NOT NULL,
        qty INTEGER NOT NULL DEFAULT 0,
        updated_at TEXT
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS transactions(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        created_at TEXT NOT NULL,
        tx_type TEXT NOT NULL,
        product_name TEXT NOT NULL,
        warehouse_name TEXT,
        lot TEXT,
        exp_date TEXT,
        from_company TEXT,
        from_location TEXT,
        to_company TEXT,
        to_location TEXT,
        qty INTEGER NOT NULL,
        memo TEXT
    )
    """)
    tx_cols = {r[1] for r in cur.execute("PRAGMA table_info(transactions)").fetchall()}
    if "final_stock" not in tx_cols:
        cur.execute("ALTER TABLE transactions ADD COLUMN final_stock INTEGER")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS erp_stock(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        uploaded_at TEXT,
        company TEXT,
        product_name TEXT,
        lot TEXT,
        exp_date TEXT,
        qty INTEGER
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS erp_ambiguous_candidates(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        erp_company TEXT NOT NULL,
        erp_name TEXT NOT NULL,
        candidate_product TEXT NOT NULL,
        memo TEXT
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS customers(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_code TEXT,
        customer_name TEXT NOT NULL,
        manager TEXT,
        phone TEXT,
        address TEXT,
        memo TEXT,
        updated_at TEXT
    )
    """)
    customer_cols = {r[1] for r in cur.execute("PRAGMA table_info(customers)").fetchall()}
    for col in ["company", "customer_type"]:
        if col not in customer_cols:
            cur.execute(f"ALTER TABLE customers ADD COLUMN {col} TEXT")
    cur.execute("""
    CREATE TABLE IF NOT EXISTS erp_upload_decisions(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        decided_at TEXT NOT NULL,
        erp_company TEXT NOT NULL,
        erp_name TEXT NOT NULL,
        selected_product TEXT NOT NULL,
        memo TEXT
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS outbound_orders(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        created_at TEXT NOT NULL,
        order_date TEXT NOT NULL,
        title TEXT,
        status TEXT DEFAULT '저장됨',
        memo TEXT
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS outbound_order_items(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        inventory_id INTEGER,
        location TEXT,
        product_name TEXT,
        lot TEXT,
        exp_date TEXT,
        qty INTEGER NOT NULL,
        company TEXT,
        warehouse_name TEXT,
        FOREIGN KEY(order_id) REFERENCES outbound_orders(id)
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS product_match_conflict_approvals(
        company TEXT NOT NULL,
        source_name TEXT NOT NULL,
        approved_at TEXT,
        PRIMARY KEY(company, source_name)
    )
    """)
    # v3.8: 더미데이터 자동 생성 중단.
    # 처음 실행 시 제품/재고는 비어 있으며, 제품마스터 엑셀 또는 재고조사 엑셀 업로드로 채웁니다.
    con.commit(); con.close()

def q(sql, params=()):
    with connect() as con:
        return pd.read_sql_query(sql, con, params=params)

def exec_sql(sql, params=()):
    with connect() as con:
        con.execute(sql, params); con.commit()

def normalize_blank(v):
    v = (v or "").strip()
    return v if v else "-"

def normalize_exp_date(v):
    """유통기한 입력값을 YYYY-MM-DD로 정규화.
    예: 28/3/2, 28.3.2 -> 2028-03-02.
    pandas 자동 파싱보다 직접 파싱을 우선하여 YY/M/D가 DD/M/YY로 뒤집히는 문제를 막는다.
    """
    v = (v or "").strip()
    if not v:
        return "-"

    # 엑셀/판다스 날짜 객체
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")

    raw = str(v).strip()
    if not raw or raw.lower() == "nan" or raw == "-":
        return "-"

    # "2026-04-30 00:00:00" 같은 문자열은 날짜 부분만 우선 사용
    raw = raw.split(" ")[0].strip()
    compact = raw.replace("/", ".").replace("-", ".").replace("_", ".").replace(" ", "")

    try:
        # 20280302
        if compact.isdigit() and len(compact) == 8:
            y, m, d = int(compact[:4]), int(compact[4:6]), int(compact[6:8])
            return f"{y:04d}-{m:02d}-{d:02d}"

        # 280302
        if compact.isdigit() and len(compact) == 6:
            y, m, d = int(compact[:2]), int(compact[2:4]), int(compact[4:6])
            y = 2000 + y if y < 100 else y
            return f"{y:04d}-{m:02d}-{d:02d}"

        parts = [x for x in compact.split(".") if x != ""]
        if len(parts) == 3:
            # 반드시 년-월-일 순서로 해석한다.
            y, m, d = map(int, parts)
            y = 2000 + y if y < 100 else y
            return f"{y:04d}-{m:02d}-{d:02d}"

        if len(parts) == 2:
            y, m = map(int, parts)
            y = 2000 + y if y < 100 else y
            return f"{y:04d}-{m:02d}-01"
    except Exception:
        pass

    # 위 규칙에 걸리지 않는 경우만 pandas에 맡긴다.
    dt = pd.to_datetime(raw, errors="coerce")
    if pd.notna(dt):
        return dt.strftime("%Y-%m-%d")
    return raw





def expiry_status(exp_date):
    """정상 / 임박(6개월) / 만료 자동 판정."""
    exp = (exp_date or "").strip()
    if not exp or exp == "-":
        return "정상"
    try:
        d = datetime.strptime(exp, "%Y-%m-%d").date()
    except Exception:
        return "정상"
    today = date.today()
    if d < today:
        return "만료"
    if (d - today).days <= 183:
        return "임박(6개월)"
    return "정상"


def product_master_excel_bytes(highlight_missing=False):
    """제품 마스터를 사용자가 수정하기 쉬운 엑셀 양식으로 내보낸다.
    v3.7부터 제품코드는 노투스팜 ERP 전용 코드로 취급하고, 전산상 명칭은 제품마스터에서 제외한다.
    """
    df = q("SELECT standard_name, erp_nohtuspharm_name, product_code, erp_noh_name, erp_noh_code, erp_nohtus_name, bidata_name, aliases FROM products ORDER BY standard_name")
    out = df.rename(columns={
        "standard_name": "표준제품명",
        "erp_nohtuspharm_name": "노투스팜 ERP명",
        "product_code": "노투스팜 ERP 제품코드",
        "erp_noh_name": "NOH ERP명",
        "erp_noh_code": "NOH ERP 제품코드",
        "erp_nohtus_name": "노투스 ERP명",
        "bidata_name": "비자료명",
        "aliases": "별칭",
    })
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="제품마스터")
        ws = writer.book["제품마스터"]
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        need_fill = PatternFill("solid", fgColor="FFF2CC")
        widths = {"A":24,"B":34,"C":28,"D":34,"E":28,"F":34,"G":34,"H":34}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:H{max(1, len(out)+1)}"
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                # ERP 제품코드는 계산값이 아니라 텍스트다. 003 같은 앞자리 0을 보존한다.
                if cell.column_letter in ["C", "E"]:
                    cell.number_format = "@"
                    if cell.value is not None:
                        cell.value = str(cell.value)
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                elif highlight_missing:
                    row_values = [ws.cell(row=cell.row, column=i).value for i in [2,4,6,7]]
                    if all(str(v or "").strip() == "" for v in row_values):
                        cell.fill = need_fill
    bio.seek(0)
    return bio.getvalue()


def import_product_master_excel(uploaded_file):
    """업로드된 제품매칭표 엑셀을 products 테이블에 완전 교체 반영한다.
    기존 제품명/ERP명 데이터가 남지 않도록 업로드 파일 기준으로 products를 다시 만든다.
    """
    df = pd.read_excel(uploaded_file, dtype=str).fillna("")
    rename = {
        "노투스팜 ERP 제품코드": "product_code",
        "제품코드": "product_code",
        "표준제품명": "standard_name",
        "제품명": "standard_name",
        "별칭": "aliases",
        "노투스팜 ERP명": "erp_nohtuspharm_name",
        "NOH ERP명": "erp_noh_name",
        "NOH ERP 제품코드": "erp_noh_code",
        "노투스 ERP명": "erp_nohtus_name",
        "비자료명": "bidata_name",
    }
    df = df.rename(columns={c: rename.get(c, c) for c in df.columns})
    if "standard_name" not in df.columns:
        raise ValueError("엑셀에 '표준제품명' 컬럼이 필요합니다.")
    for c in ["product_code", "aliases", "erp_nohtuspharm_name", "erp_nohtus_name", "erp_noh_name", "erp_noh_code", "bidata_name"]:
        if c not in df.columns:
            df[c] = ""
    inserted = 0
    skipped = 0
    seen = set()
    rows_to_insert = []
    for _, r in df.iterrows():
        code = "" if pd.isna(r.get("product_code")) else str(r.get("product_code")).strip()
        name = "" if pd.isna(r.get("standard_name")) else str(r.get("standard_name")).strip()
        aliases = "" if pd.isna(r.get("aliases")) else str(r.get("aliases")).strip()
        erp_np = "" if pd.isna(r.get("erp_nohtuspharm_name")) else str(r.get("erp_nohtuspharm_name")).strip()
        erp_nt = "" if pd.isna(r.get("erp_nohtus_name")) else str(r.get("erp_nohtus_name")).strip()
        erp_noh = "" if pd.isna(r.get("erp_noh_name")) else str(r.get("erp_noh_name")).strip()
        erp_noh_code = "" if pd.isna(r.get("erp_noh_code")) else str(r.get("erp_noh_code")).strip()
        bidata_name = "" if pd.isna(r.get("bidata_name")) else str(r.get("bidata_name")).strip()
        if not name:
            skipped += 1
            continue
        key = name.strip()
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        rows_to_insert.append((code, name, name, aliases, erp_np, erp_nt, erp_noh, erp_noh_code, bidata_name))
    with connect() as con:
        cur = con.cursor()
        cur.execute("DELETE FROM products")
        for row in rows_to_insert:
            cur.execute("""INSERT INTO products(product_code,standard_name,warehouse_name,aliases,erp_nohtuspharm_name,erp_nohtus_name,erp_noh_name,erp_noh_code,bidata_name)
                           VALUES(?,?,?,?,?,?,?,?,?)""", row)
            inserted += 1
        con.commit()
    return 0, inserted, skipped


def add_product_mapping_record(standard_name, erp_np="", np_code="", erp_noh="", noh_code="", erp_nt="", bidata_name="", aliases=""):
    """제품 매칭 관리 화면에서 제품매칭표 행을 직접 추가한다."""
    standard_name = (standard_name or "").strip()
    if not standard_name:
        raise ValueError("표준제품명은 반드시 입력해야 합니다.")
    with connect() as con:
        cur = con.cursor()
        exists = cur.execute("SELECT id FROM products WHERE TRIM(standard_name)=?", (standard_name,)).fetchone()
        if exists:
            raise ValueError("이미 같은 표준제품명이 제품매칭표에 있습니다. 기존 행을 수정하세요.")
        cur.execute("""INSERT INTO products(
            product_code, standard_name, warehouse_name, aliases,
            erp_nohtuspharm_name, erp_nohtus_name, erp_noh_name, erp_noh_code, bidata_name
        ) VALUES(?,?,?,?,?,?,?,?,?)""", (
            str(np_code or "").strip(), standard_name, standard_name, str(aliases or "").strip(),
            str(erp_np or "").strip(), str(erp_nt or "").strip(), str(erp_noh or "").strip(),
            str(noh_code or "").strip(), str(bidata_name or "").strip()
        ))
        con.commit()

def get_erp_column(company):
    if company == "노투스팜":
        return "erp_nohtuspharm_name"
    if company == "NOH":
        return "erp_noh_name"
    if company == "노투스":
        return "erp_nohtus_name"
    return None


def match_erp_name(company, erp_name):
    """ERP명칭을 표준제품명으로 변환한다. 1건이면 자동, 여러 후보면 확인 필요."""
    erp_name = (erp_name or "").strip()
    if not erp_name:
        return {"status": "fail", "candidates": [], "message": "ERP명칭이 비어 있습니다."}
    col = get_erp_column(company)
    candidates = []
    if col:
        df = q(f"SELECT standard_name FROM products WHERE TRIM(COALESCE({col}, '')) = ?", (erp_name,))
        candidates.extend(df["standard_name"].dropna().astype(str).tolist())
    cand_df = q("SELECT candidate_product FROM erp_ambiguous_candidates WHERE erp_company=? AND erp_name=? ORDER BY candidate_product", (company, erp_name))
    candidates.extend(cand_df["candidate_product"].dropna().astype(str).tolist())
    candidates = sorted(set([c for c in candidates if c]))
    if len(candidates) == 1:
        return {"status": "auto", "candidates": candidates, "message": f"자동 매칭: {candidates[0]}"}
    if len(candidates) > 1:
        return {"status": "ambiguous", "candidates": candidates, "message": "확인 필요 품목입니다."}
    return {"status": "fail", "candidates": [], "message": "매칭 실패: 제품마스터 또는 확인필요 후보에 등록하세요."}




def ensure_standard_product(name):
    """표준제품명이 products DB에 없으면 새로 등록하고, 최종 표준제품명을 반환한다."""
    name = (name or "").strip()
    if not name:
        return ""
    with connect() as con:
        cur = con.cursor()
        existing = cur.execute("SELECT standard_name FROM products WHERE TRIM(standard_name)=?", (name,)).fetchone()
        if existing:
            return str(existing[0])
        cur.execute("""
            INSERT INTO products(
                product_code, standard_name, warehouse_name, aliases,
                erp_nohtuspharm_name, erp_nohtus_name, erp_noh_name, erp_noh_code, bidata_name
            ) VALUES(?,?,?,?,?,?,?,?,?)
        """, ("", name, name, "", "", "", "", "", ""))
        con.commit()
    return name


def ensure_standard_product_only(name):
    """최초 입고 등록용: 표준제품명만 products DB에 추가한다.
    ERP명/비자료명/별칭은 비워 두어 제품 매칭 관리에서 보완할 수 있게 한다.
    """
    name = (name or "").strip()
    if not name:
        return ""
    with connect() as con:
        cur = con.cursor()
        existing = cur.execute("SELECT standard_name FROM products WHERE TRIM(standard_name)=?", (name,)).fetchone()
        if existing:
            return str(existing[0])
        cur.execute("""
            INSERT INTO products(
                product_code, standard_name, warehouse_name, aliases,
                erp_nohtuspharm_name, erp_nohtus_name, erp_noh_name, erp_noh_code, bidata_name
            ) VALUES(?,?,?,?,?,?,?,?,?)
        """, ("", name, "", "", "", "", "", "", ""))
        con.commit()
    return name


def apply_standard_name_change(old_name, new_name):
    """제품 매칭표에서 표준제품명이 바뀌면 이미 올라간 재고/이력에도 즉시 반영한다."""
    old_name = (old_name or "").strip()
    new_name = (new_name or "").strip()
    if not old_name or not new_name or old_name == new_name:
        return
    with connect() as con:
        cur = con.cursor()
        for table, col in [
            ("inventory", "product_name"),
            ("transactions", "product_name"),
            ("outbound_order_items", "product_name"),
            ("erp_ambiguous_candidates", "candidate_product"),
        ]:
            try:
                cur.execute(f"UPDATE {table} SET {col}=? WHERE {col}=?", (new_name, old_name))
            except Exception:
                pass
        con.commit()

def delete_product(product_id):
    """제품 매칭표에서 제품을 삭제한다. 이미 재고/이력/출고지시에 사용된 제품이면 삭제하지 않는다."""
    with connect() as con:
        cur = con.cursor()
        row = cur.execute("SELECT standard_name FROM products WHERE id=?", (int(product_id),)).fetchone()
        if not row:
            raise ValueError("삭제할 제품을 찾을 수 없습니다.")
        name = row[0]
        used = 0
        for table, col in [("inventory", "product_name"), ("transactions", "product_name"), ("outbound_order_items", "product_name")]:
            try:
                cnt = cur.execute(f"SELECT COUNT(*) FROM {table} WHERE {col}=?", (name,)).fetchone()[0]
                used += int(cnt or 0)
            except Exception:
                pass
        if used > 0:
            raise ValueError(f"이미 재고/이력/출고지시에 사용된 제품이라 삭제할 수 없습니다. 사용 건수: {used}건")
        cur.execute("DELETE FROM erp_ambiguous_candidates WHERE candidate_product=?", (name,))
        cur.execute("DELETE FROM products WHERE id=?", (int(product_id),))
        con.commit()


def outbound_erp_note_for_row(row, cache=None):
    """출고지시서 비고용 ERP명/비자료명.
    매출등록은 ERP명 기준으로 하므로, 사업장+표준제품명에 맞는 제품매칭표 값을 표시한다.
    매칭값이 없으면 빈칸으로 둔다.
    """
    cache = cache if cache is not None else {}
    company = str(row.get("사업장") or row.get("company") or row.get("사업체") or "").strip()
    product = str(row.get("제품명") or row.get("product_name") or "").strip()
    if not company or not product:
        return ""
    key = (company, product)
    if key in cache:
        return cache[key]
    try:
        val = product_mapping_name_for(company, product) or ""
    except Exception:
        val = ""
    cache[key] = val
    return val

def outbound_excel_bytes(rows, title="출고지시서"):
    rows = sort_outbound_rows_for_picking(rows)
    note_cache = {}
    rows = [{**(r or {}), "비고": outbound_erp_note_for_row(r or {}, note_cache)} for r in rows]
    df = pd.DataFrame(rows)
    cols = ["로케이션", "제품명", "LOT", "유통기한", "요청수량", "비고"]
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    out = df[cols].copy()
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="출고지시서", startrow=2)
        ws = writer.book["출고지시서"]
        ws["A1"] = title or "출고지시서"
        ws.merge_cells("A1:F1")
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")
        widths = {"A":18,"B":30,"C":18,"D":18,"E":12,"F":32}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows(min_row=3, max_row=3 + len(out), min_col=1, max_col=6):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", horizontal="center" if cell.column in [1,3,4,5] else "left", wrap_text=True)
                if cell.row == 3:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
        ws.auto_filter.ref = f"A3:F{3+len(out)}"
    bio.seek(0)
    return bio.getvalue()

def _find_korean_font():
    candidates = [
        r"C:\Windows\Fonts\malgun.ttf",
        r"C:\Windows\Fonts\malgunbd.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return str(path)
    return None

def outbound_pdf_bytes(rows, title="출고지시서"):
    rows = sort_outbound_rows_for_picking(rows)
    note_cache = {}
    rows = [{**(r or {}), "비고": outbound_erp_note_for_row(r or {}, note_cache)} for r in rows]
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    bio = BytesIO()
    font_name = "Helvetica"
    font_path = _find_korean_font()
    if font_path:
        try:
            pdfmetrics.registerFont(TTFont("NOHTUS_KR", font_path))
            font_name = "NOHTUS_KR"
        except Exception:
            font_name = "Helvetica"
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4), leftMargin=22, rightMargin=22, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = font_name
    styles["Normal"].fontName = font_name
    story = [Paragraph(title or "출고지시서", styles["Title"]), Spacer(1, 12)]
    cols = ["로케이션", "제품명", "LOT", "유통기한", "요청수량", "비고"]
    data = [cols]
    for r in rows:
        data.append([str(r.get(c, "")) for c in cols])
    table = Table(data, colWidths=[80, 220, 105, 105, 70, 160], repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), font_name),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E5E7EB")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#111827")),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
        ("ALIGN", (0,1), (0,-1), "CENTER"),
        ("ALIGN", (2,1), (4,-1), "CENTER"),
        ("ALIGN", (1,1), (1,-1), "LEFT"),
        ("ALIGN", (5,1), (5,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(table)
    doc.build(story)
    bio.seek(0)
    return bio.getvalue()

def full_inventory_excel_bytes(exclude_zero=True):
    where_sql = "WHERE qty > 0" if exclude_zero else ""
    df = q(f"""
        SELECT location, product_name, warehouse_name, lot, exp_date, qty
        FROM inventory
        {where_sql}
        ORDER BY location, product_name, lot, exp_date
    """)
    out = pd.DataFrame()
    # 재고 실사용 엑셀은 현장에서 바로 입력하기 쉬운 최소 컬럼만 제공한다.
    out["로케이션"] = df["location"] if not df.empty else []
    out["제품명(표준제품명)"] = df["product_name"] if not df.empty else []
    out["제조번호"] = df["lot"] if not df.empty else []
    out["유통기한"] = df["exp_date"].apply(display_date_only) if not df.empty else []
    out["전산수량"] = df["qty"] if not df.empty else []
    out["실물수량"] = ""
    from io import BytesIO
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="전체재고실사")
        ws = writer.book["전체재고실사"]
        widths = {"A":16,"B":30,"C":18,"D":16,"E":12,"F":12}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
    bio.seek(0)
    return bio.getvalue()


def baseline_stock_template_excel_bytes():
    """기준재고 업로드용 샘플 양식.
    사용자가 ERP/비자료 내용을 가공해 만든 초기 재고자료를 WMS DB에 넣기 위한 최소 양식이다.
    표준제품명은 비워서 올려도 제품매칭표 기준으로 자동 보완한다.
    """
    sample = pd.DataFrame([
        {
            "사업장": "노투스팜",
            "ERP제품코드": "003",
            "ERP제품명": "JS Tox 100U",
            "표준제품명": "",
            "LOT/제조번호": "NF20CL0901",
            "유통기한": "2027-09-30",
            "로케이션": "A1-01-01",
            "수량": 100,
        },
        {
            "사업장": "비자료",
            "ERP제품코드": "",
            "ERP제품명": "홍보 브로슈어",
            "표준제품명": "",
            "LOT/제조번호": "-",
            "유통기한": "-",
            "로케이션": "홍보물랙",
            "수량": 20,
        },
    ])
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        sample.to_excel(writer, index=False, sheet_name="기준재고업로드")
        ws = writer.book["기준재고업로드"]
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        optional_fill = PatternFill("solid", fgColor="EEF2FF")
        widths = {"A":14,"B":18,"C":34,"D":30,"E":18,"F":16,"G":18,"H":10}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:H{len(sample)+1}"
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if cell.column_letter == "B":
                    cell.number_format = "@"
                    if cell.value is not None:
                        cell.value = str(cell.value)
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = optional_fill if cell.value == "표준제품명" else header_fill
    bio.seek(0)
    return bio.getvalue()


def _baseline_get_product_raw(row):
    return first_nonblank(
        row.get("ERP제품명"), row.get("제품명"), row.get("비자료명"),
        row.get("노투스팜 ERP명"), row.get("NOH ERP명"), row.get("노투스 ERP명")
    )

def _baseline_match_standard(company, product_raw):
    company = (company or "").strip()
    product_raw = (product_raw or "").strip()
    if not company or not product_raw:
        return ""
    if company in ["노투스팜", "NOH", "노투스"]:
        m = match_erp_name(company, product_raw)
        if m.get("status") == "auto" and m.get("candidates"):
            return m["candidates"][0]
        return ""
    if company == "비자료":
        df = q("SELECT standard_name FROM products WHERE TRIM(COALESCE(bidata_name, '')) = ?", (product_raw,))
        if len(df) == 1:
            return str(df.iloc[0]["standard_name"] or "")
        if df.empty:
            same = q("SELECT standard_name FROM products WHERE TRIM(standard_name)=?", (product_raw,))
            if len(same) == 1:
                return str(same.iloc[0]["standard_name"] or "")
    return ""

def _baseline_mapping_payload(company, code, product_raw, standard_name):
    payload = {
        "표준제품명": standard_name or "",
        "노투스팜 ERP명": "",
        "노투스팜 ERP 제품코드": "",
        "NOH ERP명": "",
        "NOH ERP 제품코드": "",
        "노투스 ERP명": "",
        "비자료명": "",
    }
    if company == "노투스팜":
        payload["노투스팜 ERP명"] = product_raw or ""
        payload["노투스팜 ERP 제품코드"] = code or ""
    elif company == "NOH":
        payload["NOH ERP명"] = product_raw or ""
        payload["NOH ERP 제품코드"] = code or ""
    elif company == "노투스":
        payload["노투스 ERP명"] = product_raw or ""
    elif company == "비자료":
        payload["비자료명"] = product_raw or ""
    return payload

def prepare_baseline_stock_dataframe(uploaded_file):
    """기준재고 파일을 제품매칭표 기준으로 정제한다.
    별도 검증으로 막지 않고, 제품매칭표에 따라 표준제품명을 자동 보완한다.
    표준제품명이 직접 입력되어 있으면 절대 덮어쓰지 않는다.
    """
    df = pd.read_excel(uploaded_file, dtype=str).fillna("")
    col_alias = {
        "구분": "사업장",
        "ERP제품코드": "ERP제품코드",
        "ERP 제품코드": "ERP제품코드",
        "전산제품코드": "ERP제품코드",
        "제품코드": "ERP제품코드",
        "노투스팜 ERP 제품코드": "ERP제품코드",
        "NOH ERP 제품코드": "ERP제품코드",
        "ERP상제품명": "ERP제품명",
        "ERP제품명": "ERP제품명",
        "제품명": "ERP제품명",
        "전산상명칭": "ERP제품명",
        "전산상 명칭": "ERP제품명",
        "전산상제품명": "ERP제품명",
        "비자료명": "비자료명",
        "LOT": "LOT/제조번호",
        "제조번호": "LOT/제조번호",
        "수량": "수량",
        "기준수량": "수량",
        "현재재고": "수량",
        "실재고": "수량",
    }
    df = df.rename(columns={c: col_alias.get(c, c) for c in df.columns})
    for c in ["사업장", "ERP제품코드", "ERP제품명", "비자료명", "표준제품명", "LOT/제조번호", "유통기한", "로케이션", "수량"]:
        if c not in df.columns:
            df[c] = ""

    rows = []
    for _, r in df.iterrows():
        company = first_nonblank(r.get("사업장"))
        code = first_nonblank(r.get("ERP제품코드"), r.get("노투스팜 ERP 제품코드"), r.get("NOH ERP 제품코드"))
        product_raw = _baseline_get_product_raw(r)
        standard = first_nonblank(r.get("표준제품명"), r.get("WMS표준제품명"), r.get("실제제품명"), r.get("실제품명"))
        if not standard:
            standard = _baseline_match_standard(company, product_raw)
        if not standard:
            standard = product_raw
        lot = first_nonblank(r.get("LOT/제조번호")) or "-"
        exp_raw = first_nonblank(r.get("유통기한")) or "-"
        loc = first_nonblank(r.get("로케이션")) or "-"
        qty_text = first_nonblank(r.get("수량"))
        try:
            qty = int(float(str(qty_text).replace(",", "")))
        except Exception:
            qty = 0
        if not company or company not in COMPANIES or not standard or qty <= 0:
            # 검증 화면은 없애되, DB에 넣을 수 없는 핵심 누락행은 조용히 제외한다.
            continue
        rows.append({
            "사업장": company,
            "ERP제품코드": code,
            "ERP제품명": product_raw,
            "표준제품명": standard,
            "LOT/제조번호": lot,
            "유통기한": _excel_date_to_iso(exp_raw),
            "로케이션": loc,
            "수량": qty,
        })
    normal_df = pd.DataFrame(rows, columns=["사업장","ERP제품코드","ERP제품명","표준제품명","LOT/제조번호","유통기한","로케이션","수량"])
    issue_df = pd.DataFrame(columns=["보완사유", "사업장", "ERP제품코드", "ERP제품명", "표준제품명", "LOT/제조번호", "유통기한", "로케이션", "수량"])
    return normal_df, issue_df

def baseline_stock_supplement_excel_bytes(issue_df):
    """보완이 필요한 기준재고 행만 내려받기 위한 엑셀 파일."""
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        out = issue_df.copy() if issue_df is not None else pd.DataFrame()
        out.to_excel(writer, index=False, sheet_name="보완필요")
        ws = writer.book["보완필요"]
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        need_fill = PatternFill("solid", fgColor="FFF2CC")
        widths = {
            "A":34,"B":14,"C":18,"D":34,"E":30,"F":34,"G":24,"H":34,"I":22,
            "J":34,"K":34,"L":18,"M":16,"N":18,"O":10
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        max_row = max(1, len(out) + 1)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:O{max_row}"
        required_headers = {"사업장", "ERP제품명", "표준제품명", "LOT/제조번호", "유통기한", "로케이션", "수량"}
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                elif ws.cell(row=cell.row, column=1).value and (cell.value is None or str(cell.value).strip() == ""):
                    header = ws.cell(row=1, column=cell.column).value
                    if header in required_headers or header in ["노투스팜 ERP명", "NOH ERP명", "노투스 ERP명", "비자료명"]:
                        cell.fill = need_fill
                if cell.column_letter in ["C", "G", "I"]:
                    cell.number_format = "@"
    bio.seek(0)
    return bio.getvalue()


def _excel_date_to_iso(v):
    """엑셀 날짜값/문자열을 YYYY-MM-DD로 정규화한다."""
    if pd.isna(v):
        return "-"
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    # Excel serial date
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            d = pd.to_datetime(v, unit="D", origin="1899-12-30")
            return d.strftime("%Y-%m-%d")
        except Exception:
            pass
    text = str(v).strip()
    if not text or text.lower() == "nan":
        return "-"
    if text == "-":
        return "-"
    return normalize_exp_date(text)



def display_date_only(v):
    """화면 표시용: YYYY-MM-DD만 남긴다. 빈값은 '-'로 표시."""
    if v is None:
        return "-"
    text = str(v).strip()
    if not text or text.lower() == "nan" or text == "-":
        return "-"
    dt = pd.to_datetime(text, errors="coerce")
    if pd.notna(dt):
        return dt.strftime("%Y-%m-%d")
    return normalize_exp_date(text)

def first_nonblank(*values):
    for v in values:
        if v is None:
            continue
        text = str(v).strip()
        if text and text.lower() != "nan" and text != "-":
            return text
    return ""

def product_mapping_name_for(company, standard_name):
    if not standard_name:
        return ""
    col = {
        "노투스팜": "erp_nohtuspharm_name",
        "NOH": "erp_noh_name",
        "노투스": "erp_nohtus_name",
        "비자료": "bidata_name",
    }.get(company)
    if not col:
        return ""
    df = q(f"SELECT {col} AS nm FROM products WHERE standard_name=?", (standard_name,))
    if df.empty:
        return ""
    return first_nonblank(df.iloc[0].get("nm"))


def product_compare_name_for(company, standard_name):
    """ERP 비교용 제품명.
    제품매칭표에 해당 사업장의 ERP명이 있으면 ERP명 기준으로 비교하고,
    없으면 표준제품명 기준으로 비교한다.
    """
    return product_mapping_name_for(company, standard_name) or (standard_name or "")


def _row_override_standard_name(row):
    """재고 업로드 파일에 표준제품명/실제제품명 컬럼이 있으면 그 값을 우선 사용한다.
    모든 품목을 매핑할 필요는 없고, JS Tox 100U처럼 ERP명만으로 실제 제품을 알 수 없는 행에만 넣으면 된다.
    """
    for col in ["표준제품명", "WMS표준제품명", "실제제품명", "실제품명"]:
        if col in row.index and not pd.isna(row.get(col)):
            v = str(row.get(col)).strip()
            if v and v.lower() != "nan":
                return v
    return ""


def import_stock_survey_excel(uploaded_file, replace_current=True):
    """기준재고 엑셀을 현재 WMS 재고로 불러온다.

    기준재고 파일은 사용자가 ERP/비자료 내용을 가공한 초기 DB 투입용 자료다.
    업로드 시 제품매칭표를 기준으로 표준제품명을 자동 보완하고,
    필수값 누락 또는 매칭 실패 행은 DB에 반영하지 않는다.
    """
    normal_df, issue_df = prepare_baseline_stock_dataframe(uploaded_file)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    inserted = 0
    skipped = int(len(issue_df)) if issue_df is not None else 0
    product_inserted = 0
    with connect() as con:
        cur = con.cursor()
        if replace_current:
            cur.execute("DELETE FROM inventory")
            cur.execute("DELETE FROM transactions WHERE tx_type='재고조사불러오기'")
        for _, r in normal_df.iterrows():
            company = str(r.get("사업장") or "").strip()
            code = str(r.get("ERP제품코드") or "").strip()
            product_raw = str(r.get("ERP제품명") or "").strip()
            product = str(r.get("표준제품명") or "").strip()
            lot = str(r.get("LOT/제조번호") or "").strip() or "-"
            exp = _excel_date_to_iso(r.get("유통기한"))
            loc = str(r.get("로케이션") or "").strip()
            qty = int(float(r.get("수량") or 0))
            if not company or not product or not loc or qty <= 0:
                skipped += 1
                continue

            exists = cur.execute("SELECT id FROM products WHERE standard_name=?", (product,)).fetchone()
            if not exists:
                cur.execute("""INSERT INTO products(product_code, standard_name, warehouse_name, aliases, erp_nohtuspharm_name, erp_noh_name, erp_noh_code, erp_nohtus_name, bidata_name)
                               VALUES(?,?,?,?,?,?,?,?,?)""", (
                    code if company == "노투스팜" else "",
                    product,
                    product_raw,
                    "",
                    product_raw if company == "노투스팜" else "",
                    product_raw if company == "NOH" else "",
                    code if company == "NOH" else "",
                    product_raw if company == "노투스" else "",
                    product_raw if company == "비자료" else "",
                ))
                product_inserted += 1
            else:
                pid = int(exists[0])
                if company == "노투스팜":
                    cur.execute("UPDATE products SET erp_nohtuspharm_name=COALESCE(NULLIF(erp_nohtuspharm_name,''), ?), product_code=COALESCE(NULLIF(product_code,''), ?) WHERE id=?", (product_raw, code, pid))
                elif company == "NOH":
                    cur.execute("UPDATE products SET erp_noh_name=COALESCE(NULLIF(erp_noh_name,''), ?), erp_noh_code=COALESCE(NULLIF(erp_noh_code,''), ?) WHERE id=?", (product_raw, code, pid))
                elif company == "노투스":
                    cur.execute("UPDATE products SET erp_nohtus_name=COALESCE(NULLIF(erp_nohtus_name,''), ?) WHERE id=?", (product_raw, pid))
                elif company == "비자료":
                    cur.execute("UPDATE products SET bidata_name=COALESCE(NULLIF(bidata_name,''), ?) WHERE id=?", (product_raw, pid))

            cur.execute("""INSERT INTO inventory(company, product_name, warehouse_name, lot, exp_date, location, qty, updated_at)
                           VALUES(?,?,?,?,?,?,?,?)""", (company, product, product_raw, lot, exp, loc, qty, now))
            insert_transaction_log(cur, created_at=now, tx_type="재고조사불러오기", product_name=product, warehouse_name=product_raw,
                                   lot=lot, exp_date=exp, from_company=None, from_location=None,
                                   to_company=company, to_location=loc, qty=qty, memo=f"기준재고 엑셀 업로드 / 원본명: {product_raw}")
            inserted += 1
        con.commit()
    return inserted, skipped, product_inserted, skipped

def make_location(area, line=None, level=None):
    area = area or ""; line = line or ""; level = level or ""
    # 기타 위치는 DB에 "기타 위치-홍보물랙" 형태로 저장하지 않고 실제 위치명만 저장한다.
    if area == "N" and line in SPECIAL_LOCATIONS:
        return line
    if area == "Q" and line in ["Q1", "Q2"]:
        return line
    if line and level:
        return f"{area}-{line}-{level}"
    if line:
        return f"{area}-{line}"
    return area

# ---------------- common lookup ----------------
def parse_location(loc):
    loc = (loc or "").strip()
    if loc in SPECIAL_LOCATIONS:
        return "N", loc, ""
    parts = loc.split("-")
    area = parts[0] if parts else ""
    line = parts[1] if len(parts) >= 2 else ""
    level = parts[2] if len(parts) >= 3 else ""
    return area, line, level


def _location_picking_key(loc):
    """REC에서 시작하는 피킹 동선용 정렬 키.
    실제 거리 계산 대신 창고 구역 순서 + 라인 + 단 기준으로 안정적으로 정렬한다.
    """
    loc = (loc or "").strip()
    area, line, level = parse_location(loc)
    area_order = [
        "REC", "A1", "A2", "B1", "B2", "C1", "C2", "D1", "E1", "F1",
        "G1", "G2", "X1", "X2", "Q", "N", "T1", "T2", "P", "R1", "R2"
    ]
    try:
        area_idx = area_order.index(area)
    except ValueError:
        area_idx = 999

    def _num(v):
        text = str(v or "").strip()
        m = re.search(r"\d+", text)
        return int(m.group()) if m else 999

    special_idx = SPECIAL_LOCATIONS.index(line) if area == "N" and line in SPECIAL_LOCATIONS else 999
    return (area_idx, special_idx, _num(line), _num(level), loc)


def sort_outbound_rows_for_picking(rows):
    """출고지시서 출력용 피킹 순서 정렬. 화면 장바구니 순서는 건드리지 않는다."""
    if not rows:
        return rows
    copied = [dict(r) for r in rows]
    return sorted(copied, key=lambda r: _location_picking_key(r.get("로케이션") or r.get("location") or ""))

def has_stock_map():
    df = q("SELECT location, SUM(qty) qty FROM inventory WHERE qty>0 GROUP BY location")
    return {r.location: int(r.qty) for r in df.itertuples()}

def loc_has_stock(loc, stock=None):
    stock = stock or has_stock_map()
    return loc in stock or any(k.startswith(loc + "-") for k in stock)

def product_options(term=""):
    term = (term or "").strip().lower()
    df = q("""SELECT standard_name, warehouse_name, aliases,
                    erp_nohtuspharm_name, erp_nohtus_name, erp_noh_name, bidata_name
             FROM products ORDER BY standard_name""")
    if term:
        search_cols = ["standard_name", "warehouse_name", "aliases", "erp_nohtuspharm_name", "erp_nohtus_name", "erp_noh_name", "bidata_name"]
        mask = df.apply(lambda r: any(term in str(r.get(c, "")).lower() for c in search_cols), axis=1)
        df = df[mask]
    return df

def warehouse_name_options(term=""):
    term = (term or "").strip().lower()
    df = q("""
        SELECT standard_name AS name FROM products
        UNION
        SELECT warehouse_name AS name FROM products WHERE IFNULL(warehouse_name,'')<>''
        UNION
        SELECT product_name AS name FROM inventory
        UNION
        SELECT warehouse_name AS name FROM inventory WHERE IFNULL(warehouse_name,'')<>''
        ORDER BY name
    """)
    df = df.dropna()
    df = df[df["name"].astype(str).str.strip() != ""]
    if term:
        df = df[df["name"].astype(str).str.lower().str.contains(term, regex=False)]
    return df["name"].drop_duplicates().tolist()

def location_picker(prefix, default_area="A1", stock_only=False):
    """구역/라인/단 선택.
    입고 등록과 같은 조합 규칙을 사용한다.
    - 라인/단이 있는 구역은 선택 없음 없이 실제 값만 표시한다.
    - Q 구역은 라인/단 선택 없이 Q로 고정한다.
    - 라인/단이 없는 특수 구역만 비활성 선택 없음으로 표시한다.
    """
    picker_defaults = st.session_state.get(f"_{prefix}_picker_defaults", {}) or {}
    widget_suffix = ""
    if picker_defaults:
        # 외부 도면 클릭으로 key 재생성이 필요한 입고 화면에만 토큰 suffix를 붙인다.
        # 이동 등록은 사용자가 콤보박스에서 직접 타이핑/선택한 값을 안정적으로 유지하기 위해 고정 key를 사용한다.
        if prefix == "inbound":
            widget_suffix = f"_{st.session_state.get(f'_{prefix}_picker_token', 0)}"
        default_area = picker_defaults.get("area") or default_area

    if stock_only:
        stock_df = q("SELECT DISTINCT location FROM inventory WHERE qty>0 ORDER BY location")
        locs = stock_df["location"].tolist()
        areas = sorted({parse_location(x)[0] for x in locs}) or ["A1"]
    else:
        locs = []
        areas = list(AREA_CONFIG.keys())
    if default_area not in areas:
        default_area = areas[0]

    c1, c2, c3 = st.columns(3)
    with c1:
        area = st.selectbox("구역", areas, index=areas.index(default_area), key=f"{prefix}_area{widget_suffix}")

    cfg = AREA_CONFIG.get(area, {"lines": [], "levels": []})
    if area == "Q":
        lines = []
        levels = []
    elif stock_only:
        lines = sorted({parse_location(x)[1] for x in locs if parse_location(x)[0] == area and parse_location(x)[1]})
        levels = []
    else:
        lines = list(cfg.get("lines", []))
        levels = list(cfg.get("levels", []))

    default_line = str(picker_defaults.get("line", "") or "") if picker_defaults else ""
    with c2:
        if lines:
            if default_line not in lines:
                default_line = lines[0]
            line = st.selectbox("라인", lines, index=lines.index(default_line), key=f"{prefix}_line{widget_suffix}")
        else:
            st.selectbox("라인", ["선택 없음"], key=f"{prefix}_line_disabled{widget_suffix}", disabled=True)
            line = ""

    if area != "Q" and stock_only:
        if line:
            levels = sorted({parse_location(x)[2] for x in locs if parse_location(x)[0] == area and parse_location(x)[1] == line and parse_location(x)[2]})
        else:
            levels = sorted({parse_location(x)[2] for x in locs if parse_location(x)[0] == area and parse_location(x)[2]})

    default_level = str(picker_defaults.get("level", "") or "") if picker_defaults else ""
    with c3:
        if levels:
            if default_level not in levels:
                default_level = levels[0]
            level = st.selectbox("단", levels, index=levels.index(default_level), key=f"{prefix}_level{widget_suffix}")
        else:
            st.selectbox("단", ["선택 없음"], key=f"{prefix}_level_disabled{widget_suffix}", disabled=True)
            level = ""

    if prefix == "inbound":
        st.session_state["_inbound_selected_loc"] = make_location(area, line, level)
    st.session_state[f"_{prefix}_picker_defaults"] = {"area": area, "line": line, "level": level}
    return make_location(area, line, level)

def inbound_location_picker(default_area="REC"):
    """입고 등록 전용 위치 선택기.
    도면 클릭값은 _inbound_picker_defaults/_inbound_picker_token으로 받아서
    기존 Streamlit widget key를 직접 수정하지 않고 다음 렌더에서 콤보박스 값을 동기화한다.

    운영 규칙:
    - 라인/단이 있는 구역은 "선택 없음"을 표시하지 않는다.
    - 라인만 있는 구역(Q 등)은 라인만 선택한다.
    - 라인/단이 없는 구역(REC/P/R1/R2/N 등)에만 선택 없음/비활성 표시가 나온다.
    """
    defaults = st.session_state.get("_inbound_picker_defaults", {}) or {}
    area_default = str(defaults.get("area") or default_area or "REC")
    line_default = str(defaults.get("line") or "")
    level_default = str(defaults.get("level") or "")
    token = int(st.session_state.get("_inbound_picker_token", 0) or 0)

    areas = list(AREA_CONFIG.keys())
    if area_default not in areas:
        area_default = default_area if default_area in areas else areas[0]

    c1, c2, c3 = st.columns(3)
    with c1:
        area = st.selectbox("구역", areas, index=areas.index(area_default), key=f"inbound_area_{token}")

    cfg = AREA_CONFIG.get(area, {"lines": [], "levels": []})
    if area == "Q":
        lines = []
    else:
        lines = list(cfg.get("lines", []))
    with c2:
        if lines:
            if line_default not in lines:
                line_default = lines[0]
            line = st.selectbox("라인", lines, index=lines.index(line_default), key=f"inbound_line_{token}")
        else:
            st.selectbox("라인", ["선택 없음"], key=f"inbound_line_disabled_{token}", disabled=True)
            line = ""

    levels = [] if area == "Q" else list(cfg.get("levels", []))
    with c3:
        if levels:
            if level_default not in levels:
                level_default = levels[0]
            level = st.selectbox("단", levels, index=levels.index(level_default), key=f"inbound_level_{token}")
        else:
            st.selectbox("단", ["선택 없음"], key=f"inbound_level_disabled_{token}", disabled=True)
            level = ""

    loc = make_location(area, line, level)
    st.session_state["_inbound_selected_loc"] = loc
    st.session_state["_inbound_picker_defaults"] = {"area": area, "line": line, "level": level}
    return loc

def set_loc(loc):
    st.session_state["selected_loc"] = loc
    st.session_state.pop("selected_product_for_detail", None)

def get_loc():
    try:
        loc = st.query_params.get("loc", "")
        if isinstance(loc, list):
            loc = loc[0] if loc else ""
        if loc:
            st.session_state["selected_loc"] = loc
            return loc
    except Exception:
        pass
    return st.session_state.get("selected_loc", "")


# ---------------- inventory ledger helpers ----------------
def product_total_stock(cur, product_name):
    """현재 inventory 기준 표준제품명 전체 총재고.
    사업장/LOT/유통기한/로케이션을 모두 무시하고 product_name만 기준으로 합산한다.
    거래 이력의 final_stock은 이 함수의 값을 작업 직후 스냅샷으로 저장한다.
    """
    product_name = str(product_name or "").strip()
    if not product_name:
        return 0
    row = cur.execute("SELECT COALESCE(SUM(qty), 0) FROM inventory WHERE product_name=?", (product_name,)).fetchone()
    return int((row[0] if row else 0) or 0)


def insert_transaction_log(cur, *, created_at, tx_type, product_name, warehouse_name=None,
                           lot="-", exp_date="-", from_company=None, from_location=None,
                           to_company=None, to_location=None, qty=0, memo="", final_stock=None):
    """거래 이력을 한 곳에서 기록한다.
    final_stock을 넘기지 않으면 현재 inventory 기준 표준제품명 총재고를 저장한다.
    """
    if final_stock is None:
        final_stock = product_total_stock(cur, product_name)
    cur.execute("""INSERT INTO transactions(created_at,tx_type,product_name,warehouse_name,lot,exp_date,
                   from_company,from_location,to_company,to_location,qty,memo,final_stock)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (created_at, tx_type, product_name, warehouse_name, lot or "-", exp_date or "-",
                 from_company, from_location, to_company, to_location, int(qty or 0), memo, int(final_stock or 0)))


def strip_company_stock_label(label):
    """'노투스팜 (30 EA)'처럼 표시된 사업장 라벨에서 실제 사업장명만 추출한다."""
    text = str(label or "").strip()
    return re.sub(r"\s*\(\s*[-+]?\d+\s*EA\s*\)\s*$", "", text).strip()


def inbound_company_options_for(product_name):
    """입고등록 전용 사업장 선택지.
    등록대기는 입고등록 전용 임시값이므로 수량을 붙이지 않는다.
    """
    product_name = str(product_name or "").strip()
    stock = {c: 0 for c in COMPANIES}
    if product_name:
        df = q("""SELECT company, COALESCE(SUM(qty),0) AS qty
                  FROM inventory
                  WHERE product_name=?
                  GROUP BY company""", (product_name,))
        if not df.empty:
            for r in df.itertuples(index=False):
                c = str(getattr(r, "company", "") or "").strip()
                if c in stock:
                    stock[c] = int(getattr(r, "qty", 0) or 0)
    return [f"{c} ({stock.get(c, 0)} EA)" for c in COMPANIES] + ["등록대기"]

# ---------------- inventory operations ----------------
def add_inventory(company, product, warehouse, lot, exp, location, qty, memo="입고 등록"):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with connect() as con:
        cur = con.cursor()
        row = cur.execute("""SELECT id, qty FROM inventory WHERE company=? AND product_name=? AND IFNULL(warehouse_name,'')=? AND lot=? AND exp_date=? AND location=?""",
                          (company, product, warehouse or "", lot, exp, location)).fetchone()
        if row:
            cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (int(row[1] or 0) + int(qty), now, row[0]))
        else:
            cur.execute("""INSERT INTO inventory(company,product_name,warehouse_name,lot,exp_date,location,qty,updated_at)
                           VALUES(?,?,?,?,?,?,?,?)""", (company, product, warehouse, lot, exp, location, int(qty), now))
        insert_transaction_log(cur, created_at=now, tx_type="입고", product_name=product, warehouse_name=warehouse,
                               lot=lot, exp_date=exp, from_company=None, from_location=None,
                               to_company=company, to_location=location, qty=qty, memo=memo)
        con.commit()

def move_inventory(src_id, to_company, to_location, qty, memo=""):
    """재고 이동. 사업장 이동 시 전산상명칭은 도착 사업장 기준으로 다시 계산한다."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with connect() as con:
        cur = con.cursor()
        src = cur.execute("SELECT * FROM inventory WHERE id=?", (src_id,)).fetchone()
        cols = [d[0] for d in cur.description]
        src = dict(zip(cols, src)) if src else None
        if not src:
            raise ValueError("출발 재고를 찾을 수 없습니다.")
        qty = int(qty)
        if qty <= 0 or qty > int(src["qty"] or 0):
            raise ValueError("이동 수량이 현재 재고보다 많거나 올바르지 않습니다.")

        product_name = src["product_name"]
        old_warehouse = src.get("warehouse_name") or ""
        dest_warehouse = product_mapping_name_for(to_company, product_name) or product_name

        cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (int(src["qty"] or 0)-qty, now, src_id))
        row = cur.execute("""SELECT id, qty FROM inventory WHERE company=? AND product_name=? AND IFNULL(warehouse_name,'')=? AND lot=? AND exp_date=? AND location=?""",
                          (to_company, product_name, dest_warehouse or "", src["lot"], src["exp_date"], to_location)).fetchone()
        if row:
            cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (int(row[1] or 0)+qty, now, row[0]))
        else:
            cur.execute("""INSERT INTO inventory(company,product_name,warehouse_name,lot,exp_date,location,qty,updated_at)
                           VALUES(?,?,?,?,?,?,?,?)""", (to_company, product_name, dest_warehouse, src["lot"], src["exp_date"], to_location, qty, now))
        tx_type = "사업장+위치이동"
        if src["company"] == to_company and src["location"] != to_location: tx_type = "위치이동"
        elif src["company"] != to_company and src["location"] == to_location: tx_type = "사업장이동"
        if to_company == "비자료": tx_type = "비자료전환"

        move_memo = str(memo or "").strip()
        if str(old_warehouse or "").strip() != str(dest_warehouse or "").strip():
            erp_note = f"전산상명칭 변경: {old_warehouse or '-'} → {dest_warehouse or '-'}"
            move_memo = f"{move_memo} / {erp_note}" if move_memo else erp_note

        insert_transaction_log(cur, created_at=now, tx_type=tx_type, product_name=product_name, warehouse_name=dest_warehouse,
                               lot=src["lot"], exp_date=src["exp_date"], from_company=src["company"], from_location=src["location"],
                               to_company=to_company, to_location=to_location, qty=qty, memo=move_memo)
        con.commit()


def update_inventory_metadata(inv_id, new_lot, new_exp, memo=""):
    """기존 재고 행의 제조번호/유통기한을 정정한다. 수량은 변경하지 않는다."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lot2 = normalize_blank(new_lot)
    exp2 = normalize_exp_date(new_exp)
    with connect() as con:
        con.row_factory = sqlite3.Row
        cur = con.cursor()
        src = cur.execute("SELECT * FROM inventory WHERE id=?", (int(inv_id),)).fetchone()
        if not src:
            raise ValueError("수정할 재고를 찾을 수 없습니다.")
        old_lot = src["lot"] or "-"
        old_exp = src["exp_date"] or "-"
        if lot2 == old_lot and exp2 == old_exp:
            raise ValueError("변경된 제조번호/유통기한이 없습니다.")

        target = cur.execute("""
            SELECT id, qty FROM inventory
            WHERE id<>? AND company=? AND product_name=? AND IFNULL(warehouse_name,'')=?
              AND lot=? AND exp_date=? AND location=?
        """, (int(inv_id), src["company"], src["product_name"], src["warehouse_name"] or "", lot2, exp2, src["location"])).fetchone()

        merge_note = ""
        qty = int(src["qty"] or 0)
        if target:
            cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (int(target["qty"] or 0) + qty, now, int(target["id"])))
            cur.execute("DELETE FROM inventory WHERE id=?", (int(inv_id),))
            merge_note = f" / 동일 재고행 #{int(target['id'])}에 수량 합산"
        else:
            cur.execute("UPDATE inventory SET lot=?, exp_date=?, updated_at=? WHERE id=?", (lot2, exp2, now, int(inv_id)))

        reason = f"재고정보수정: LOT {old_lot} → {lot2}, 유통기한 {old_exp} → {exp2}"
        if str(memo or "").strip():
            reason += f" / {str(memo).strip()}"
        reason += merge_note
        insert_transaction_log(cur, created_at=now, tx_type="재고정보수정", product_name=src["product_name"], warehouse_name=src["warehouse_name"],
                               lot=lot2, exp_date=exp2, from_company=src["company"], from_location=src["location"],
                               to_company=src["company"], to_location=src["location"], qty=0, memo=reason)
        con.commit()
        return True

def adjust_inventory(inv_id, actual_qty, reason, memo=""):
    """실사 결과 기준으로 해당 재고 행의 수량을 실제 수량으로 조정한다."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with connect() as con:
        con.row_factory = sqlite3.Row
        cur = con.cursor()
        src = cur.execute("SELECT * FROM inventory WHERE id=?", (inv_id,)).fetchone()
        if not src:
            raise ValueError("조정할 재고를 찾을 수 없습니다.")
        actual_qty = int(actual_qty)
        if actual_qty < 0:
            raise ValueError("실물수량은 0 이상이어야 합니다.")
        before = int(src["qty"])
        diff = actual_qty - before
        cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (actual_qty, now, inv_id))
        reason_memo = reason if not memo else f"{reason} / {memo}"
        insert_transaction_log(cur, created_at=now, tx_type="재고조정", product_name=src["product_name"], warehouse_name=src["warehouse_name"],
                               lot=src["lot"], exp_date=src["exp_date"], from_company=src["company"], from_location=src["location"],
                               to_company=src["company"], to_location=src["location"], qty=diff, memo=reason_memo)
        con.commit()
        return before, actual_qty, diff

def create_outbound_instruction(src_id, qty, memo="출고지시"):
    """출고지시는 피킹 지시서만 남기고 실제 inventory 수량은 차감하지 않는다."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with connect() as con:
        cur = con.cursor()
        src = cur.execute("SELECT * FROM inventory WHERE id=?", (src_id,)).fetchone()
        cols = [d[0] for d in cur.description]
        src = dict(zip(cols, src)) if src else None
        if not src:
            raise ValueError("출고 지시할 재고를 찾을 수 없습니다.")
        if qty <= 0 or qty > src["qty"]:
            raise ValueError("지시 수량이 현재 재고보다 많거나 올바르지 않습니다.")
        cur.execute("""INSERT INTO transactions(created_at,tx_type,product_name,warehouse_name,lot,exp_date,from_company,from_location,to_company,to_location,qty,memo)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (now,"출고지시",src["product_name"],src.get("warehouse_name"),src["lot"],src["exp_date"],src["company"],src["location"],None,None,qty,memo))
        con.commit()

def save_outbound_order(cart, title="", memo=""):
    """장바구니를 출고지시서로 저장한다.
    출고지시 저장 시점에 inventory 현재고를 즉시 차감한다.
    같은 inventory_id가 장바구니에 여러 번 들어와도 합산 검증 후 차감하여 중복 출고를 막는다.
    """
    if not cart:
        raise ValueError("저장할 출고지시 품목이 없습니다.")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    order_date = datetime.now().strftime("%Y-%m-%d")
    valid_cart = [item for item in (cart or []) if int(item.get("요청수량", 0) or 0) > 0]
    if not valid_cart:
        raise ValueError("저장할 출고지시 품목이 없습니다.")

    # 출고지시는 반드시 특정 재고행(id)에 묶여야 한다.
    # 제조번호/유통기한/로케이션이 같은 제품만 정확히 차감하기 위해 id가 없는 행은 저장하지 않는다.
    missing_id = [item for item in valid_cart if not item.get("id")]
    if missing_id:
        names = ", ".join(sorted({str(x.get("제품명") or "-") for x in missing_id}))
        raise ValueError(f"재고ID가 없는 장바구니 행이 있어 출고지시를 저장할 수 없습니다: {names}")

    inv_ids = sorted({int(item.get("id")) for item in valid_cart})
    requested_by_inv = {}
    for item in valid_cart:
        inv_key = int(item.get("id"))
        requested_by_inv[inv_key] = requested_by_inv.get(inv_key, 0) + int(item.get("요청수량", 0) or 0)

    with connect() as con:
        cur = con.cursor()
        placeholders = ",".join(["?"] * len(inv_ids))
        rows = cur.execute(f"SELECT * FROM inventory WHERE id IN ({placeholders})", inv_ids).fetchall()
        cols = [d[0] for d in cur.description]
        inv_map = {int(row[cols.index("id")]): dict(zip(cols, row)) for row in rows}

        missing = [x for x in inv_ids if x not in inv_map]
        if missing:
            raise ValueError(f"현재고 DB에서 찾을 수 없는 재고ID가 있습니다: {missing}")

        # 저장 전에 합산 수량을 먼저 검증한다. 하나라도 부족하면 아무것도 저장/차감하지 않는다.
        for inv_key, req_qty in requested_by_inv.items():
            src = inv_map[inv_key]
            before_qty = int(src.get("qty", 0) or 0)
            if req_qty <= 0:
                raise ValueError("출고 요청 수량이 올바르지 않습니다.")
            if req_qty > before_qty:
                product = src.get("product_name", "-")
                loc = src.get("location", "-")
                lot = src.get("lot", "-")
                exp = display_date_only(src.get("exp_date", "-"))
                raise ValueError(f"{product} / {loc} / {lot} / {exp} 재고가 부족합니다. 현재 {before_qty}EA, 요청 {req_qty}EA")

        cur.execute("INSERT INTO outbound_orders(created_at, order_date, title, status, memo) VALUES(?,?,?,?,?)", (now, order_date, title or f"출고지시 {now}", "저장됨", memo))
        order_id = cur.lastrowid

        # inventory는 재고행별 합산 수량으로 한 번만 차감한다.
        final_by_inv = {}
        for inv_key, req_qty in requested_by_inv.items():
            before_qty = int(inv_map[inv_key].get("qty", 0) or 0)
            final_stock = before_qty - int(req_qty)
            cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (final_stock, now, inv_key))
            final_by_inv[inv_key] = final_stock

        # 지시서 품목과 거래 이력은 장바구니 행 단위로 남긴다.
        running_final = {k: int(inv_map[k].get("qty", 0) or 0) for k in inv_map}
        for item in valid_cart:
            qty = int(item.get("요청수량", 0) or 0)
            inv_key = int(item.get("id"))
            src = inv_map[inv_key]
            company = src.get("company", item.get("사업장", item.get("사업체", "")))
            wh = src.get("warehouse_name", item.get("전산상 명칭", item.get("warehouse_name", "")))
            loc = src.get("location", item.get("로케이션", ""))
            product = src.get("product_name", item.get("제품명", ""))
            lot = src.get("lot", item.get("LOT", "-"))
            exp = src.get("exp_date", item.get("유통기한", "-"))
            cur.execute("""INSERT INTO outbound_order_items(order_id, inventory_id, location, product_name, lot, exp_date, qty, company, warehouse_name)
                           VALUES(?,?,?,?,?,?,?,?,?)""", (order_id, inv_key, loc, product, lot, exp, qty, company, wh))
            insert_transaction_log(cur, created_at=now, tx_type="출고지시", product_name=product, warehouse_name=wh,
                                   lot=lot, exp_date=exp, from_company=company, from_location=loc,
                                   to_company=None, to_location=None, qty=qty, memo=f"출고지시서 #{order_id} / 재고차감")
        con.commit()
        return order_id

def load_outbound_order(order_id):
    df = q("""SELECT inventory_id AS id, location AS 로케이션, product_name AS 제품명, lot AS LOT, exp_date AS 유통기한, qty AS 요청수량, company AS 사업장, warehouse_name AS '전산상 명칭'
              FROM outbound_order_items WHERE order_id=? ORDER BY id""", (order_id,))
    return df.to_dict("records")

def update_outbound_order(order_id, title_or_cart, maybe_cart=None):
    """저장된 출고지시서를 수정한다.
    기존 호출 호환: update_outbound_order(id, cart) 또는 update_outbound_order(id, title, cart).

    운영 기준:
    - 저장된 출고지시는 이미 inventory에서 차감된 상태다.
    - 수정 저장 시 기존 지시 수량을 먼저 원복한 뒤 새 장바구니 수량을 다시 차감한다.
    - 제조번호/유통기한/로케이션까지 같은 inventory_id 기준으로만 처리한다.
    """
    if maybe_cart is None:
        title = None
        cart = title_or_cart
    else:
        title = title_or_cart
        cart = maybe_cart

    order_id = int(order_id)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    valid_cart = [item for item in (cart or []) if int(item.get("요청수량", 0) or 0) > 0]
    if not valid_cart:
        raise ValueError("저장할 출고지시 품목이 없습니다.")
    missing_id = [item for item in valid_cart if not item.get("id")]
    if missing_id:
        names = ", ".join(sorted({str(x.get("제품명") or "-") for x in missing_id}))
        raise ValueError(f"재고ID가 없는 장바구니 행이 있어 출고지시를 수정할 수 없습니다: {names}")

    new_requested_by_inv = {}
    for item in valid_cart:
        inv_key = int(item.get("id"))
        new_requested_by_inv[inv_key] = new_requested_by_inv.get(inv_key, 0) + int(item.get("요청수량", 0) or 0)

    with connect() as con:
        cur = con.cursor()
        order = cur.execute("SELECT id, status FROM outbound_orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            raise ValueError("수정할 출고지시서를 찾을 수 없습니다.")
        if str(order[1] or "") == "취소됨":
            raise ValueError("취소된 출고지시서는 수정할 수 없습니다.")

        old_rows = cur.execute("""SELECT inventory_id, location, product_name, lot, exp_date, qty, company, warehouse_name
                                  FROM outbound_order_items WHERE order_id=? ORDER BY id""", (order_id,)).fetchall()
        old_by_inv = {}
        for inv_id, location, product_name, lot, exp_date, qty, company, warehouse_name in old_rows:
            if inv_id:
                old_by_inv[int(inv_id)] = old_by_inv.get(int(inv_id), 0) + int(qty or 0)

        all_inv_ids = sorted(set(old_by_inv.keys()) | set(new_requested_by_inv.keys()))
        if not all_inv_ids:
            raise ValueError("수정할 재고행을 찾을 수 없습니다.")
        placeholders = ",".join(["?"] * len(all_inv_ids))
        rows = cur.execute(f"SELECT * FROM inventory WHERE id IN ({placeholders})", all_inv_ids).fetchall()
        cols = [d[0] for d in cur.description]
        inv_map = {int(row[cols.index("id")]): dict(zip(cols, row)) for row in rows}

        # 기존 차감분을 원복했다고 가정한 가용수량으로 새 수량을 검증한다.
        for inv_key, new_qty in new_requested_by_inv.items():
            src = inv_map.get(inv_key)
            if not src:
                raise ValueError(f"현재고 DB에서 찾을 수 없는 재고ID가 있습니다: {inv_key}")
            available_after_restore = int(src.get("qty", 0) or 0) + int(old_by_inv.get(inv_key, 0) or 0)
            if new_qty > available_after_restore:
                product = src.get("product_name", "-")
                loc = src.get("location", "-")
                lot = src.get("lot", "-")
                exp = display_date_only(src.get("exp_date", "-"))
                raise ValueError(f"{product} / {loc} / {lot} / {exp} 재고가 부족합니다. 원복 후 가능 {available_after_restore}EA, 요청 {new_qty}EA")

        # 기존 출고지시 차감분 원복
        for inv_key, old_qty in old_by_inv.items():
            src = inv_map.get(inv_key)
            if src:
                restored = int(src.get("qty", 0) or 0) + int(old_qty or 0)
                cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (restored, now, inv_key))
                src["qty"] = restored

        # 기존 품목 삭제 후 새 품목 차감/저장
        cur.execute("DELETE FROM outbound_order_items WHERE order_id=?", (order_id,))
        final_by_inv = {}
        for inv_key, new_qty in new_requested_by_inv.items():
            src = inv_map[inv_key]
            final_stock = int(src.get("qty", 0) or 0) - int(new_qty or 0)
            cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (final_stock, now, inv_key))
            final_by_inv[inv_key] = final_stock

        running_final = {k: int(inv_map[k].get("qty", 0) or 0) for k in inv_map}
        for item in valid_cart:
            qty = int(item.get("요청수량", 0) or 0)
            inv_key = int(item.get("id"))
            src = inv_map[inv_key]
            company = src.get("company", item.get("사업장", ""))
            wh = src.get("warehouse_name", item.get("전산상 명칭", ""))
            loc = src.get("location", item.get("로케이션", ""))
            product = src.get("product_name", item.get("제품명", ""))
            lot = src.get("lot", item.get("LOT", "-"))
            exp = src.get("exp_date", item.get("유통기한", "-"))
            cur.execute("""INSERT INTO outbound_order_items(order_id, inventory_id, location, product_name, lot, exp_date, qty, company, warehouse_name)
                           VALUES(?,?,?,?,?,?,?,?,?)""", (order_id, inv_key, loc, product, lot, exp, qty, company, wh))
            insert_transaction_log(cur, created_at=now, tx_type="출고지시수정", product_name=product, warehouse_name=wh,
                                   lot=lot, exp_date=exp, from_company=company, from_location=loc,
                                   to_company=None, to_location=None, qty=qty, memo=f"출고지시서 #{order_id} 수정 / 재고 재차감")

        if title is not None:
            cur.execute("UPDATE outbound_orders SET title=?, status='수정됨', memo=IFNULL(memo,'') || ? WHERE id=?", (title or f"출고지시서 #{order_id}", "\n" + now + " 수정", order_id))
        else:
            cur.execute("UPDATE outbound_orders SET status='수정됨', memo=IFNULL(memo,'') || ? WHERE id=?", ("\n" + now + " 수정", order_id))
        con.commit()

def build_outbound_order_title(customer_name, cart_items, fallback_title=""):
    """출고지시서 제목 기본값 생성.
    형식: [매출처] [1번째 제품명] 외 x품목
    """
    customer = str(customer_name or "").strip()
    items = list(cart_items or [])
    if not items:
        return str(fallback_title or "").strip()
    first = items[0] or {}
    first_name = str(first.get("제품명") or first.get("product_name") or "").strip()
    if customer and first_name:
        title = f"{customer} - {first_name}"
    elif customer:
        title = customer
    else:
        title = first_name
    rest_count = max(0, len(items) - 1)
    if rest_count:
        title = f"{title} 외 {rest_count}품목"
    return title or str(fallback_title or "").strip()


def _fallback_restore_outbound_from_items(cur, order_id, now):
    """출고지시 품목 테이블 기준으로 취소 수량을 현재 재고에 더한다.
    transactions.final_stock은 재고행 수량이 아니라 표준제품명 전체 총재고 스냅샷으로 저장한다.
    """
    item_rows = cur.execute("""SELECT inventory_id, location, product_name, lot, exp_date, qty, company, warehouse_name
                                   FROM outbound_order_items WHERE order_id=? ORDER BY id""", (order_id,)).fetchall()
    restored_count = 0
    for inv_id, location, product_name, lot, exp_date, qty, company, warehouse_name in item_rows:
        qty = int(qty or 0)
        if qty <= 0:
            continue
        inv = None
        if inv_id:
            inv = cur.execute("SELECT id, qty FROM inventory WHERE id=?", (int(inv_id),)).fetchone()
        if not inv:
            inv = cur.execute("""SELECT id, qty FROM inventory
                                     WHERE company=? AND product_name=? AND IFNULL(warehouse_name,'')=? AND IFNULL(lot,'')=? AND IFNULL(exp_date,'')=? AND location=?""",
                                  (company or "", product_name or "", warehouse_name or "", lot or "", exp_date or "", location or "")).fetchone()
        if inv:
            row_qty_after = int(inv[1] or 0) + qty
            cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (row_qty_after, now, int(inv[0])))
        else:
            row_qty_after = qty
            cur.execute("""INSERT INTO inventory(company, product_name, warehouse_name, lot, exp_date, location, qty, updated_at)
                               VALUES(?,?,?,?,?,?,?,?)""", (company or "", product_name or "", warehouse_name or "", lot or "-", exp_date or "-", location or "", qty, now))
        restored_count += 1
        insert_transaction_log(cur, created_at=now, tx_type="출고지시취소", product_name=product_name or "", warehouse_name=warehouse_name or "",
                               lot=lot or "-", exp_date=exp_date or "-", from_company=company or "", from_location=location or "",
                               to_company=company or "", to_location=location or "", qty=qty, memo=f"출고지시서 #{order_id} 취소 / 원복")
    return len(item_rows), restored_count


def cancel_outbound_order(order_id):
    """저장된 출고지시를 취소 처리하고 출고 품목 수량을 현재 재고에 되돌린다.

    주의: transactions.final_stock은 표준제품명 전체 총재고 스냅샷이므로
    재고행 원복 계산에 사용하지 않는다. 원복은 outbound_order_items의 품목/수량 기준으로 처리한다.
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    order_id = int(order_id)
    with connect() as con:
        cur = con.cursor()
        order = cur.execute("SELECT id, status FROM outbound_orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            raise ValueError("취소할 출고지시서를 찾을 수 없습니다.")
        if str(order[1] or "") == "취소됨":
            raise ValueError("이미 취소된 출고지시서입니다.")

        already_cancelled = cur.execute("""SELECT COUNT(*) FROM transactions
                                         WHERE tx_type='출고지시취소' AND memo LIKE ?""", (f"%출고지시서 #{order_id}%",)).fetchone()[0]
        if int(already_cancelled or 0) > 0:
            cur.execute("UPDATE outbound_orders SET status='취소됨', memo=IFNULL(memo,'') || ? WHERE id=?", ("\n" + now + " 출고지시 취소 상태 보정", order_id))
            con.commit()
            raise ValueError("이미 이 출고지시서의 재고 원복 이력이 있습니다.")

        item_count, restored_count = _fallback_restore_outbound_from_items(cur, order_id, now)
        cur.execute("UPDATE outbound_orders SET status='취소됨', memo=IFNULL(memo,'') || ? WHERE id=?", ("\n" + now + " 출고지시 취소", order_id))
        con.commit()
        return item_count, restored_count


def restore_inventory_from_log(order_id):
    """출고지시 거래이력(final_stock + qty)을 기준으로 전체 취소/원복한다."""
    return cancel_outbound_order(int(order_id))


def cancel_saved_order(order_id):
    """저장된 출고지시 취소 버튼에서 호출하는 명시적 래퍼."""
    return restore_inventory_from_log(int(order_id))


def partial_cancel_outbound_order(order_id, cancel_qty_by_item_id):
    """저장된 출고지시 품목 일부를 취소하고 해당 수량만 재고에 되돌린다.

    cancel_qty_by_item_id: {outbound_order_items.id: cancel_qty}
    - 전체 취소는 cancel_saved_order/restore_inventory_from_log를 사용한다.
    - 부분취소는 현재 지시서 품목 기준으로 수량을 줄이고, inventory에는 취소 수량만 더한다.
    """
    order_id = int(order_id)
    clean = {}
    for k, v in (cancel_qty_by_item_id or {}).items():
        try:
            iid = int(k)
            qty = int(float(v or 0))
        except Exception:
            continue
        if qty > 0:
            clean[iid] = qty
    if not clean:
        raise ValueError("부분취소할 수량이 없습니다.")

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with connect() as con:
        con.row_factory = sqlite3.Row
        cur = con.cursor()
        order = cur.execute("SELECT id, status FROM outbound_orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            raise ValueError("출고지시서를 찾을 수 없습니다.")
        if str(order["status"] or "") == "취소됨":
            raise ValueError("이미 취소된 출고지시서는 부분취소할 수 없습니다.")

        item_ids = list(clean.keys())
        placeholders = ",".join(["?"] * len(item_ids))
        rows = cur.execute(f"""
            SELECT id, order_id, inventory_id, location, product_name, lot, exp_date, qty, company, warehouse_name
            FROM outbound_order_items
            WHERE order_id=? AND id IN ({placeholders})
            ORDER BY id
        """, [order_id] + item_ids).fetchall()
        row_map = {int(r["id"]): r for r in rows}
        missing = [iid for iid in item_ids if iid not in row_map]
        if missing:
            raise ValueError(f"출고지시 품목을 찾을 수 없습니다: {missing}")

        restored_lines = 0
        restored_qty_total = 0
        for iid, cancel_qty in clean.items():
            r = row_map[iid]
            original_qty = int(r["qty"] or 0)
            if cancel_qty > original_qty:
                raise ValueError(f"{r['product_name']} 취소수량이 지시수량보다 큽니다. 지시 {original_qty}, 취소 {cancel_qty}")

        for iid, cancel_qty in clean.items():
            r = row_map[iid]
            original_qty = int(r["qty"] or 0)
            remain_qty = original_qty - cancel_qty
            inv_id = r["inventory_id"]
            inv = None
            if inv_id:
                inv = cur.execute("SELECT id, qty FROM inventory WHERE id=?", (int(inv_id),)).fetchone()
            if not inv:
                inv = cur.execute("""
                    SELECT id, qty FROM inventory
                    WHERE company=? AND product_name=? AND IFNULL(warehouse_name,'')=?
                      AND IFNULL(lot,'')=? AND IFNULL(exp_date,'')=? AND location=?
                """, (r["company"] or "", r["product_name"] or "", r["warehouse_name"] or "", r["lot"] or "-", r["exp_date"] or "-", r["location"] or "")).fetchone()
            if inv:
                row_qty_after = int(inv["qty"] or 0) + int(cancel_qty)
                cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (row_qty_after, now, int(inv["id"])))
            else:
                row_qty_after = int(cancel_qty)
                cur.execute("""
                    INSERT INTO inventory(company, product_name, warehouse_name, lot, exp_date, location, qty, updated_at)
                    VALUES(?,?,?,?,?,?,?,?)
                """, (r["company"] or "", r["product_name"] or "", r["warehouse_name"] or "", r["lot"] or "-", r["exp_date"] or "-", r["location"] or "", row_qty_after, now))

            if remain_qty > 0:
                cur.execute("UPDATE outbound_order_items SET qty=? WHERE id=?", (remain_qty, iid))
            else:
                cur.execute("DELETE FROM outbound_order_items WHERE id=?", (iid,))

            insert_transaction_log(cur, created_at=now, tx_type="출고지시부분취소", product_name=r["product_name"] or "", warehouse_name=r["warehouse_name"] or "",
                                   lot=r["lot"] or "-", exp_date=r["exp_date"] or "-", from_company=r["company"] or "", from_location=r["location"] or "",
                                   to_company=r["company"] or "", to_location=r["location"] or "", qty=int(cancel_qty), memo=f"출고지시서 #{order_id} 부분취소")
            restored_lines += 1
            restored_qty_total += int(cancel_qty)

        remaining_count = cur.execute("SELECT COUNT(*) FROM outbound_order_items WHERE order_id=?", (order_id,)).fetchone()[0]
        if int(remaining_count or 0) == 0:
            new_status = "취소됨"
            memo_add = f"\n{now} 출고지시 전체 부분취소 완료"
        else:
            new_status = "수정됨"
            memo_add = f"\n{now} 출고지시 부분취소: {restored_qty_total}EA 원복"
        cur.execute("UPDATE outbound_orders SET status=?, memo=IFNULL(memo,'') || ? WHERE id=?", (new_status, memo_add, order_id))
        con.commit()
        return restored_lines, restored_qty_total, int(remaining_count or 0)

def outbound_inventory(src_id, qty, memo="출고지시 완료"):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with connect() as con:
        cur = con.cursor()
        src = cur.execute("SELECT * FROM inventory WHERE id=?", (src_id,)).fetchone()
        cols = [d[0] for d in cur.description]
        src = dict(zip(cols, src)) if src else None
        if not src:
            raise ValueError("출고 재고를 찾을 수 없습니다.")
        if qty <= 0 or qty > src["qty"]:
            raise ValueError("출고 수량이 현재 재고보다 많거나 올바르지 않습니다.")
        cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?", (src["qty"]-qty, now, src_id))
        insert_transaction_log(cur, created_at=now, tx_type="출고", product_name=src["product_name"], warehouse_name=src.get("warehouse_name"),
                               lot=src["lot"], exp_date=src["exp_date"], from_company=src["company"], from_location=src["location"],
                               to_company=None, to_location=None, qty=qty, memo=memo)
        con.commit()


# ---------------- files / product images ----------------
def asset_dir():
    d = Path(__file__).parent / "data" / "product_images"
    d.mkdir(parents=True, exist_ok=True)
    return d

def safe_filename(name):
    keep = []
    for ch in str(name):
        if ch.isalnum() or ch in ("-", "_", "."):
            keep.append(ch)
        else:
            keep.append("_")
    return "".join(keep)[:80] or "product"

def save_product_image(product_name, uploaded_file):
    if uploaded_file is None or not product_name:
        return None
    suffix = Path(uploaded_file.name).suffix.lower() or ".jpg"
    fname = safe_filename(product_name) + "_" + datetime.now().strftime("%Y%m%d_%H%M%S") + suffix
    path = asset_dir() / fname
    path.write_bytes(uploaded_file.getvalue())
    rel = str(path.relative_to(Path(__file__).parent))
    exec_sql("UPDATE products SET image_path=? WHERE standard_name=?", (rel, product_name))
    return rel

def get_product_image_path(product_name):
    df = q("SELECT image_path FROM products WHERE standard_name=?", (product_name,))
    if df.empty:
        return ""
    value = str(df.iloc[0].get("image_path") or "")
    full = Path(__file__).parent / value
    return str(full) if value and full.exists() else ""

def product_image_placeholder(size=60):
    st.markdown(f"""
    <div style='width:{size}px;height:{size}px;border:1.5px dashed #cbd5e1;border-radius:14px;background:#f8fafc;display:flex;align-items:center;justify-content:center;color:#94a3b8;font-size:22px;'>📷</div>
    """, unsafe_allow_html=True)

# ---------------- map UI ----------------
def _loc_group_from_df(df):
    data = {}
    for r in df.itertuples():
        loc = str(r.location)
        data.setdefault(loc, []).append({
            "id": int(r.id),
            "company": str(r.company),
            "product_name": str(r.product_name),
            "warehouse_name": str(r.warehouse_name or "-"),
            "lot": str(r.lot or "-"),
            "exp_date": display_date_only(r.exp_date),
            "qty": int(r.qty),
        })
    return data

def render_location_map():
    """새로고침 없는 로케이션 맵.
    Streamlit 버튼/링크 대신 components.html 내부 JavaScript로 오른쪽 상세패널만 갱신한다.
    """
    df = q("SELECT id, company, product_name, warehouse_name, lot, exp_date, location, qty FROM inventory WHERE qty>0 ORDER BY location, company, product_name")
    loc_data = _loc_group_from_df(df)
    tx = q("""SELECT created_at, tx_type, product_name, lot, exp_date, from_location, to_location, qty
              FROM transactions ORDER BY id DESC LIMIT 300""")
    tx_data = tx.to_dict("records") if not tx.empty else []
    selected_loc = st.session_state.get("selected_location", "")
    payload = json.dumps({"inventory": loc_data, "tx": tx_data, "selected_location": selected_loc}, ensure_ascii=False)

    def dot(loc):
        if loc == "N":
            has = any(x in loc_data for x in SPECIAL_LOCATIONS)
        else:
            has = loc in loc_data or any(k.startswith(loc + "-") for k in loc_data)
        return '<span class="stock-dot"></span>' if has else ''

    def cell(loc, text=None):
        text = text or loc
        return f'<button type="button" class="map-cell" data-loc="{escape(loc)}">{escape(text)}{dot(loc)}</button>'

    def rack(area, labels, left, top, cls):
        cells = ''.join(cell(x) for x in labels)
        return f'<div class="rack {cls}" style="left:{left}px;top:{top}px;">{cells}</div>'

    def zone(loc, text, left, top, w, h, cls="white", extra=""):
        return f'<button type="button" class="zone {cls}" data-loc="{escape(loc)}" style="left:{left}px;top:{top}px;width:{w}px;height:{h}px;{extra}">{text}{dot(loc)}</button>'

    html = f"""
<!doctype html><html><head><meta charset="utf-8">
<style>
*{{box-sizing:border-box}} body{{margin:0;background:#f8fafc;font-family:Inter,Segoe UI,Arial,'Noto Sans KR',sans-serif;color:#0f172a;}}
.wms-wrap{{display:grid;grid-template-columns:minmax(0,3fr) minmax(300px,1fr);gap:22px;align-items:start;width:100%;}}
.map-card,.side-card{{background:#fff;border:1px solid #dbe4f0;border-radius:18px;padding:16px;box-shadow:0 8px 24px rgba(15,23,42,.06);}}
.legend-wrap{{display:flex;gap:12px;flex-wrap:wrap;margin:0 0 14px 0;}}
.legend-chip{{display:flex;align-items:center;gap:8px;border:1px solid #dbe4f0;background:#fff;border-radius:12px;padding:9px 14px;font-weight:800;color:#111827;font-size:14px;}}
.swatch{{width:18px;height:18px;border-radius:5px;border:1px solid rgba(15,23,42,.12);display:inline-block;}}
.swatch.y{{background:#fff39b}} .swatch.b{{background:#68d2e7}} .swatch.p{{background:#f0a7e6}} .swatch.g{{background:#f3f4f6}}
.map-scroll{{overflow:hidden;padding-bottom:0;height:654px;}}
.map-stage{{position:relative;width:1064px;height:618px;min-width:1064px;background:#fff;border-radius:14px;transform:scale(0.982);transform-origin:top left;}}
.rack{{position:absolute;width:116px;height:154px;display:grid;grid-template-columns:1fr 1fr;grid-template-rows:1fr 1fr 1fr;border:1px solid #334155;border-radius:9px;overflow:hidden;box-shadow:0 6px 14px rgba(15,23,42,.06);}}
.map-cell,.zone{{appearance:none;position:relative;display:flex;align-items:center;justify-content:center;text-decoration:none;color:#0f172a;font-weight:900;font-size:14px;border:0;border-right:1px solid rgba(51,65,85,.38);border-bottom:1px solid rgba(51,65,85,.38);cursor:pointer;font-family:inherit;}}
.map-cell:hover,.zone:hover{{outline:3px solid rgba(37,99,235,.22);z-index:2;}}
.map-cell:nth-child(2n){{border-right:none;}}
.map-cell:nth-child(n+5){{border-bottom:none;}}

.special-menu{{position:absolute;display:none;z-index:30;background:#fff;border:1px solid #cbd5e1;border-radius:12px;box-shadow:0 12px 28px rgba(15,23,42,.18);padding:6px;}}
.special-menu.open{{display:grid;gap:5px;}}
.special-menu button,.special-menu a{{appearance:none;border:1px solid #e2e8f0;background:#f8fafc;border-radius:9px;padding:8px 7px;font-size:12px;font-weight:900;color:#0f172a;cursor:pointer;font-family:inherit;text-align:center;text-decoration:none;}}
.special-menu button:hover,.special-menu button.selected,.special-menu a:hover,.special-menu a.selected{{background:#22c55e;color:#fff;border-color:#16a34a;}}
.map-cell.selected,.zone.selected{{background:#22c55e!important;color:#ffffff!important;outline:3px solid rgba(34,197,94,.35)!important;box-shadow:0 0 0 3px rgba(255,255,255,.8),0 0 18px rgba(34,197,94,.75)!important;border-color:#16a34a!important;z-index:4;}}
.yellow{{background:#fff39b;}} .blue{{background:#68d2e7;}} .pink{{background:#f0a7e6;}} .gray{{background:#f7f8fa;}} .bidata{{background:#d1d5db;}} .white{{background:#fff;}} .yellow .map-cell,.zone.yellow{{background:#fff39b;}} .blue .map-cell,.zone.blue{{background:#68d2e7;}} .pink .map-cell,.zone.pink{{background:#f0a7e6;}} .gray .map-cell,.zone.gray{{background:#f7f8fa;}} .bidata .map-cell,.zone.bidata{{background:#d1d5db;}} .white .map-cell,.zone.white{{background:#fff;}}
.stock-dot{{position:absolute;right:7px;top:7px;width:9px;height:9px;background:#65d84f;border:1.5px solid #166534;border-radius:999px;box-shadow:0 0 0 2px rgba(255,255,255,.65);}}
.zone{{position:absolute;border:1px solid #334155;border-radius:9px;box-shadow:0 6px 14px rgba(15,23,42,.04);}}
.big-left{{position:absolute;left:0;top:0;width:170px;height:258px;border:1px solid #334155;border-radius:10px;overflow:hidden;background:#fff;}}
.big-left button,.big-left a{{appearance:none;position:relative;display:flex;align-items:center;justify-content:center;width:100%;border:0;border-bottom:1px solid #cbd5e1;background:#f7f8fa;color:#0f172a;font-weight:900;cursor:pointer;font-family:inherit;text-decoration:none;}}
.big-left button:hover,.big-left a:hover{{outline:3px solid rgba(37,99,235,.22);z-index:2;}}
.big-left button.selected,.big-left a.selected{{background:#22c55e!important;color:#fff!important;outline:3px solid rgba(34,197,94,.35)!important;box-shadow:0 0 0 3px rgba(255,255,255,.8),0 0 18px rgba(34,197,94,.75)!important;z-index:4;}}
.g2{{height:205px;background:#f7f8fa;}} .g1row{{height:52px;display:grid;grid-template-columns:1fr 1fr 1fr;}}
.g1row button,.g1row a{{height:52px;border-right:1px solid #cbd5e1;border-bottom:none;}} .g1row button:last-child,.g1row a:last-child{{border-right:none;}}

.label{{position:absolute;text-align:center;font-weight:900;color:#111827;font-size:14px;}}
.memo{{position:absolute;color:#334155;font-size:15px;line-height:2.6;}}
.qp{{position:absolute;left:0;top:480px;width:150px;height:135px;border:1px solid #cbd5e1;border-radius:10px;overflow:hidden;background:#fff;}}
.qp button,.qp a{{position:relative;display:grid;grid-template-columns:54px 1fr;align-items:center;width:100%;height:67px;border:0;border-bottom:1px solid #e2e8f0;background:#fff;color:#111827;font-weight:900;cursor:pointer;text-align:left;font-family:inherit;text-decoration:none;}}
.qp button:last-child{{border-bottom:none;}}
.qp button.selected{{background:#22c55e!important;color:#ffffff!important;outline:3px solid rgba(34,197,94,.35)!important;box-shadow:0 0 0 3px rgba(255,255,255,.8),0 0 18px rgba(34,197,94,.75)!important;border-color:#16a34a!important;z-index:4;}}
.qp button.selected .qp-key{{background:#16a34a!important;color:#fff!important;}}
.qp-key{{height:100%;display:flex;align-items:center;justify-content:center;color:#ff221a;font-weight:900;font-size:18px;border-right:1px solid #e2e8f0;}}
.qp .qkey{{background:#f186ca;color:#ff0d0d;}}
.rec-red{{color:#ff1e12;font-weight:900;}}
.small-title{{position:absolute;font-size:14px;font-weight:900;color:#111827;text-align:center;}}
.side-card{{height:705px;overflow:auto;}}
.side-title{{font-size:22px;font-weight:900;margin:0 0 4px;}} .caption{{color:#64748b;font-size:13px;margin-bottom:10px;}}
.zone-pill{{display:inline-block;background:#e8f5ee;color:#15803d;font-weight:900;border-radius:10px;padding:6px 10px;margin:8px 0 12px;}}
.metric{{background:#f8fafc;border:1px solid #e2e8f0;border-radius:14px;padding:12px;margin-bottom:12px;}}
.metric .n{{font-size:24px;font-weight:900;}}
.level{{font-size:17px;font-weight:900;margin:14px 0 6px;}}
.level-tabs{{display:flex;align-items:flex-end;gap:0;margin:8px 0 12px;border-bottom:2px solid #111827;}}
.level-tab{{appearance:none;border:2px solid #111827;border-bottom:none;background:#f8fafc;color:#111827;font-weight:900;font-size:15px;padding:9px 18px;border-radius:11px 11px 0 0;margin-right:-1px;cursor:pointer;font-family:inherit;}}
.level-tab.active{{background:#fff;transform:translateY(2px);padding-top:14px;padding-bottom:12px;}}
.level-panel{{display:none;}}
.level-panel.active{{display:block;}}
.detail-card{{background:white;border:1px solid #dbe4f0;border-radius:14px;padding:12px;margin:8px 0;box-shadow:0 5px 16px rgba(15,23,42,.05);}}
.product-inline-detail{{margin:8px 0 14px 0;}}
.card-top{{display:flex;justify-content:space-between;align-items:center;gap:8px;}} .company-badge{{display:inline-block;background:#f1f5f9;color:#475569;font-weight:500;border:1px solid #e2e8f0;border-radius:999px;padding:3px 8px;font-size:12px;}}
.product-title{{font-weight:400;font-size:18px;line-height:1.25;margin-top:8px;margin-bottom:6px;color:#111827;}}
.lot-exp{{font-weight:400;font-size:14px;color:#334155;line-height:1.55;margin-top:6px;}}
.loc-line{{font-weight:400;font-size:14px;color:#334155;line-height:1.55;margin-top:6px;}}
.muted{{color:#64748b;font-size:12px;line-height:1.6;}}
.qty-text{{font-weight:900;color:#111827;white-space:nowrap;}}
.loc-link{{border:1px solid #e2e8f0;background:#ffffff;border-radius:10px;padding:7px 10px;margin:5px 0;width:100%;display:flex;justify-content:space-between;align-items:center;cursor:pointer;color:#0f172a;font-weight:700;}}
.loc-link:hover{{background:#f1f5f9;border-color:#94a3b8;}}
.prod-name-large{{font-size:12pt;font-weight:400;line-height:1.25;color:#111827;margin:8px auto 6px;}}
.prod-search-form{{margin:0;padding:0;}}
.prod-search-title{{appearance:none;border:0;background:transparent;display:block;text-align:center;cursor:pointer;padding:0;font-family:inherit;text-decoration:none;width:100%;}}
.prod-search-title:hover{{color:#1d4ed8;text-decoration:underline;}}
.prod-btn{{display:block;margin:10px auto 0;border:1px solid #bfdbfe;background:#eff6ff;color:#1d4ed8;border-radius:10px;padding:7px 16px;font-weight:800;cursor:pointer;min-width:150px;text-align:center;}}
.prod-box{{border-top:1px solid #e2e8f0;margin-top:14px;padding-top:14px;text-align:center;}} .photo-box{{width:150px;height:150px;margin:0 auto 10px;border:1px dashed #cbd5e1;border-radius:16px;background:#f8fafc;display:flex;align-items:center;justify-content:center;color:#94a3b8;font-weight:700;}}

.detail-total-text{{display:flex;gap:8px;align-items:baseline;justify-content:center;color:#334155;font-size:13px;margin:2px auto 10px;}}
.detail-total-text strong{{font-weight:600;color:#111827;}}
.loc-metric .caption{{text-align:left!important;}}
.recent-title{{text-align:center;margin:8px 0 4px!important;font-size:14px;}}
.recent-list{{text-align:left;}}
.tx-row{{font-size:12px;border-bottom:1px solid #f1f5f9;padding:6px 0;color:#334155;text-align:left;}}
@media(max-width:1100px){{.wms-wrap{{grid-template-columns:1fr}}.side-card{{height:auto}}}}
</style></head><body>
<div class="wms-wrap">
  <div class="map-card">
    <div class="legend-wrap">
      <div class="legend-chip"><span class="swatch y"></span>노투스팜</div>
      <div class="legend-chip"><span class="swatch b"></span>노투스</div>
      <div class="legend-chip"><span class="swatch p"></span>NOH</div>
      <div class="legend-chip"><span class="swatch g"></span>비자료</div>
    </div>
    <div class="map-scroll"><div class="map-stage">
      <div class="big-left">
        <button type="button" class="g2 gray" data-loc="G2">G2{dot('G2')}</button>
        <div class="g1row">
          <button type="button" data-loc="G1-01">G1-01{dot('G1-01')}</button>
          <button type="button" data-loc="G1-02">G1-02{dot('G1-02')}</button>
          <button type="button" data-loc="G1-03">G1-03{dot('G1-03')}</button>
        </div>
      </div>
      {rack('A2',['A2-03','A2-04','A2-02','A2-05','A2-01','A2-06'],210,0,'yellow')}
      {rack('B2',['B2-03','B2-04','B2-02','B2-05','B2-01','B2-06'],340,0,'yellow')}
      {rack('C2',['C2-03','C2-04','C2-02','C2-05','C2-01','C2-06'],470,0,'blue')}
      {rack('D1',['D1-03','D1-04','D1-02','D1-05','D1-01','D1-06'],600,0,'blue')}
      {zone('T1','T1',600,154,116,48,'white')}
      {rack('E1',['E1-03','E1-04','E1-02','E1-05','E1-01','E1-06'],730,0,'pink')}
      {zone('T2','T2',730,154,116,48,'white')}
      {zone('F1-01','F1-01',875,0,58,48,'bidata')}
      {zone('F1-02','F1-02',933,0,58,48,'bidata')}
      {zone('F1-03','F1-03',991,0,58,48,'bidata')}
      <div class="small-title" style="left:915px;top:66px;width:100px;">비자료</div>
      {zone('X2','X2',1070,0,64,48,'gray')}
      {rack('A1',['A1-03','A1-04','A1-02','A1-05','A1-01','A1-06'],210,245,'yellow')}
      {rack('B1',['B1-03','B1-04','B1-02','B1-05','B1-01','B1-06'],340,245,'yellow')}
      {rack('C1',['C1-03','C1-04','C1-02','C1-05','C1-01','C1-06'],470,245,'yellow')}
      <div class="memo" style="left:760px;top:270px;line-height:1.55;">X1-01~03 : 폐기<br>X1-01-01 : 대표님 시술용</div>
      {zone('X1-01','X1-01',1010,245,58,52,'gray')}
      {zone('X1-02','X1-02',1010,297,58,52,'gray')}
      {zone('X1-03','X1-03',1010,349,58,52,'gray')}
      <div class="qp">
        <button type="button" data-loc="Q"><span class="qp-key qkey">Q</span><span>유통기간임박</span>{dot('Q')}</button>
        <button type="button" data-loc="P"><span class="qp-key">P</span><span>수출대기</span>{dot('P')}</button>
      </div>
      {zone('REC','<span><span class="rec-red">REC</span>eiving</span>',340,520,130,52,'white')}
      <div class="label" style="left:340px;top:582px;width:130px;">매입등록대기</div>
      {zone('R2','R2',725,420,58,52,'white')}
      {zone('R1','R1',783,420,58,52,'white')}
      <div class="label" style="left:706px;top:482px;width:170px;">R2 비자료 / R1 자료</div>
      {zone('N','기타 위치',930,565,155,60,'white')}
      <div class="special-menu" id="specialMenu" style="left:930px;top:428px;width:155px;"><button type="button" data-special-loc="홍보물랙">홍보물랙</button><button type="button" data-special-loc="회색 카트">회색 카트</button><button type="button" data-special-loc="오른쪽 창고">오른쪽 창고</button><button type="button" data-special-loc="사무실(4층)">사무실(4층)</button></div>
    </div></div>
  </div>
  <div class="side-card" id="detail"><div class="side-title">위치 상세 정보</div><div class="caption">맵에서 로케이션을 선택하면 상세 재고가 여기에 표시됩니다.</div></div>
</div>
<script>
const DATA = {payload};
const inventory = DATA.inventory || {{}};
const txData = DATA.tx || [];
const specialLocations = ["홍보물랙", "회색 카트", "오른쪽 창고", "사무실(4층)"];
const initialSelectedLocation = DATA.selected_location || "";
function zoneName(loc){{
  if(specialLocations.includes(loc||'')) return '기타 위치';
  const area=(loc||'').split('-')[0];
  if(['A1','A2','B1','B2','C1'].includes(area)) return '노투스팜';
  if(['C2','D1'].includes(area)) return '노투스';
  if(area==='E1') return 'NOH';
  if(area==='Q') return '유통기간임박';
  if(area==='F1') return '비자료';
  if(area==='X1') return '폐기';
  if(area==='X2') return '기타 보관 구역';
  if(area==='R1') return '냉장고(자료)';
  if(area==='R2') return '냉장고(비자료)';
  if(area==='REC') return '매입등록대기';
  if(area==='P') return '수출대기';
  if(area==='G2') return '패키지 창고';
  if(area==='N') return '기타 위치';
  return '기타 보관 구역';
}}
function levelLabel(loc){{
  if(specialLocations.includes(loc||'')) return '단 구분 없음';
  const p=(loc||'').split('-');
  if(p[0]==='X1' && p.length===2) return '1단';
  if(p.length>=3 && /^\d+$/.test(p[2])) return parseInt(p[2],10)+'단';
  return '단 구분 없음';
}}
function rowsFor(loc){{
  let rows=[];
  Object.entries(inventory).forEach(([k,v])=>{{ if(k===loc || k.startsWith(loc+'-')) rows=rows.concat(v.map(x=>({{...x, location:k}}))); }});
  return rows;
}}
function formatRecentHistory(tx, currentTotal){{
  if(!tx.length) return '<div class="muted">최근 이력이 없습니다.</div>';
  let running = Number(currentTotal || 0);
  return tx.map(t=>{{
    const type = String(t.tx_type || '이력');
    const qty = Number(t.qty || 0);
    const fromLoc = t.from_location || '-';
    const toLoc = t.to_location || '-';
    let body = '';
    if(type.includes('이동')){{
      body = `${{esc(fromLoc)}} → ${{esc(toLoc)}} (${{qty}}EA)`;
    }} else if(type.includes('입고')){{
      const after = running;
      const before = Math.max(0, after - qty);
      body = `${{before}}EA → ${{after}}EA`;
      running = before;
    }} else if(type.includes('출고')){{
      const after = running;
      const before = after + qty;
      body = `${{before}}EA → ${{after}}EA`;
      running = before;
    }} else if(type.includes('조정')){{
      body = `${{qty}}EA 조정`;
    }} else {{
      body = `${{qty}}EA`;
    }}
    const shownDate = cleanDate(String(t.created_at || '').slice(0,10));
    return `<div class="tx-row">${{esc(shownDate)}} <b>[${{esc(type)}}]</b> ${{body}}</div>`;
  }}).join('');
}}
function productDetail(name){{
  const all=[]; Object.values(inventory).forEach(arr=>arr.forEach(x=>{{ if(x.product_name===name) all.push(x); }}));
  const total=all.reduce((a,b)=>a+(b.qty||0),0);
  const locMap={{}}; const locCompanies={{}};
  Object.entries(inventory).forEach(([loc,arr])=>arr.forEach(x=>{{
    if(x.product_name===name){{
      locMap[loc]=(locMap[loc]||0)+(x.qty||0);
      if(!locCompanies[loc]) locCompanies[loc]=new Set();
      locCompanies[loc].add(x.company||'-');
    }}
  }}));
  const locRows=Object.entries(locMap).sort((a,b)=>a[0].localeCompare(b[0])).map(([loc,qty])=>{{
    const comp=Array.from(locCompanies[loc]||[]).join(', ');
    return `<button class="loc-link" type="button" data-jump-loc="${{esc(loc)}}"><span>${{esc(loc)}} <em style="font-style:normal;color:#64748b;font-size:12px;">${{esc(comp)}}</em></span><span>${{qty}} EA</span></button>`;
  }}).join('');
  const tx=txData.filter(t=>t.product_name===name).slice(0,5);
  return `<div class="prod-box"><div class="photo-box">📷</div><form class="prod-search-form" method="get" target="_top" action="" data-search-form="1"><input type="hidden" name="map_search_product" value="${{esc(name)}}"><button type="submit" class="prod-name-large prod-search-title" data-search-product="${{esc(name)}}">${{esc(name)}}</button></form><div class="detail-total-text"><span>창고 총재고</span><strong>${{total}} EA</strong></div><div class="metric loc-metric"><div class="caption">분산 로케이션</div>${{locRows||'<div class="muted">재고 위치가 없습니다.</div>'}}</div><h4 class="recent-title">최근 이력 5건</h4><div class="recent-list">${{formatRecentHistory(tx,total)}}</div></div>`;
}}
function productCardsHtml(rows){{
  const products={{}};
  rows.forEach(r=>{{
    const key=r.product_name||'-';
    if(!products[key]) products[key]=[];
    products[key].push(r);
  }});
  return Object.entries(products).map(([name,items])=>{{
    const total=items.reduce((a,b)=>a+(Number(b.qty)||0),0);
    const companies=Array.from(new Set(items.map(x=>x.company||'-'))).sort().join(', ');
    const lines=items
      .slice()
      .sort((a,b)=>String(a.exp_date||'').localeCompare(String(b.exp_date||'')) || String(a.lot||'').localeCompare(String(b.lot||'')) || String(a.company||'').localeCompare(String(b.company||'')))
      .map(x=>{{
        const companyInfo = companies.includes(',') ? `<span class="company-badge">${{esc(x.company||'-')}}</span> ` : '';
        return `<div class="lot-exp">${{companyInfo}}${{Number(x.qty)||0}}EA&nbsp;&nbsp;${{esc(x.lot||'-')}} | ${{esc(cleanDate(x.exp_date||'-'))}}</div>`;
      }}).join('');
    return `<div class="detail-card"><div class="card-top"><span class="product-title">${{esc(name)}}</span><span class="qty-text">${{total}} EA</span></div><div class="muted">사업장: ${{esc(companies||'-')}}</div>${{lines}}<button class="prod-btn" type="button" data-product="${{esc(name)}}">제품 상세 보기</button></div>`;
  }}).join('');
}}
function cleanDate(v){{
  const s=String(v ?? '').trim();
  if(!s || s==='-' || s.toLowerCase()==='nan') return '-';
  const m=s.match(/^(\d{{4}})[-/.](\d{{1,2}})[-/.](\d{{1,2}})/);
  if(m) return `${{m[1]}}-${{String(m[2]).padStart(2,'0')}}-${{String(m[3]).padStart(2,'0')}}`;
  return s;
}}
function esc(v){{return String(v ?? '').replaceAll('&','&amp;').replaceAll('<','&lt;').replaceAll('>','&gt;').replaceAll('"','&quot;').replaceAll("'",'&#39;')}}
function parentBaseHref(){{
  try {{ return window.top.location.href; }} catch(e) {{}}
  try {{ return window.parent.location.href; }} catch(e) {{}}
  return document.referrer || window.location.href;
}}
function buildParentUrl(key, value, removeKey){{
  const url = new URL(parentBaseHref());
  url.searchParams.set(key, value);
  if(removeKey) url.searchParams.delete(removeKey);
  return url.toString();
}}

function tryParentSearch(productName){{
  try {{
    const doc = window.parent.document;
    const inputs = Array.from(doc.querySelectorAll('input'));
    const input = inputs.find(x => (x.getAttribute('aria-label')||'').includes('제품명 검색'));
    if(input){{
      const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
      setter.call(input, productName);
      input.dispatchEvent(new InputEvent('input', {{bubbles:true, inputType:'insertText', data:productName}}));
      input.dispatchEvent(new Event('change', {{bubbles:true}}));
      const buttons = Array.from(doc.querySelectorAll('button'));
      const searchBtn = buttons.find(b => (b.innerText||'').trim()==='검색');
      if(searchBtn){{ searchBtn.click(); return true; }}
    }}
  }} catch(e) {{}}
  return false;
}}
function setFormsToParent(){{
  document.querySelectorAll('form[data-search-form]').forEach(f=>{{
    try {{ f.action = buildParentUrl('map_search_product', f.querySelector('[name="map_search_product"]').value || '', 'inbound_loc'); }} catch(e) {{}}
  }});
}}

function navigateTop(href){{
  try {{ window.top.location.assign(href); return; }} catch(e) {{}}
  try {{ window.parent.location.assign(href); return; }} catch(e) {{}}
  try {{ window.open(href, '_top'); return; }} catch(e) {{}}
  const a = document.createElement('a');
  a.href = href;
  a.target = '_top';
  a.rel = 'noopener';
  document.body.appendChild(a);
  a.click();
}}
function showDetail(loc){{
  document.querySelectorAll('[data-loc]').forEach(b=>{{
    const cellLoc = b.dataset.loc || '';
    b.classList.toggle('selected', cellLoc===loc || String(loc||'').startsWith(cellLoc+'-') || (cellLoc==='N' && specialLocations.includes(loc||'')));
  }});
  const rows=rowsFor(loc);
  const d=document.getElementById('detail');
  let title=loc;
  const p=loc.split('-'); if(p.length>=2) title=p[0]+'-'+p[1];
  if(!rows.length){{d.innerHTML=`<div class="side-title">${{esc(title)}}</div><div class="zone-pill">${{esc(zoneName(loc))}}</div><div class="caption">현재 이 로케이션에는 표시할 재고가 없습니다.</div>`; return;}}
  const total=rows.reduce((a,b)=>a+(b.qty||0),0);
  const order={{'1단':0,'2단':1,'3단':2,'단 구분 없음':9}};
  rows.sort((a,b)=>{{return (order[levelLabel(a.location)]??5)-(order[levelLabel(b.location)]??5) || a.company.localeCompare(b.company) || a.product_name.localeCompare(b.product_name);}});
  const grouped={{}};
  rows.forEach(r=>{{const lvl=levelLabel(r.location); if(!grouped[lvl]) grouped[lvl]=[]; grouped[lvl].push(r);}});
  const levels=['1단','2단','3단'].filter(l=>grouped[l]);
  Object.keys(grouped).forEach(l=>{{if(!levels.includes(l)) levels.push(l);}});
  let html=`<div class="side-title">${{esc(title)}}</div><div class="zone-pill">${{esc(zoneName(loc))}}</div><div class="metric"><div class="caption">현재 총재고</div><div class="n">${{total}} EA</div></div>`;
  if(levels.length>1){{
    html+=`<div class="level-tabs">${{levels.map((lvl,i)=>`<button class="level-tab ${{i===0?'active':''}}" type="button" data-level-tab="${{esc(lvl)}}">${{esc(lvl)}}</button>`).join('')}}</div>`;
  }}
  levels.forEach((lvl,i)=>{{
    html+=`<div class="level-panel ${{i===0?'active':''}}" data-level-panel="${{esc(lvl)}}">`;
    if(levels.length===1) html+=`<div class="level">${{esc(lvl)}}</div>`;
    html+=productCardsHtml(grouped[lvl]);
    html+=`</div>`;
  }});
  d.innerHTML=html;
  d.querySelectorAll('[data-level-tab]').forEach(tab=>tab.addEventListener('click',()=>{{
    const lvl=tab.dataset.levelTab;
    d.querySelectorAll('[data-level-tab]').forEach(x=>x.classList.toggle('active', x.dataset.levelTab===lvl));
    d.querySelectorAll('[data-level-panel]').forEach(x=>x.classList.toggle('active', x.dataset.levelPanel===lvl));
  }}));
  d.querySelectorAll('[data-product]').forEach(btn=>btn.addEventListener('click',()=>{{
    const card=btn.closest('.detail-card');
    const old=card.nextElementSibling;
    if(old && old.classList.contains('product-inline-detail')){{old.remove(); btn.disabled=false; return;}}
    d.querySelectorAll('.product-inline-detail').forEach(x=>x.remove());
    d.querySelectorAll('[data-product]').forEach(x=>x.disabled=false);
    card.insertAdjacentHTML('afterend', `<div class="product-inline-detail">${{productDetail(btn.dataset.product)}}</div>`);
    btn.disabled=true;
    const box=card.nextElementSibling;
    box.querySelectorAll('[data-jump-loc]').forEach(j=>j.addEventListener('click',()=>showDetail(j.dataset.jumpLoc)));
    box.querySelectorAll('[data-search-product]').forEach(t=>t.addEventListener('click',(ev)=>{{
      ev.preventDefault();
      ev.stopPropagation();
      const productName = t.dataset.searchProduct || '';
      if(!productName) return;
      if(tryParentSearch(productName)) return;
      const href = buildParentUrl('map_search_product', productName, 'inbound_loc');
      const form = t.closest('form');
      if(form) form.action = href;
      navigateTop(href);
    }}));
    setFormsToParent();
  }}));
}}
function toggleSpecialMenu(forceClose=false){{
  const menu=document.getElementById('specialMenu');
  if(!menu) return;
  if(forceClose){{menu.classList.remove('open'); return;}}
  menu.classList.toggle('open');
}}
document.querySelectorAll('[data-special-loc]').forEach(btn=>btn.addEventListener('click',(ev)=>{{
  ev.preventDefault(); ev.stopPropagation();
  const loc=btn.dataset.specialLoc || '';
  toggleSpecialMenu(true);
  document.querySelectorAll('[data-special-loc]').forEach(x=>x.classList.toggle('selected', x.dataset.specialLoc===loc));
  showDetail(loc);
}}));
document.querySelectorAll('[data-loc]').forEach(btn=>btn.addEventListener('click',(ev)=>{{
  const loc=btn.dataset.loc || '';
  if(loc==='N'){{
    ev.preventDefault(); ev.stopPropagation();
    toggleSpecialMenu(false);
    return;
  }}
  toggleSpecialMenu(true);
  showDetail(loc);
}}));
if(initialSelectedLocation){{setTimeout(()=>showDetail(initialSelectedLocation), 80);}}
</script></body></html>
"""
    components.html(html, height=790, scrolling=False)
# ---------------- detail ----------------
def location_zone_name(loc):
    if (loc or '') in SPECIAL_LOCATIONS:
        return '기타 위치'
    area = (loc or '').split('-')[0]
    if area in ["A1","A2","B1","B2","C1"]: return "노투스팜"
    if area in ["C2","D1"]: return "노투스"
    if area == "E1": return "NOH"
    if area == "Q": return "유통기간임박"
    if area == "F1": return "비자료"
    if area == "X1": return "폐기"
    if area == "X2": return "기타 보관 구역"
    if area == "G2": return "패키지 창고"
    if area == "R1": return "냉장고(자료)"
    if area == "R2": return "냉장고(비자료)"
    if area == "REC": return "매입등록대기"
    if area == "P": return "수출대기"
    if area == "N": return "기타 위치"
    return "기타 보관 구역"

def level_label(location):
    if (location or '') in SPECIAL_LOCATIONS:
        return '단 구분 없음'
    parts = (location or '').split('-')
    if len(parts) == 2 and parts[0] == "X1":
        return "1단"
    if len(parts) >= 3 and parts[2].isdigit(): return f"{int(parts[2])}단"
    return "단 구분 없음"

def format_history_rows(tx, current_total):
    rows = []
    running = int(current_total or 0)
    for r in tx.itertuples():
        typ = str(getattr(r, "tx_type", "이력") or "이력")
        qty = int(getattr(r, "qty", 0) or 0)
        if "이동" in typ:
            desc = f"{getattr(r, 'from_location', '-') or '-'} → {getattr(r, 'to_location', '-') or '-'} ({qty}EA)"
        elif "입고" in typ:
            after = running
            before = max(0, after - qty)
            desc = f"{before}EA → {after}EA"
            running = before
        elif "출고" in typ:
            after = running
            before = after + qty
            desc = f"{before}EA → {after}EA"
            running = before
        elif "조정" in typ:
            desc = f"{qty}EA 조정"
        else:
            desc = f"{qty}EA"
        rows.append({"일자": str(getattr(r, "created_at", "") or "")[:10], "이력": f"[{typ}] {desc}"})
    return pd.DataFrame(rows)

def render_product_detail(product_name):
    if not product_name:
        return
    inv = q("""SELECT company, product_name, warehouse_name, lot, exp_date, location, qty
             FROM inventory WHERE product_name=? AND qty>0 ORDER BY location, company""", (product_name,))
    tx = q("""SELECT created_at, tx_type, lot, exp_date, from_location, to_location, qty
             FROM transactions WHERE product_name=? ORDER BY id DESC LIMIT 5""", (product_name,))
    total = int(inv["qty"].sum()) if not inv.empty else 0
    st.markdown("---")
    st.markdown("### 제품 상세")
    st.markdown("<div style='width:150px;height:150px;margin:0 auto 10px;border:1px dashed #cbd5e1;border-radius:16px;background:#f8fafc;display:flex;align-items:center;justify-content:center;color:#94a3b8;'>📷</div>", unsafe_allow_html=True)
    st.markdown("<div class='map-detail-title-wrap'>", unsafe_allow_html=True)
    if st.button(product_name, key=f"detail_search_product_{product_name}", use_container_width=True):
        # iframe/URL 파라미터에 의존하지 않고, 로케이션맵 검색 결과 상태를 직접 지정한다.
        # 검색 입력칸은 빈 상태로 렌더링하고, 결과만 해당 제품명으로 표시한다.
        st.session_state["_map_forced_search_term"] = product_name
        st.session_state["map_view_mode"] = "search"
        st.session_state["_last_map_product_search"] = ""
        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown(f"<div class='detail-total-text'><span>창고 총재고</span><strong>{total} EA</strong></div>", unsafe_allow_html=True)
    st.markdown("<h5 style='text-align:left;margin:8px 0 6px;'>분산 로케이션</h5>", unsafe_allow_html=True)
    if inv.empty:
        st.caption("재고 위치가 없습니다.")
    else:
        loc_sum = (
            inv.groupby("location", as_index=False)
               .agg(qty=("qty", "sum"), company=("company", lambda x: ", ".join(sorted(set([str(v) for v in x if str(v).strip()])))))
               .sort_values("location")
        )
        for r in loc_sum.itertuples():
            label = f"{r.location}  |  {r.company or '-'}    {int(r.qty)} EA"
            if st.button(label, key=f"jump_loc_{product_name}_{r.location}", use_container_width=True):
                st.session_state["selected_location"] = r.location
                st.rerun()
    st.markdown("<h5 style='text-align:center;margin:8px 0 4px;'>최근 이력 5건</h5>", unsafe_allow_html=True)
    if tx.empty:
        st.caption("최근 이력이 없습니다.")
    else:
        st.dataframe(format_history_rows(tx, total), hide_index=True, use_container_width=True)

def render_detail(loc):
    st.markdown("## 위치 상세 정보")
    if not loc:
        st.info("맵에서 로케이션을 선택하면 상세 재고가 여기에 표시됩니다.")
        return
    area, line, level = parse_location(loc)
    title = f"{area}-{line}" if line else area
    st.caption("선택 위치")
    st.markdown(f"### {title}")
    st.markdown(f"<div class='zone-pill'>{location_zone_name(loc)}</div>", unsafe_allow_html=True)
    df = q("""SELECT id, company, product_name, warehouse_name, lot, exp_date, location, qty
              FROM inventory
              WHERE qty>0 AND (location=? OR location LIKE ?)
              ORDER BY location DESC, company, product_name""", (loc, loc+'-%'))
    if df.empty:
        st.info("현재 이 로케이션에는 표시할 재고가 없습니다.")
        return
    st.metric("현재 총재고", f"{int(df['qty'].sum())} EA")
    df["단"] = df["location"].apply(level_label)
    order = {"3단": 0, "2단": 1, "1단": 2, "단 구분 없음": 9}
    df["_order"] = df["단"].map(order).fillna(5)
    df = df.sort_values(["_order", "company", "product_name"])
    for lvl, g in df.groupby("단", sort=False):
        st.markdown(f"#### {lvl}")
        for product_name, pg in g.groupby("product_name", sort=True):
            total_qty = int(pg["qty"].sum())
            companies = ", ".join(sorted({str(x) for x in pg["company"].dropna().tolist() if str(x).strip()})) or "-"
            lines = []
            for rr in pg.sort_values(["exp_date", "lot", "company"]).itertuples():
                prefix = f"<span class='company-badge'>{rr.company}</span> " if "," in companies else ""
                lines.append(f"<div class='lot-exp'>{prefix}{int(rr.qty)}EA&nbsp;&nbsp;{rr.lot or '-'} | {display_date_only(rr.exp_date)}</div>")
            html_lines = "".join(lines)
            st.markdown(f"""
            <div class="detail-card">
              <div class="card-top"><span class="product-title">{product_name}</span><span class="qty-text">{total_qty} EA</span></div>
              <div class="muted">사업장: {companies}</div>
              {html_lines}
            </div>
            """, unsafe_allow_html=True)
            _btn_l, _btn_c, _btn_r = st.columns([1, 2, 1])
            with _btn_c:
                if st.button("제품 상세 보기", key=f"prod_detail_group_{lvl}_{product_name}", use_container_width=True):
                    st.session_state["selected_product_for_detail"] = product_name
                    st.rerun()
    render_product_detail(st.session_state.get("selected_product_for_detail"))

# ---------------- pages ----------------
def page_map_search_results(term):
    """로케이션맵 > 제품명 검색 결과.
    제품 카드 내부의 재고분포를 웹 레이아웃 방식으로 재정렬한다.
    """
    term = (term or "").strip()
    st.markdown("### 제품 검색 결과")
    opts = product_options(term)
    if opts.empty:
        st.info("검색 결과가 없습니다.")
        return

    if len(opts) >= 2:
        st.markdown("""
        <style>
        .wms-floating-top{
            position:fixed;right:28px;bottom:28px;width:46px;height:46px;border-radius:999px;
            background:#0f172a;color:white!important;text-decoration:none!important;
            display:flex;align-items:center;justify-content:center;font-size:24px;font-weight:900;
            box-shadow:0 10px 24px rgba(15,23,42,.28);z-index:9999;line-height:1;
        }
        .wms-floating-top:hover{background:#2563eb;color:white!important;text-decoration:none!important;}
        </style>
        <a class="wms-floating-top" href="#wms-top-anchor" title="맨위로">↑</a>
        """, unsafe_allow_html=True)

    inv = q("""SELECT id, company, product_name, warehouse_name, lot, exp_date, location, qty
             FROM inventory WHERE qty>0""")
    if not inv.empty:
        inv["exp_date"] = inv["exp_date"].apply(display_date_only)
        for col in ["company", "warehouse_name", "lot", "location", "product_name"]:
            inv[col] = inv[col].fillna("-").astype(str)
        inv["qty"] = pd.to_numeric(inv["qty"], errors="coerce").fillna(0).astype(int)

    totals = pd.DataFrame(columns=["product_name", "total_qty"])
    if not inv.empty:
        totals = inv.groupby("product_name", as_index=False)["qty"].sum().rename(columns={"qty": "total_qty"})
    merged = opts.merge(totals, left_on="standard_name", right_on="product_name", how="left")
    merged["total_qty"] = merged["total_qty"].fillna(0).astype(int)

    st.markdown("""
    <style>
    .product-main-name{font-size:18px;font-weight:400;color:#111827;line-height:1.35;margin:14px 0 9px;word-break:keep-all;text-align:center;}
    .product-photo-panel{
        width:250px;height:250px;max-width:100%;border:1.5px dashed #d6dee9;border-radius:20px;background:linear-gradient(180deg,#ffffff,#f8fafc);
        display:flex;align-items:center;justify-content:center;color:#94a3b8;font-weight:600;font-size:20px;line-height:1.55;
        margin:0 auto 10px;overflow:hidden;
    }
    .total-card-small{width:50%;min-width:180px;border:1.5px solid #e5e7eb;border-radius:20px;padding:12px 17px;margin:4px auto 48px;display:flex;flex-direction:column;justify-content:center;align-items:center;text-align:center;background:#fafafa;box-shadow:0 2px 8px rgba(15,23,42,.025);}
    .total-label{font-size:15px;font-weight:500;color:#6b7280;text-align:center;}
    .total-value{font-size:24px;font-weight:800;color:#111827;text-align:center;}
    .dist-wrap{padding:0 0 38px 0;}
    .dist-header{font-size:18px;font-weight:800;color:#111827;margin:2px 0 12px;}
    .dist-rule{height:1px;background:#e5e7eb;margin:0 0 14px;}
    .company-section{margin:0 0 26px 0;}
    .company-section:last-child{margin-bottom:42px;}
    .company-head{display:flex;align-items:center;gap:10px;margin:0 0 10px;flex-wrap:wrap;}
    .company-pill{display:inline-flex;align-items:center;border-radius:12px;background:#e8f8ef;color:#118445;font-size:20px;font-weight:500;padding:7px 14px;white-space:nowrap;}
    .company-erp-name{font-size:14px;color:#9ca3af;font-weight:400;word-break:keep-all;}
    .company-total-blue{font-size:20px;color:#4f6fff;font-weight:700;white-space:nowrap;margin-left:2px;}
    .dist-row{display:grid;grid-template-columns:128px 160px 185px 72px;align-items:center;column-gap:24px;margin:7px 0;max-width:650px;}
    .loc-pill-btn{display:flex;align-items:center;justify-content:center;height:34px;border:1px solid #d8dee8;border-radius:9px;background:#fff;color:#334155;text-decoration:none!important;font-size:15px;font-weight:400;box-shadow:0 1px 1px rgba(15,23,42,.02);}
    .loc-pill-btn:hover{background:#f8fafc;border-color:#94a3b8;color:#111827;text-decoration:none!important;}
    .dist-row a,.dist-row a:visited,.dist-row a:hover,.dist-row a:active{text-decoration:none!important;}
    .dist-row a *{text-decoration:none!important;}
    .lot-info,.exp-info{font-size:16px;font-weight:400;color:#111827;line-height:1.35;white-space:nowrap;}
    .qty-blue{font-size:16px;font-weight:400;color:#4f6fff;text-align:right;white-space:nowrap;line-height:1.35;}
    .no-stock-box{border:1px dashed #cbd5e1;border-radius:16px;background:#f8fafc;padding:22px;color:#64748b;font-weight:800;}
    .detail-total-center{border:1px solid #e5e7eb;border-radius:16px;background:#fafafa;padding:12px 14px;margin:8px auto 14px;text-align:center;} .detail-total-narrow{width:150px;max-width:150px;min-width:150px;}
    .detail-total-label{font-size:13px;color:#6b7280;margin-bottom:4px;}
    .detail-total-value{font-size:24px;color:#111827;font-weight:600;}
    .dist-row-streamlit{margin:0 0 1px;}
    .dist-cell-text{display:flex;align-items:center;height:28px;font-size:14px;font-weight:400;color:#111827;white-space:nowrap;}
    .dist-cell-qty{display:flex;align-items:center;justify-content:flex-end;height:28px;font-size:14px;font-weight:400;color:#4f6fff;white-space:nowrap;}
    div[data-testid="stButton"] > button[kind="secondary"]{text-decoration:none;min-height:28px;height:28px;padding:0 10px;border-radius:8px;font-size:13px;}
    section[data-testid="stSidebar"] div[data-testid="stButton"] > button, section[data-testid="stSidebar"] div[data-testid="stButton"] > button[kind="secondary"], section[data-testid="stSidebar"] div[data-testid="stButton"] > button[kind="primary"]{background:transparent!important;color:white!important;border:0!important;min-height:auto!important;height:auto!important;padding:8px 10px!important;border-radius:10px!important;font-weight:800!important;font-size:123%!important;text-align:left!important;justify-content:flex-start!important;}
    section[data-testid="stSidebar"] div[data-testid="stButton"] > button p{color:white!important;text-align:left!important;width:100%!important;}
    /* 제품검색 결과 전용 CSS가 사이드바/상단 타이틀 카드까지 침범하지 않도록
       전역 stVerticalBlockBorderWrapper 배경 강제 지정은 제거한다. */
    @media(max-width:1200px){.dist-row{grid-template-columns:110px 130px 150px 65px;column-gap:16px}.lot-info,.exp-info,.qty-blue{font-size:14px}}
    </style>
    """, unsafe_allow_html=True)

    company_order = {"노투스팜": 0, "노투스": 1, "NOH": 2, "비자료": 3}

    def mapping_name_for_company(row_obj, company, product_name):
        company = str(company or "")
        attr_map = {
            "노투스팜": "erp_nohtuspharm_name",
            "NOH": "erp_noh_name",
            "노투스": "erp_nohtus_name",
            "비자료": "bidata_name",
        }
        attr = attr_map.get(company)
        if attr and hasattr(row_obj, attr):
            v = getattr(row_obj, attr, "")
            if str(v or "").strip() and str(v).strip().lower() != "nan":
                return str(v).strip()
        v = product_mapping_name_for(company, product_name)
        return str(v).strip() if str(v or "").strip() else "-"

    for r in merged.itertuples():
        product_name = str(r.standard_name)
        rows = inv[inv["product_name"] == product_name].copy() if not inv.empty else pd.DataFrame()
        total_qty = int(getattr(r, "total_qty", 0) or 0)

        with st.container(border=True):
            left, right = st.columns([0.95, 2.35], gap="large")
            with left:
                img_path = get_product_image_path(product_name)
                if img_path:
                    st.markdown(
                        f"<div class='product-photo-panel' style=\"border-style:solid;padding:0;\"><img src='file:///{img_path}' style='width:100%;height:100%;object-fit:contain;display:block;' /></div>",
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown("<div class='product-photo-panel'>제품 사진<br>(업로드 가능)</div>", unsafe_allow_html=True)
                st.markdown(f"<div class='product-main-name'>{escape(product_name)}</div>", unsafe_allow_html=True)
                st.markdown(
                    f"<div class='total-card-small'><span class='total-label'>총 재고</span><span class='total-value'>{total_qty} EA</span></div>",
                    unsafe_allow_html=True,
                )

            with right:
                st.markdown("<div class='dist-header'>재고 분포</div><div class='dist-rule'></div>", unsafe_allow_html=True)
                if rows.empty:
                    st.markdown("<div class='no-stock-box'>현재 재고가 없습니다.</div>", unsafe_allow_html=True)
                else:
                    rows["_company_order"] = rows["company"].map(company_order).fillna(9)
                    rows = rows.sort_values(["_company_order", "company", "lot", "exp_date", "location"])
                    for company, cg in rows.groupby("company", sort=False):
                        company_total = int(cg["qty"].sum())
                        erp_name = mapping_name_for_company(r, company, product_name)
                        st.markdown(
                            f"<div class='company-head'><span class='company-pill'>{escape(str(company))}</span>"
                            f"<span class='company-erp-name'>{escape(erp_name)}</span>"
                            f"<span class='company-total-blue'>{company_total} EA</span></div>",
                            unsafe_allow_html=True,
                        )
                        cg = cg.sort_values(["location", "lot", "exp_date", "warehouse_name"])
                        for rr in cg.itertuples():
                            loc = str(rr.location)
                            c_loc, c_lot, c_exp, c_qty, c_blank = st.columns([1.08, 1.05, 1.05, 0.55, 3.25], gap="small")
                            with c_loc:
                                if st.button(loc, key=f"map_dist_loc_{product_name}_{rr.id}_{loc}", use_container_width=True):
                                    st.session_state["selected_location"] = loc
                                    st.session_state["map_view_mode"] = "map"
                                    st.rerun()
                            with c_lot:
                                st.markdown(f"<div class='dist-cell-text'>제조번호: {escape(str(rr.lot or '-'))}</div>", unsafe_allow_html=True)
                            with c_exp:
                                st.markdown(f"<div class='dist-cell-text'>유통기한: {escape(display_date_only(rr.exp_date))}</div>", unsafe_allow_html=True)
                            with c_qty:
                                st.markdown(f"<div class='dist-cell-qty'>{int(rr.qty)} EA</div>", unsafe_allow_html=True)
                        st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)


def _map_search_changed():
    # 검색어를 새로 입력하면 검색결과 화면으로 즉시 돌아간다.
    st.session_state["map_view_mode"] = "search"


def page_map():
    if st.session_state.pop("_scroll_map_top", False):
        components.html("""<script>
        try { window.parent.scrollTo({top:0,left:0,behavior:'auto'}); } catch(e) {}
        try { window.parent.document.documentElement.scrollTop = 0; window.parent.document.body.scrollTop = 0; } catch(e) {}
        </script>""", height=0, scrolling=False)
    st.markdown("<div id='wms-top-anchor'></div>", unsafe_allow_html=True)
    forced_search_term = ""
    try:
        qprod = st.session_state.pop("_map_forced_search_term", "") or st.session_state.pop("_pending_map_search_product", "") or st.query_params.get("map_search_product", "")
        if isinstance(qprod, list):
            qprod = qprod[0] if qprod else ""
        forced_search_term = str(qprod or "").strip()
        if forced_search_term:
            st.session_state["map_view_mode"] = "search"
            st.session_state["_last_map_product_search"] = ""
            # 제품 상세에서 표준제품명을 눌러 들어온 검색은 결과만 표시하고 입력칸은 비운다.
            # 위젯 생성 전에만 session_state 값을 정리하므로 widget key 직접 수정 오류가 나지 않는다.
            for _k in ["map_product_search", "map_product_search_forced_blank"]:
                if _k in st.session_state:
                    st.session_state[_k] = ""
            try:
                del st.query_params["map_search_product"]
            except Exception:
                pass
    except Exception:
        forced_search_term = ""
    h1, h2 = st.columns([1.2, 1.8], gap="large")
    with h1:
        st.title("📍로케이션 맵")
    with h2:
        # form을 쓰면 같은 검색어가 남아 있는 상태에서도 Enter/검색 버튼으로 다시 검색결과 화면으로 돌아갈 수 있다.
        with st.form("map_product_search_form", clear_on_submit=False):
            search_col, btn_col = st.columns([8, 1], gap="small")
            with search_col:
                term = st.text_input(
                    "제품명 검색",
                    value="",
                    placeholder="제품명/ERP명/별칭 일부 입력",
                    key="map_product_search_forced_blank" if forced_search_term else "map_product_search",
                )
            with btn_col:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                search_submitted = st.form_submit_button("검색", use_container_width=True)

    if "map_view_mode" not in st.session_state:
        st.session_state["map_view_mode"] = "search"

    term_clean = forced_search_term or (term or "").strip()
    last_term = st.session_state.get("_last_map_product_search", "")

    # 로케이션 버튼에서 넘어온 경우에는 맵을 보여주되, 검색창 값은 유지한다.
    if st.session_state.get("selected_location_from_search"):
        st.session_state["selected_location"] = st.session_state.pop("selected_location_from_search")
        st.session_state["map_view_mode"] = "map"

    # 검색어가 비면 즉시 기본 로케이션맵으로 복귀한다.
    if not term_clean:
        st.session_state["map_view_mode"] = "search"
    # 검색 버튼/Enter 또는 검색어 변경이 있으면 검색결과 화면으로 돌아간다.
    elif search_submitted or term_clean != last_term:
        st.session_state["map_view_mode"] = "search"

    st.session_state["_last_map_product_search"] = term_clean

    if term_clean and st.session_state.get("map_view_mode") != "map":
        page_map_search_results(term_clean)
    else:
        render_location_map()


def _inbound_js_loc_changed():
    """입고 도면 iframe에서 부모 페이지의 숨김 입력칸으로 넘긴 위치값을 받는다."""
    loc = str(st.session_state.get("_inbound_js_loc_buffer", "") or "").strip()
    if loc:
        st.session_state["_pending_inbound_loc"] = loc


def _apply_inbound_location_pending():
    pending = st.session_state.pop("_pending_inbound_loc", None)
    if not pending:
        try:
            qloc = st.query_params.get("inbound_loc", "")
            if isinstance(qloc, list):
                qloc = qloc[0] if qloc else ""
            pending = str(qloc or "").strip()
            if pending:
                try:
                    del st.query_params["inbound_loc"]
                except Exception:
                    pass
        except Exception:
            pending = ""
    if not pending:
        return
    if pending in ["Q1", "Q2", "Q"]:
        area, line, level = "Q", "", ""
    else:
        area, line, level = parse_location(pending)

    # 위젯 key(inbound_area/line/level)를 직접 수정하지 않고,
    # 별도 기본값 + 위젯 토큰으로 다음 렌더에서 콤보박스 값을 맞춘다.
    st.session_state["_inbound_picker_defaults"] = {"area": area or "REC", "line": line or "", "level": level or ""}
    st.session_state["_inbound_selected_loc"] = make_location(area or "REC", line or "", level or "")
    st.session_state["_inbound_picker_token"] = int(st.session_state.get("_inbound_picker_token", 0) or 0) + 1


def render_inbound_quick_location_map():
    """입고 등록용 로케이션 도면.
    로케이션맵과 같은 components.html 기반 도면을 유지하되,
    클릭은 target="_top" 링크로 부모 Streamlit 앱에 inbound_loc 파라미터를 넘긴다.
    """
    selected = st.session_state.get("_inbound_selected_loc", "") or "REC"

    def is_selected(loc):
        return selected == loc or (selected and selected.startswith(loc + "-"))

    def href(loc):
        return f"?inbound_loc={quote(loc)}"

    def cell(loc, text=None):
        text = text or loc
        cls = " selected" if is_selected(loc) else ""
        return f'<a class="map-cell{cls}" href="{href(loc)}" target="_top" data-inbound-loc="{escape(loc)}" data-loc="{escape(loc)}">{escape(text)}</a>'

    def rack(labels, left, top, cls):
        cells = ''.join(cell(x) for x in labels)
        return f'<div class="rack {cls}" style="left:{left}px;top:{top}px;">{cells}</div>'

    def zone(loc, text, left, top, w, h, cls="white", extra=""):
        selected_cls = " selected" if is_selected(loc) else ""
        return f'<a class="zone {cls}{selected_cls}" href="{href(loc)}" target="_top" data-inbound-loc="{escape(loc)}" data-loc="{escape(loc)}" style="left:{left}px;top:{top}px;width:{w}px;height:{h}px;{extra}">{text}</a>'

    html = f"""
<!doctype html><html><head><meta charset="utf-8">
<style>
*{{box-sizing:border-box}} body{{margin:0;background:#f8fafc;font-family:Inter,Segoe UI,Arial,'Noto Sans KR',sans-serif;color:#0f172a;}}
.inbound-map-card{{background:#fff;border:1px solid #dbe4f0;border-radius:18px;padding:14px 14px 18px;box-shadow:0 8px 24px rgba(15,23,42,.05);width:100%;}}
.title{{font-weight:900;font-size:18px;margin:0 0 10px;color:#111827;}}
.map-scroll{{overflow:visible;height:760px;padding:0;}}
.map-stage{{position:relative;width:1160px;height:704px;min-width:1160px;background:#fff;border-radius:14px;transform:scale(.98);transform-origin:top left;}}
.rack{{position:absolute;width:126px;height:168px;display:grid;grid-template-columns:1fr 1fr;grid-template-rows:1fr 1fr 1fr;border:1px solid #334155;border-radius:9px;overflow:hidden;box-shadow:0 6px 14px rgba(15,23,42,.06);}}
.map-cell,.zone{{appearance:none;position:relative;display:flex;align-items:center;justify-content:center;text-decoration:none;color:#0f172a;font-weight:900;font-size:14px;border:0;border-right:1px solid rgba(51,65,85,.38);border-bottom:1px solid rgba(51,65,85,.38);cursor:pointer;font-family:inherit;}}
.map-cell:hover,.zone:hover{{outline:3px solid rgba(37,99,235,.22);z-index:2;}}
.map-cell:nth-child(2n){{border-right:none;}}
.map-cell:nth-child(n+5){{border-bottom:none;}}

.special-menu{{position:absolute;display:none;z-index:30;background:#fff;border:1px solid #cbd5e1;border-radius:12px;box-shadow:0 12px 28px rgba(15,23,42,.18);padding:6px;}}
.special-menu.open{{display:grid;gap:5px;}}
.special-menu button,.special-menu a{{appearance:none;border:1px solid #e2e8f0;background:#f8fafc;border-radius:9px;padding:8px 7px;font-size:12px;font-weight:900;color:#0f172a;cursor:pointer;font-family:inherit;text-align:center;text-decoration:none;}}
.special-menu button:hover,.special-menu button.selected,.special-menu a:hover,.special-menu a.selected{{background:#22c55e;color:#fff;border-color:#16a34a;}}
.map-cell.selected,.zone.selected{{background:#22c55e!important;color:#ffffff!important;outline:3px solid rgba(34,197,94,.35)!important;box-shadow:0 0 0 3px rgba(255,255,255,.8),0 0 18px rgba(34,197,94,.75)!important;border-color:#16a34a!important;z-index:4;}}
.yellow{{background:#fff39b;}} .blue{{background:#68d2e7;}} .pink{{background:#f0a7e6;}} .gray{{background:#f7f8fa;}} .bidata{{background:#d1d5db;}} .white{{background:#fff;}}
.yellow .map-cell,.zone.yellow{{background:#fff39b;}} .blue .map-cell,.zone.blue{{background:#68d2e7;}} .pink .map-cell,.zone.pink{{background:#f0a7e6;}} .gray .map-cell,.zone.gray{{background:#f7f8fa;}} .bidata .map-cell,.zone.bidata{{background:#d1d5db;}} .white .map-cell,.zone.white{{background:#fff;}}
.zone{{position:absolute;border:1px solid #334155;border-radius:9px;box-shadow:0 6px 14px rgba(15,23,42,.04);}}
.big-left{{position:absolute;left:0;top:0;width:185px;height:282px;border:1px solid #334155;border-radius:10px;overflow:hidden;background:#fff;}}
.big-left a{{appearance:none;position:relative;display:flex;align-items:center;justify-content:center;width:100%;border:0;border-bottom:1px solid #cbd5e1;background:#f7f8fa;color:#0f172a;font-weight:900;cursor:pointer;font-family:inherit;text-decoration:none;}}
.big-left a:hover{{outline:3px solid rgba(37,99,235,.22);z-index:2;}}
.big-left a.selected{{background:#22c55e!important;color:#fff!important;outline:3px solid rgba(34,197,94,.35)!important;box-shadow:0 0 0 3px rgba(255,255,255,.8),0 0 18px rgba(34,197,94,.75)!important;z-index:4;}}
.g2{{height:225px;background:#f7f8fa;}} .g1row{{height:57px;display:grid;grid-template-columns:1fr 1fr 1fr;}}
.g1row a{{height:57px;border-right:1px solid #cbd5e1;border-bottom:none;}} .g1row a:last-child{{border-right:none;}}
.label{{position:absolute;text-align:center;font-weight:900;color:#111827;font-size:14px;}}
.memo{{position:absolute;color:#334155;font-size:15px;line-height:1.65;}}
.qp{{position:absolute;left:0;top:525px;width:165px;height:148px;border:1px solid #cbd5e1;border-radius:10px;overflow:hidden;background:#fff;}}
.qp a{{position:relative;display:grid;grid-template-columns:58px 1fr;align-items:center;width:100%;height:74px;border:0;border-bottom:1px solid #e2e8f0;background:#fff;color:#111827;font-weight:900;cursor:pointer;text-align:left;font-family:inherit;text-decoration:none;}}
.qp a:last-child{{border-bottom:none;}}
.qp a.selected{{background:#22c55e!important;color:#ffffff!important;outline:3px solid rgba(34,197,94,.35)!important;box-shadow:0 0 0 3px rgba(255,255,255,.8),0 0 18px rgba(34,197,94,.75)!important;border-color:#16a34a!important;z-index:4;}}
.qp-key{{height:100%;display:flex;align-items:center;justify-content:center;color:#ff221a;font-weight:900;font-size:18px;border-right:1px solid #e2e8f0;}}
.qp .qkey{{background:#f186ca;color:#ff0d0d;}}
.rec-red{{color:#ff1e12;font-weight:900;}}
.small-title{{position:absolute;font-size:14px;font-weight:900;color:#111827;text-align:center;}}
</style></head><body>
<div class="inbound-map-card">
  <div class="title">도면에서 입고 위치 선택</div>
  <div class="map-scroll"><div class="map-stage">
    <div class="big-left">
      <a class="g2 gray{' selected' if is_selected('G2') else ''}" href="?inbound_loc=G2" target="_top" data-inbound-loc="G2">G2</a>
      <div class="g1row">
        {cell('G1-01')}{cell('G1-02')}{cell('G1-03')}
      </div>
    </div>
    {rack(['A2-03','A2-04','A2-02','A2-05','A2-01','A2-06'],230,0,'yellow')}
    {rack(['B2-03','B2-04','B2-02','B2-05','B2-01','B2-06'],372,0,'yellow')}
    {rack(['C2-03','C2-04','C2-02','C2-05','C2-01','C2-06'],514,0,'blue')}
    {rack(['D1-03','D1-04','D1-02','D1-05','D1-01','D1-06'],656,0,'blue')}
    {zone('T1','T1',656,168,126,52,'white')}
    {rack(['E1-03','E1-04','E1-02','E1-05','E1-01','E1-06'],798,0,'pink')}
    {zone('T2','T2',798,168,126,52,'white')}
    {zone('F1-01','F1-01',955,0,64,52,'bidata')}
    {zone('F1-02','F1-02',1019,0,64,52,'bidata')}
    {zone('F1-03','F1-03',1083,0,64,52,'bidata')}
    <div class="small-title" style="left:996px;top:72px;width:110px;">비자료</div>
    {zone('X2','X2',1070,78,70,52,'gray')}
    {rack(['A1-03','A1-04','A1-02','A1-05','A1-01','A1-06'],230,268,'yellow')}
    {rack(['B1-03','B1-04','B1-02','B1-05','B1-01','B1-06'],372,268,'yellow')}
    {rack(['C1-03','C1-04','C1-02','C1-05','C1-01','C1-06'],514,268,'yellow')}
    <div class="memo" style="left:800px;top:292px;">X1-01~03 : 폐기<br>X1-01-01 : 대표님 시술용</div>
    {zone('X1-01','X1-01',1090,268,64,56,'gray')}
    {zone('X1-02','X1-02',1090,324,64,56,'gray')}
    {zone('X1-03','X1-03',1090,380,64,56,'gray')}
    <div class="qp">
      <a class="{'selected' if is_selected('Q1') or is_selected('Q2') else ''}" href="?inbound_loc=Q" target="_top" data-inbound-loc="Q"><span class="qp-key qkey">Q</span><span>유통기간임박</span></a>
      <a class="{'selected' if is_selected('P') else ''}" href="?inbound_loc=P" target="_top" data-inbound-loc="P"><span class="qp-key">P</span><span>수출대기</span></a>
    </div>
    {zone('REC','<span><span class="rec-red">REC</span>eiving</span>',372,568,142,56,'white')}
    <div class="label" style="left:372px;top:635px;width:142px;">매입등록대기</div>
    {zone('R2','R2',790,460,64,56,'white')}
    {zone('R1','R1',854,460,64,56,'white')}
    <div class="label" style="left:770px;top:526px;width:190px;">R2 비자료 / R1 자료</div>
    {zone('N','기타 위치',975,628,168,60,'white')}
    <div class="special-menu" id="inboundSpecialMenu" style="left:975px;top:492px;width:168px;"><a href="?inbound_loc=홍보물랙" target="_top" data-inbound-loc="홍보물랙">홍보물랙</a><a href="?inbound_loc=회색 카트" target="_top" data-inbound-loc="회색 카트">회색 카트</a><a href="?inbound_loc=오른쪽 창고" target="_top" data-inbound-loc="오른쪽 창고">오른쪽 창고</a><a href="?inbound_loc=사무실(4층)" target="_top" data-inbound-loc="사무실(4층)">사무실(4층)</a></div>
  </div></div>
</div>
<script>
function parentBaseHref(){{
  try {{ return window.top.location.href; }} catch(e) {{}}
  try {{ return window.parent.location.href; }} catch(e) {{}}
  return document.referrer || window.location.href;
}}
function buildParentUrl(key, value){{
  const url = new URL(parentBaseHref());
  url.searchParams.set(key, value);
  url.searchParams.delete('map_search_product');
  return url.toString();
}}
function navigateTop(href){{
  try {{ window.top.location.assign(href); return; }} catch(e) {{}}
  try {{ window.parent.location.assign(href); return; }} catch(e) {{}}
  try {{ window.open(href, '_top'); return; }} catch(e) {{}}
  const a = document.createElement('a');
  a.href = href;
  a.target = '_top';
  document.body.appendChild(a);
  a.click();
}}
function markSelected(loc){{
  document.querySelectorAll('[data-inbound-loc]').forEach(x => {{
    const v = x.getAttribute('data-inbound-loc') || '';
    x.classList.toggle('selected', v === loc || (loc && loc.startsWith(v + '-')) || (v === 'N' && ['홍보물랙','회색 카트','오른쪽 창고','사무실(4층)'].includes(loc)));
  }});
}}
function tryParentInbound(loc){{
  try {{
    const doc = window.parent.document;
    // 핵심: Streamlit text_input 값 확정 타이밍에 의존하지 않는다.
    // 먼저 부모 URL의 query parameter에 클릭 위치를 심고, 숨김 버튼은 rerun 트리거로만 사용한다.
    try {{
      const url = new URL(parentBaseHref());
      url.searchParams.set('inbound_loc', loc);
      url.searchParams.delete('map_search_product');
      if (window.parent && window.parent.history && window.parent.history.replaceState) {{
        window.parent.history.replaceState(null, '', url.toString());
      }} else if (window.top && window.top.history && window.top.history.replaceState) {{
        window.top.history.replaceState(null, '', url.toString());
      }}
    }} catch(e) {{}}

    const input = doc.querySelector('input[aria-label="__입고도면선택값"]');
    if (input) {{
      try {{
        const setter = Object.getOwnPropertyDescriptor(window.parent.HTMLInputElement.prototype, 'value').set;
        setter.call(input, loc);
      }} catch(e) {{
        input.value = loc;
      }}
      input.dispatchEvent(new InputEvent('input', {{bubbles:true, inputType:'insertText', data:loc}}));
      input.dispatchEvent(new Event('change', {{bubbles:true}}));
    }}

    let applied = false;
    doc.querySelectorAll('button').forEach(btn => {{
      if ((btn.innerText || '').trim() === '__입고도면적용') {{
        setTimeout(() => btn.click(), 30);
        applied = true;
      }}
    }});
    return applied;
  }} catch(e) {{
    return false;
  }}
}}
function toggleInboundSpecialMenu(forceClose=false){{
  const menu=document.getElementById('inboundSpecialMenu');
  if(!menu) return;
  if(forceClose){{menu.classList.remove('open'); return;}}
  menu.classList.toggle('open');
}}
document.querySelectorAll('[data-inbound-loc]').forEach(el => {{
  el.addEventListener('click', (ev) => {{
    ev.preventDefault();
    ev.stopPropagation();
    const loc = el.getAttribute('data-inbound-loc') || '';
    if(!loc) return;
    if(loc === 'N'){{
      markSelected('N');
      toggleInboundSpecialMenu(false);
      return;
    }}
    toggleInboundSpecialMenu(true);
    markSelected(loc);
    if(tryParentInbound(loc)) return;
    const href = buildParentUrl('inbound_loc', loc);
    el.setAttribute('href', href);
    el.setAttribute('target', '_top');
    navigateTop(href);
  }});
}});
</script>
</body></html>
"""
    components.html(html, height=780, scrolling=False)



def page_inbound():
    _apply_inbound_location_pending()
    st.title("입고 등록")
    
    # 입고 도면 클릭값은 query parameter(inbound_loc)로 직접 처리한다.
    # "__입고도면적용" 버튼/숨김 입력창은 생성하지 않는다.
    # 버튼을 만들면 Streamlit이 먼저 렌더링해서 화면에 순간 노출된다.
    
    _apply_inbound_location_pending()
    inbound_product_term = st.text_input("제품 검색", placeholder="제품명, ERP명, 비자료명, 별칭 일부 입력", key="inbound_product_term")
    products = product_options(inbound_product_term)
    product_list = products["standard_name"].dropna().astype(str).tolist() if not products.empty else []

    def inbound_product_label(value):
        # 검색은 표준제품명/ERP명/비자료명/별칭 전체를 대상으로 하되,
        # 콤보박스에는 현장 혼선을 줄이기 위해 표준제품명만 표시한다.
        if value == "":
            return "제품명을 입력하거나 선택하세요"
        return str(value)

    top_left, top_right = st.columns(2, gap="large")
    with top_left:
        in_src_col, in_company_col = st.columns(2, gap="small")
        with in_src_col:
            inbound_source = st.text_input("매입처", value="", placeholder="예: 거래처명/수입처", key="inbound_source")
        with in_company_col:
            _inbound_selected_product_for_stock = st.session_state.get("inbound_product", "")
            company_label = st.selectbox("사업장", inbound_company_options_for(_inbound_selected_product_for_stock), key="inbound_company")
            company = strip_company_stock_label(company_label)
        first_product_state = bool(st.session_state.get("inbound_first_product", False))
        selected_product = st.selectbox(
            "제품",
            [""] + product_list,
            index=0,
            key="inbound_product",
            format_func=inbound_product_label,
            disabled=first_product_state,
        )
        first_product = st.checkbox("최초 등록", key="inbound_first_product")
        if first_product:
            product = st.text_input("제품명 직접 입력", value="", placeholder="신규 표준제품명 입력", key="inbound_new_product_name").strip()
        else:
            product = selected_product
        wh = product_mapping_name_for(company, product) or product
    with top_right:
        lot = st.text_input("LOT/제조번호", value="", placeholder="미입력 시 '-' 저장", key="inbound_lot")
        exp = st.text_input("유통기한", value="", placeholder="예: 28/3/2, 28.3.2, 2028-03-02 / 미입력 시 '-' 저장", key="inbound_exp")

    st.markdown("---")
    map_col, pos_col = st.columns([7.9, 2.1], gap="large")
    with map_col:
        render_inbound_quick_location_map()
    with pos_col:
        st.markdown("#### 입고 위치")
        loc = inbound_location_picker("REC")
        qty = st.number_input("수량", min_value=1, step=1, key="inbound_qty")
        memo = st.text_input("메모", value="", key="inbound_memo")
        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
        _save_left, save_col, _save_right = st.columns([1, 2, 1])
        with save_col:
            save_clicked = st.button("입고 저장", type="primary", use_container_width=True)
        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
        save_msg = st.empty()
        if save_clicked:
            if not product:
                save_msg.error("제품을 선택하세요.")
            else:
                if first_product:
                    product = ensure_standard_product_only(product)
                    wh = product
                memo_parts = []
                if inbound_source:
                    memo_parts.append(f"매입처: {inbound_source}")
                if memo:
                    memo_parts.append(memo)
                inbound_memo = " / ".join(memo_parts) if memo_parts else "입고 등록"
                add_inventory(company, product, wh, normalize_blank(lot), normalize_exp_date(exp), loc, int(qty), inbound_memo)
                save_msg.success(f"입고 저장 완료: {company} / {product} / {loc} / {qty}EA")


def page_move():
    st.title("이동 등록")
    st.caption("제품 → LOT/유통기한을 선택하면 출발 재고가 자동 표시됩니다.")
    term = st.text_input("제품 검색", placeholder="제품명, 전산상 명칭, 별칭 일부 입력")
    opts = product_options(term)
    if opts.empty:
        st.warning("일치하는 제품이 없습니다."); return
    product = st.selectbox("추천 제품", [""] + opts["standard_name"].tolist(), index=0, format_func=lambda x: "제품명을 입력하거나 선택하세요" if x == "" else x)
    if not product:
        st.info("이동할 제품을 선택하세요.")
        return
    lot_df = q("SELECT DISTINCT lot FROM inventory WHERE product_name=? AND qty>0 ORDER BY lot", (product,))
    if lot_df.empty:
        st.info("현재 재고가 0이 아닌 LOT/제조번호가 없습니다."); return
    lot = st.selectbox("LOT/제조번호", lot_df["lot"].tolist())
    exp_df = q("SELECT DISTINCT exp_date FROM inventory WHERE product_name=? AND lot=? AND qty>0 ORDER BY exp_date", (product, lot))
    exp = st.selectbox("유통기한", exp_df["exp_date"].tolist())
    src_df = q("""SELECT id, company AS 출발사업장, location AS 출발위치, qty AS 현재수량, warehouse_name AS 전산상명칭
                  FROM inventory WHERE product_name=? AND lot=? AND exp_date=? AND qty>0 ORDER BY company, location""", (product, lot, exp))
    left, right = st.columns(2, gap="large")
    with left:
        st.markdown("#### 출발 재고")
        if len(src_df) == 1:
            src_id = int(src_df.iloc[0]["id"])
            r = src_df.iloc[0]
            st.info(f"{r['출발사업장']} / {r['출발위치']} / 현재 {int(r['현재수량'])}EA")
        else:
            labels = [f"{r.출발사업장} / {r.출발위치} / {r.현재수량}EA" for r in src_df.itertuples()]
            selected = st.selectbox("출발 재고 선택", labels)
            src_id = int(src_df.iloc[labels.index(selected)]["id"])
        src_row = src_df[src_df["id"]==src_id].iloc[0]
        src_company = str(src_row["출발사업장"])
        max_qty = int(src_row["현재수량"])
        st.dataframe(src_df.drop(columns=["id"]), use_container_width=True, hide_index=True)
    with right:
        st.markdown("#### 도착 재고")
        default_idx = COMPANIES.index(src_company) if src_company in COMPANIES else 0
        to_company = st.selectbox("도착 사업장", COMPANIES, index=default_idx, key=f"move_company_{src_id}")
        if to_company != src_company:
            st.warning("정말로 다른 사업장으로 재고를 이동하시겠습니까?")
        to_location = location_picker("move", "A1")
        qty = st.number_input("이동 수량", min_value=1, max_value=max_qty, value=min(1,max_qty), step=1)
        memo = st.text_input("메모", value="")
        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
        if st.button("이동 저장", type="primary", use_container_width=True):
            try:
                move_inventory(src_id, to_company, to_location, int(qty), memo)
                st.success(f"이동 저장 완료: {product} / {qty}EA → {to_company} {to_location}")
                st.rerun()
            except Exception as e:
                st.error(str(e))



def recommend_picks(pick_df, request_qty):
    """출고 요청 수량만큼 재고 행을 추천한다.
    유통기한이 빠른 것, 로케이션, LOT 순으로 선택한다.
    반환: (추천행 list, 부족수량)
    """
    rows = []
    try:
        need = int(request_qty or 0)
    except Exception:
        need = 0
    if need <= 0 or pick_df is None or pick_df.empty:
        return rows, max(0, need)

    df = pick_df.copy()
    df["_exp_sort"] = pd.to_datetime(df.get("exp_date"), errors="coerce")
    df["_exp_sort"] = df["_exp_sort"].fillna(pd.Timestamp.max)
    df = df.sort_values(["_exp_sort", "location", "lot", "company"], na_position="last")

    for r in df.itertuples():
        if need <= 0:
            break
        available = int(getattr(r, "qty", 0) or 0)
        if available <= 0:
            continue
        take = min(available, need)
        rows.append({
            "id": int(getattr(r, "id")),
            "로케이션": getattr(r, "location", ""),
            "사업장": getattr(r, "company", ""),
            "제품명": getattr(r, "product_name", ""),
            "LOT": getattr(r, "lot", "-") or "-",
            "유통기한": display_date_only(getattr(r, "exp_date", "-")),
            "요청수량": int(take),
        })
        need -= take
    return rows, max(0, need)




def _clear_outbound_inputs_before_render():
    """출고지시 저장/수정 완료 후 다음 렌더에서 입력 위젯 값을 초기화한다.
    Streamlit widget key를 생성된 뒤 직접 수정하지 않기 위해 page_outbound 시작부에서만 실행한다.
    """
    if not st.session_state.pop("_outbound_reset_inputs_pending", False):
        return
    for k in [
        "out_customer_term", "out_customer_select", "out_selected_customer",
        "out_product_term", "out_req_qty", "out_rec_editor",
        "outbound_cart", "editing_order_id", "editing_order_title",
        "pending_outbound_save", "pending_outbound_expiry_warnings",
    ]:
        st.session_state.pop(k, None)
    st.session_state["out_cart_editor_token"] = int(st.session_state.get("out_cart_editor_token", 0) or 0) + 1

def get_cart():
    """출고지시 장바구니를 안전하게 반환한다.
    리팩토링 중 함수 누락으로 page_outbound가 깨지는 것을 막기 위해 app.py 내부에 유지한다.
    """
    cart = st.session_state.get("outbound_cart")
    if not isinstance(cart, list):
        st.session_state["outbound_cart"] = []
    return st.session_state["outbound_cart"]


def _cart_expiry_warnings(cart):
    """출고지시 장바구니 투입 전 유통기한 경고 목록을 만든다. 만료 또는 30일 미만 남은 품목."""
    warnings = []
    today = date.today()
    for item in cart or []:
        exp = display_date_only(item.get("유통기한"))
        if not exp or exp == "-":
            continue
        try:
            d = datetime.strptime(exp, "%Y-%m-%d").date()
        except Exception:
            continue
        days = (d - today).days
        if days < 0:
            status = "유통기한 만료"
        elif days < 30:
            status = f"유통기한 {days}일 남음"
        else:
            continue
        warnings.append({
            "제품명": item.get("제품명", "-"),
            "LOT": item.get("LOT", "-"),
            "유통기한": exp,
            "수량": item.get("요청수량", ""),
            "상태": status,
        })
    return warnings



def _add_rows_to_outbound_cart(rows):
    """추천 행을 출고지시 장바구니에 추가한다.
    같은 로케이션/사업장/제품/LOT/유통기한이면 수량을 합산한다.
    """
    cart = get_cart()
    added = 0
    for row in rows or []:
        qty2 = int(row.get("요청수량", 0) or 0)
        if qty2 <= 0:
            continue
        key = (row.get("로케이션"), row.get("사업장"), row.get("제품명"), row.get("LOT"), row.get("유통기한"))
        merged = False
        for existing in cart:
            ekey = (existing.get("로케이션"), existing.get("사업장"), existing.get("제품명"), existing.get("LOT"), existing.get("유통기한"))
            if ekey == key:
                existing["요청수량"] = int(existing.get("요청수량", 0) or 0) + qty2
                merged = True
                break
        if not merged:
            cart.append({
                "id": int(row.get("id")),
                "로케이션": row.get("로케이션", ""),
                "사업장": row.get("사업장", ""),
                "제품명": row.get("제품명", ""),
                "LOT": row.get("LOT", "-"),
                "유통기한": row.get("유통기한", "-"),
                "요청수량": qty2,
            })
        added += 1
    return added

def _save_outbound_cart_action(cart, title):
    """장바구니 출고지시 저장/수정 공통 처리."""
    if st.session_state.get("editing_order_id"):
        update_outbound_order(st.session_state["editing_order_id"], title, cart)
        msg = f"출고지시서 #{st.session_state['editing_order_id']} 수정 저장 완료"
        st.session_state.pop("editing_order_id", None)
        st.session_state.pop("editing_order_title", None)
    else:
        oid = save_outbound_order(cart, title)
        msg = f"출고지시서 #{oid} 저장 완료"
    st.session_state["outbound_cart"] = []
    st.session_state["_outbound_reset_inputs_pending"] = True
    st.session_state["_outbound_last_success"] = msg
    st.rerun()

def page_outbound():
    _clear_outbound_inputs_before_render()
    st.title("출고지시")
    st.caption("출고지시 저장 시 해당 제조번호/유통기한/로케이션의 현재고가 즉시 차감됩니다.")
    last_msg = st.session_state.pop("_outbound_last_success", None)
    if last_msg:
        st.success(last_msg)

    st.markdown("### 매출처")
    cust_term = st.text_input("매출처", placeholder="거래처명을 입력하세요", key="out_customer_term")
    cust_df = pd.DataFrame()
    if cust_term.strip():
        like = f"%{cust_term.strip()}%"
        cust_df = q("""SELECT * FROM customers WHERE customer_name LIKE ? ORDER BY customer_name LIMIT 30""", (like,))
    else:
        cust_df = q("""SELECT * FROM customers ORDER BY customer_name LIMIT 30""")
    selected_customer = None
    if not cust_df.empty:
        labels = [f"{r.customer_name} | {r.company or '-'}" for r in cust_df.itertuples()]
        label = st.selectbox("거래처 선택", labels, key="out_customer_select")
        selected_customer = cust_df.iloc[labels.index(label)]
        st.session_state["out_selected_customer"] = selected_customer.to_dict()
        st.markdown(f"**사업장 :** {selected_customer.get('company') or '-'} &nbsp;&nbsp;&nbsp; **유형 :** {selected_customer.get('customer_type') or '-'} &nbsp;&nbsp;&nbsp; **담당자 :** {selected_customer.get('manager') or '-'}")
        with st.expander("거래처 상세정보", expanded=False):
            st.write(f"주소 : {selected_customer.get('address') or '-'}")
            st.write(f"연락처 : {selected_customer.get('phone') or '-'}")
    else:
        st.info("거래처를 검색하거나 거래처 관리에서 먼저 등록하세요.")
    st.markdown("---")
    st.markdown("""
    <style>
      /* 출고지시 상단 카드: 총재고 숫자와 출고 요청 수량 입력값의 시각 크기를 맞춤 */
      div[data-testid="stMetricValue"] {font-size: 2.35rem; text-align:center;}
      div[data-testid="stMetricLabel"] {text-align:center; width:100%; display:flex; justify-content:center;}
      div[data-testid="stMetric"] label, div[data-testid="stMetric"] [data-testid="stMetricLabel"] {width:100%; justify-content:center; text-align:center;}
      div[data-testid="stNumberInput"] input {font-size: 2.15rem !important; font-weight: 600 !important; height: 3.25rem !important; text-align:center !important; padding-left:19px !important;}
      .out-req-label {font-size: 0.92rem; color: #64748b; margin: 0 0 0.25rem 0; text-align:left !important; width:100%; display:block;}
    </style>
    """, unsafe_allow_html=True)

    left, right = st.columns([1, 1], gap="large")

    selected_product = None
    pick_df = pd.DataFrame()
    req = 1
    rule = "유통기한 짧은 것 먼저"
    company_filter = None

    with left:
        st.markdown("### 제품 선택")
        term = st.text_input("제품 검색", placeholder="제품명/전산상 명칭/별칭을 입력하세요", key="out_product_term")
        opts = product_options(term)
        if opts.empty:
            st.info("제품을 검색하세요.")
        else:
            selected_product = st.selectbox("제품 선택", opts["standard_name"].tolist())
            df = q("SELECT * FROM inventory WHERE product_name=? AND qty>0", (selected_product,))
            total_qty = int(df.qty.sum()) if not df.empty else 0

            total_col, req_col = st.columns([1, 1], gap="medium")
            with total_col:
                st.metric("총재고", f"{total_qty} EA")
            with req_col:
                st.markdown('<div class="out-req-label">출고 요청 수량</div>', unsafe_allow_html=True)
                req = st.number_input("출고 요청 수량", min_value=1, step=1, key="out_req_qty", label_visibility="collapsed")

            rule = st.radio("추천 기준", ["유통기한 짧은 것 먼저", "특정 사업장만"], horizontal=True)
            if rule == "특정 사업장만":
                company_filter = st.selectbox("사업장 선택", ["노투스팜", "NOH", "노투스"])

    with right:
        st.markdown("### 현재 출고 가능 재고")
        if selected_product:
            df = q("SELECT * FROM inventory WHERE product_name=? AND qty>0", (selected_product,))
            if df.empty:
                st.warning("출고 지시 가능한 재고가 없습니다.")
            else:
                view = df[["company", "lot", "exp_date", "location", "qty"]].copy()
                view = view.rename(columns={"company":"사업장", "lot":"LOT", "exp_date":"유통기한", "location":"로케이션", "qty":"수량"})
                view = view.sort_values(["사업장", "LOT", "유통기한", "로케이션"])
                st.dataframe(view, hide_index=True, use_container_width=True)
                pick_df = df.copy()
                if company_filter:
                    pick_df = pick_df[pick_df["company"] == company_filter]
                    if pick_df.empty:
                        st.warning(f"{company_filter}에 출고 가능한 재고가 없습니다.")

    st.markdown("---")

    st.markdown("### 이번 품목 출고 추천")
    rec_table_col, _ = st.columns([7, 3], gap="large")
    with rec_table_col:
        if selected_product and not pick_df.empty and req:
            available = int(pick_df["qty"].sum())
            if available < int(req):
                st.error(f"재고 부족: 요청 {int(req)}EA / 가능 {available}EA / 부족 {int(req)-available}EA")
            rec_rows, shortage = recommend_picks(pick_df, int(req))
            if rec_rows:
                rec = pd.DataFrame(rec_rows)
                rec_display_cols = ["로케이션", "사업장", "제품명", "LOT", "유통기한", "요청수량"]
                edited = st.data_editor(
                    rec[rec_display_cols],
                    hide_index=True,
                    use_container_width=True,
                    num_rows="fixed",
                    disabled=["로케이션", "사업장", "제품명", "LOT", "유통기한"],
                    column_config={"요청수량": st.column_config.NumberColumn("요청수량", min_value=0, step=1)},
                    key="out_rec_editor",
                )
                if st.button("장바구니에 담기", type="primary", use_container_width=True):
                    pending_rows = []
                    for idx, row in rec.iterrows():
                        qty2 = int(edited.iloc[idx]["요청수량"] or 0)
                        if qty2 > 0:
                            pending_rows.append({
                                "id": int(row["id"]),
                                "로케이션": row["로케이션"],
                                "사업장": row["사업장"],
                                "제품명": row["제품명"],
                                "LOT": row["LOT"],
                                "유통기한": row["유통기한"],
                                "요청수량": qty2,
                            })
                    warn_rows = _cart_expiry_warnings(pending_rows)
                    if warn_rows:
                        st.session_state["pending_outbound_add_rows"] = pending_rows
                        st.session_state["pending_outbound_add_warnings"] = warn_rows
                        st.rerun()
                    else:
                        added = _add_rows_to_outbound_cart(pending_rows)
                        st.success(f"{added}개 행을 출고지시 장바구니에 담았습니다.")
                        st.rerun()
            else:
                st.info("추천할 재고가 없습니다.")
        else:
            st.info("제품과 출고 요청 수량을 입력하면 추천이 표시됩니다.")

    st.markdown("### 출고지시 장바구니")
    cart_table_col, _ = st.columns([7, 3], gap="large")
    with cart_table_col:
        cart = get_cart()
        if not cart:
            st.info("아직 담긴 품목이 없습니다. 제품을 검색해서 장바구니에 담으세요.")
        else:
            cart_df = pd.DataFrame(cart)
            for c in ["로케이션", "사업장", "제품명", "LOT", "유통기한", "요청수량"]:
                if c not in cart_df.columns:
                    cart_df[c] = ""
            display_cols = ["로케이션", "사업장", "제품명", "LOT", "유통기한", "요청수량"]
            edited_cart = st.data_editor(
                cart_df[display_cols],
                hide_index=True,
                use_container_width=True,
                num_rows="dynamic",
                disabled=["로케이션", "사업장", "제품명", "LOT", "유통기한"],
                column_config={"요청수량": st.column_config.NumberColumn("요청수량", min_value=0, step=1)},
                key=f"out_cart_editor_{st.session_state.get('out_cart_editor_token', 0)}",
            )
            new_cart = []
            for i in range(min(len(cart), len(edited_cart))):
                item = dict(cart[i])
                item["요청수량"] = int(edited_cart.iloc[i]["요청수량"] or 0)
                if item["요청수량"] > 0:
                    new_cart.append(item)
            if len(new_cart) != len(cart) or any(int(a.get("요청수량", 0)) != int(b.get("요청수량", 0)) for a, b in zip(new_cart, cart)):
                st.session_state["outbound_cart"] = new_cart
                st.session_state["out_cart_editor_token"] = int(st.session_state.get("out_cart_editor_token", 0) or 0) + 1
                st.rerun()
            customer_name = st.session_state.get("out_selected_customer", {}).get("customer_name", "")
            fallback_title = datetime.now().strftime("출고지시 %Y-%m-%d %H:%M")
            default_title = st.session_state.get("editing_order_title") or build_outbound_order_title(customer_name, cart, fallback_title)
            title = st.text_input("출고지시서 제목", value=default_title, placeholder="예: A병원 디센바(1V) 외 2품목")
            b1, b2 = st.columns(2)
            with b1:
                if st.button("지시완료 저장", type="primary", use_container_width=True):
                    try:
                        _save_outbound_cart_action(cart, title)
                    except Exception as e:
                        st.error(str(e))
            with b2:
                if st.button("장바구니 비우기", use_container_width=True):
                    st.session_state["outbound_cart"] = []
                    st.session_state.pop("editing_order_id", None)
                    st.session_state.pop("editing_order_title", None)
                    st.rerun()
            dl1, dl2 = st.columns(2)
            with dl1:
                st.download_button(
                    "출고지시서 엑셀 다운로드",
                    data=outbound_excel_bytes(cart, title or "출고지시서"),
                    file_name=f"NOHTUS_출고지시서_{date.today().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with dl2:
                pdf_data = outbound_pdf_bytes(cart, title or "출고지시서")
                st.download_button(
                    "출고지시서 PDF 다운로드",
                    data=pdf_data,
                    file_name=f"NOHTUS_출고지시서_{date.today().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )



    if st.session_state.get("pending_outbound_add_rows"):
        warn_rows = st.session_state.get("pending_outbound_add_warnings", [])
        pending_rows = st.session_state.get("pending_outbound_add_rows", [])
        dialog_api = getattr(st, "dialog", None) or getattr(st, "experimental_dialog", None)
        if dialog_api:
            @dialog_api("⚠ 유통기한 경고")
            def _confirm_expiry_add_dialog():
                st.markdown("""
                <style>
                div[data-testid="stDialog"] div[data-testid="stButton"] > button{
                    min-height:46px!important;
                    min-width:180px!important;
                    border-radius:10px!important;
                    font-weight:800!important;
                    white-space:nowrap!important;
                }
                </style>
                <div style='font-size:16px;line-height:1.7;color:#334155;margin:6px 0 14px 0;font-weight:400;'>
                    유통기한이 만료되었거나 1개월 미만 남은 품목입니다.<br>
                    그래도 출고지시 장바구니에 담으시겠습니까?
                </div>
                """, unsafe_allow_html=True)
                st.dataframe(pd.DataFrame(warn_rows), hide_index=True, use_container_width=True)
                _left, c1, c2, _right = st.columns([1.0, 1.2, 1.7, 1.0], gap="medium")
                with c1:
                    if st.button("아니오", use_container_width=True, key="add_expiry_no"):
                        st.session_state.pop("pending_outbound_add_rows", None)
                        st.session_state.pop("pending_outbound_add_warnings", None)
                        st.rerun()
                with c2:
                    if st.button("예, 담습니다", type="primary", use_container_width=True, key="add_expiry_yes"):
                        added = _add_rows_to_outbound_cart(pending_rows)
                        st.session_state.pop("pending_outbound_add_rows", None)
                        st.session_state.pop("pending_outbound_add_warnings", None)
                        st.session_state["_outbound_last_success"] = f"{added}개 행을 출고지시 장바구니에 담았습니다."
                        st.rerun()
            _confirm_expiry_add_dialog()
        else:
            st.warning("유통기한이 만료되었거나 1개월 미만 남은 품목입니다. 그래도 출고지시 장바구니에 담으시겠습니까?")
            st.dataframe(pd.DataFrame(warn_rows), hide_index=True, use_container_width=True)
            c1, c2 = st.columns(2)
            with c1:
                if st.button("아니오", key="add_expiry_no_inline"):
                    st.session_state.pop("pending_outbound_add_rows", None)
                    st.session_state.pop("pending_outbound_add_warnings", None)
                    st.rerun()
            with c2:
                if st.button("예, 담습니다", type="primary", key="add_expiry_yes_inline"):
                    added = _add_rows_to_outbound_cart(pending_rows)
                    st.session_state.pop("pending_outbound_add_rows", None)
                    st.session_state.pop("pending_outbound_add_warnings", None)
                    st.success(f"{added}개 행을 출고지시 장바구니에 담았습니다.")
                    st.rerun()


    if st.session_state.get("pending_outbound_save"):
        warn_rows = st.session_state.get("pending_outbound_expiry_warnings", [])
        if hasattr(st, "dialog"):
            @st.dialog("유통기한 경고")
            def _confirm_expiry_save_dialog():
                st.warning("유통기한이 지났거나 30일 미만 남은 품목이 포함되어 있습니다.")
                st.dataframe(pd.DataFrame(warn_rows), hide_index=True, use_container_width=True)
                st.write("그래도 출고지시서를 저장하시겠습니까?")
                c1, c2 = st.columns(2)
                with c1:
                    if st.button("저장 계속", type="primary", use_container_width=True):
                        pending_title = st.session_state.get("pending_outbound_save", {}).get("title", title)
                        st.session_state.pop("pending_outbound_save", None)
                        st.session_state.pop("pending_outbound_expiry_warnings", None)
                        try:
                            _save_outbound_cart_action(get_cart(), pending_title)
                        except Exception as e:
                            st.error(str(e))
                with c2:
                    if st.button("취소", use_container_width=True):
                        st.session_state.pop("pending_outbound_save", None)
                        st.session_state.pop("pending_outbound_expiry_warnings", None)
                        st.rerun()
            _confirm_expiry_save_dialog()
        else:
            st.warning("유통기한이 지났거나 30일 미만 남은 품목이 포함되어 있습니다.")
            st.dataframe(pd.DataFrame(warn_rows), hide_index=True, use_container_width=True)
            c1, c2 = st.columns(2)
            with c1:
                if st.button("저장 계속", type="primary", use_container_width=True):
                    pending_title = st.session_state.get("pending_outbound_save", {}).get("title", title)
                    st.session_state.pop("pending_outbound_save", None)
                    st.session_state.pop("pending_outbound_expiry_warnings", None)
                    _save_outbound_cart_action(get_cart(), pending_title)
            with c2:
                if st.button("저장 취소", use_container_width=True):
                    st.session_state.pop("pending_outbound_save", None)
                    st.session_state.pop("pending_outbound_expiry_warnings", None)
                    st.rerun()



def _status_text_html(status):
    status = str(status or "저장됨")
    color = "#475569"
    if status == "취소됨":
        color = "#dc2626"
    elif status == "수정됨":
        color = "#65a30d"
    return f"<span style='font-weight:400;color:{color};'>{escape(status)}</span>"


def render_saved_orders(orders_df, selected_order_id=None):
    """저장된 출고지시서를 1컬럼 표 형태로 렌더링하고 선택된 id를 반환한다."""
    if orders_df is None or orders_df.empty:
        st.info("저장된 출고지시가 없습니다.")
        return None

    st.markdown("""
    <style>
    .saved-order-table{width:100%;border:1px solid #dbe4f0;border-radius:14px;overflow:hidden;background:#fff;margin-top:4px;}
    .saved-order-head{display:grid;grid-template-columns:78px 120px minmax(360px,1fr) 90px;gap:0;align-items:center;background:#f1f5f9;color:#334155;font-weight:800;border-bottom:1px solid #dbe4f0;}
    .saved-order-head>div{padding:10px 12px;font-size:13px;}
    .saved-order-row{display:grid;grid-template-columns:78px 120px minmax(360px,1fr) 90px;gap:0;align-items:center;border-bottom:1px solid #edf2f7;min-height:42px;}
    .saved-order-row:last-child{border-bottom:none;}
    .saved-order-cell{padding:9px 12px;font-size:13px;color:#111827;font-weight:400;min-width:0;}
    .saved-order-title{text-align:left;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
    .saved-order-date{color:#334155;}
    .saved-order-status{text-align:center;}
    div[data-testid="stButton"] > button.saved-order-num-btn{justify-content:center!important;text-align:center!important;}
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class='saved-order-table'>
      <div class='saved-order-head'>
        <div style='text-align:center;'>번호</div>
        <div>날짜</div>
        <div style='text-align:left;'>제목</div>
        <div style='text-align:center;'>상태</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    for r in orders_df.itertuples():
        oid = int(r.id)
        title = str(r.title or "-")
        created = str(r.created_at or "")[:10]
        status = str(r.status or "저장됨")
        row_cols = st.columns([0.78, 1.2, 5.2, 0.9], gap="small")
        with row_cols[0]:
            if st.button(f"#{oid}", key=f"open_order_{oid}", use_container_width=True, type=("primary" if int(selected_order_id or 0) == oid else "secondary")):
                st.session_state["selected_saved_order_id"] = oid
                st.rerun()
        with row_cols[1]:
            st.markdown(f"<div class='saved-order-cell saved-order-date'>{escape(created)}</div>", unsafe_allow_html=True)
        with row_cols[2]:
            st.markdown(f"<div class='saved-order-cell saved-order-title' title='{escape(title)}'>{escape(title)}</div>", unsafe_allow_html=True)
        with row_cols[3]:
            st.markdown(f"<div class='saved-order-cell saved-order-status'>{_status_text_html(status)}</div>", unsafe_allow_html=True)
    return st.session_state.get("selected_saved_order_id") or (int(orders_df.iloc[0]["id"]) if not orders_df.empty else None)


def _run_cancel_order(order_id):
    """출고지시 취소 실행 후 상태 정리."""
    item_count, restored_count = cancel_saved_order(int(order_id))
    st.session_state.pop("confirm_cancel_order_id", None)
    st.session_state.pop("selected_saved_order_id", None)
    st.session_state["cancel_order_done_msg"] = f"출고지시서 #{int(order_id)} 취소 완료: {item_count}개 품목 / 원복 {restored_count}건"


def _show_cancel_order_confirm_inline(order_id):
    """Streamlit 버전이 낮아 dialog API가 없을 때만 쓰는 예비 확인 카드.
    실제 모달은 st.dialog/st.experimental_dialog를 우선 사용한다.
    """
    st.markdown("""
    <div style='border:1px solid #e5e7eb;background:#ffffff;border-radius:16px;padding:18px 20px;margin:12px auto;max-width:560px;box-shadow:0 18px 40px rgba(15,23,42,.12);'>
      <div style='font-weight:900;color:#111827;font-size:19px;margin-bottom:10px;'>⚠ 출고지시 취소 확인</div>
      <div style='color:#334155;font-weight:400;line-height:1.7;'>정말로 취소하시겠습니까?<br>제품의 수량은 출고지시 이전으로 복원됩니다.</div>
    </div>
    """, unsafe_allow_html=True)
    _left, c1, c2, _right = st.columns([1.2, 1, 1.7, 1.2])
    with c1:
        if st.button("아니오", use_container_width=True, key=f"cancel_no_{int(order_id)}"):
            st.session_state.pop("confirm_cancel_order_id", None)
            st.rerun()
    with c2:
        if st.button("예, 취소합니다", type="primary", use_container_width=True, key=f"cancel_yes_{int(order_id)}"):
            try:
                _run_cancel_order(int(order_id))
                st.rerun()
            except Exception as e:
                st.error(str(e))


_dialog_api = getattr(st, "dialog", None) or getattr(st, "experimental_dialog", None)

if _dialog_api:
    @_dialog_api("⚠ 출고지시 취소 확인")
    def _show_cancel_order_confirm(order_id):
        """화면 중앙 모달에서 출고지시 취소 여부를 확인한다.
        제목은 dialog 타이틀만 굵게 두고, 본문은 일반 굵기로 둔다.
        """
        st.markdown("""
        <style>
        div[data-testid="stDialog"] div[data-testid="stMarkdownContainer"] p,
        div[data-testid="stDialog"] div[data-testid="stMarkdownContainer"] div {
            font-weight:400!important;
        }
        div[data-testid="stDialog"] div[data-testid="stHorizontalBlock"]{
            justify-content:center!important;
        }
        div[data-testid="stDialog"] div[data-testid="stButton"] > button{
            min-height:46px!important;
            min-width:180px!important;
            border-radius:10px!important;
            font-weight:800!important;
            white-space:nowrap!important;
        }
        </style>
        <div style='font-size:16px;line-height:1.7;color:#334155;margin:6px 0 18px 0;font-weight:400;'>
            정말로 취소하시겠습니까?<br>
            제품의 수량은 출고지시 이전으로 복원됩니다.
        </div>
        """, unsafe_allow_html=True)
        _left, c1, c2, _right = st.columns([1.0, 1.0, 1.7, 1.0], gap="medium")
        with c1:
            if st.button("아니오", use_container_width=True, key=f"cancel_no_{int(order_id)}"):
                st.session_state.pop("confirm_cancel_order_id", None)
                st.rerun()
        with c2:
            if st.button("예, 취소합니다", type="primary", use_container_width=True, key=f"cancel_yes_{int(order_id)}"):
                try:
                    _run_cancel_order(int(order_id))
                    st.rerun()
                except Exception as e:
                    st.error(str(e))
else:
    def _show_cancel_order_confirm(order_id):
        _show_cancel_order_confirm_inline(order_id)

def page_saved_outbound():
    st.markdown("<h1 style='text-align:left;margin-bottom:0.2em;'>저장된 출고지시</h1>", unsafe_allow_html=True)
    if st.session_state.get("cancel_order_done_msg"):
        st.success(st.session_state.pop("cancel_order_done_msg"))
    st.caption("날짜 범위와 제목 검색으로 출고지시서를 필터링합니다.")
    st.markdown("""
    <style>
    section[data-testid="stSidebar"] div[data-testid="stButton"] > button,
    section[data-testid="stSidebar"] div[data-testid="stButton"] > button[kind="primary"],
    section[data-testid="stSidebar"] div[data-testid="stButton"] > button[kind="secondary"] {
        text-align:left!important; justify-content:flex-start!important;
    }
    section[data-testid="stSidebar"] div[data-testid="stButton"] > button p {text-align:left!important;width:100%!important;}
    </style>
    """, unsafe_allow_html=True)

    use_date_range = st.checkbox("날짜 범위 사용", value=False, key="saved_use_date_range")
    f1, f2, f3 = st.columns([1, 1, 2], gap="large")
    with f1:
        start_date = st.date_input("시작일", value=st.session_state.get("saved_start_date", datetime.now().date()), disabled=not use_date_range, key="saved_start_date")
    with f2:
        end_date = st.date_input("종료일", value=st.session_state.get("saved_end_date", datetime.now().date()), disabled=not use_date_range, key="saved_end_date")
    with f3:
        search_term = st.text_input("검색", placeholder="저장된 제목 일부 입력", key="saved_outbound_search")

    all_orders = q("SELECT id, created_at, order_date, title, status FROM outbound_orders ORDER BY id DESC")
    if all_orders.empty:
        st.info("저장된 출고지시가 없습니다.")
        return

    filtered = all_orders.copy()
    if use_date_range:
        if start_date and end_date and start_date > end_date:
            st.error("시작일은 종료일보다 늦을 수 없습니다.")
            return
        if start_date:
            filtered = filtered[filtered["order_date"] >= str(start_date)]
        if end_date:
            filtered = filtered[filtered["order_date"] <= str(end_date)]
    if search_term.strip():
        term = search_term.strip().lower()
        filtered = filtered[filtered["title"].fillna("").str.lower().str.contains(term, regex=False)]
    if filtered.empty:
        st.warning("조건에 맞는 출고지시서가 없습니다.")
        return

    total = len(filtered)
    per_page = 15
    max_page = max(1, (total + per_page - 1) // per_page)
    page_no = max(1, min(int(st.session_state.get("saved_order_page", 1)), max_page))
    st.session_state["saved_order_page"] = page_no
    orders = filtered.iloc[(page_no - 1) * per_page: page_no * per_page].copy()

    st.markdown(f"#### 출고지시서 {total}건")
    list_col, _ = st.columns([7, 3], gap="large")
    with list_col:
        selected_id = st.session_state.get("selected_saved_order_id")
        selected_id = render_saved_orders(orders, selected_id)
        if max_page > 1:
            nav_cols = st.columns([1, 3, 1])
            with nav_cols[0]:
                if st.button("이전", disabled=(page_no <= 1), key="page_prev", use_container_width=True):
                    st.session_state["saved_order_page"] = page_no - 1
                    st.rerun()
            with nav_cols[1]:
                st.markdown(f"<div style='text-align:center;color:#64748b;font-weight:700;margin:8px 0;'>{page_no} / {max_page}</div>", unsafe_allow_html=True)
            with nav_cols[2]:
                if st.button("다음", disabled=(page_no >= max_page), key="page_next", use_container_width=True):
                    st.session_state["saved_order_page"] = page_no + 1
                    st.rerun()

    valid_ids = set(filtered["id"].astype(int).tolist())
    order_id = st.session_state.get("selected_saved_order_id")
    if not order_id or int(order_id) not in valid_ids:
        order_id = int(orders.iloc[0]["id"])
        st.session_state["selected_saved_order_id"] = order_id

    order_row = all_orders[all_orders["id"] == int(order_id)]
    if order_row.empty:
        st.session_state.pop("selected_saved_order_id", None)
        return
    order_status = str(order_row.iloc[0]["status"] or "저장됨")

    st.markdown("---")
    selected_col, _spacer = st.columns([7, 3], gap="large")
    with selected_col:
        st.markdown(f"### 선택된 출고지시서 #{int(order_id)}")
        item_df = q("""
            SELECT id AS 품목ID, inventory_id AS 재고ID, location AS 로케이션, product_name AS 제품명,
                   lot AS LOT, exp_date AS 유통기한, qty AS 요청수량, company AS 사업장, warehouse_name AS 전산상명칭
            FROM outbound_order_items WHERE order_id=? ORDER BY id
        """, (int(order_id),))
        if item_df.empty:
            st.info("이 출고지시서에는 품목이 없습니다.")
        else:
            view_items = item_df[["로케이션", "제품명", "LOT", "유통기한", "요청수량"]]
            st.dataframe(view_items, hide_index=True, use_container_width=True)
            rows_for_download = view_items.to_dict("records")
            title_for_download = str(order_row.iloc[0]["title"] or f"출고지시서 #{int(order_id)}")
            d1, d2 = st.columns(2)
            with d1:
                st.download_button("선택 지시서 엑셀 다운로드", data=outbound_excel_bytes(rows_for_download, title_for_download), file_name=f"NOHTUS_출고지시서_{int(order_id)}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            with d2:
                try:
                    st.download_button("선택 지시서 PDF 다운로드", data=outbound_pdf_bytes(rows_for_download, title_for_download), file_name=f"NOHTUS_출고지시서_{int(order_id)}.pdf", mime="application/pdf", use_container_width=True)
                except Exception as e:
                    st.warning(f"PDF 생성 실패: {e}")

        c_edit, c_cancel = st.columns(2)
        with c_edit:
            if st.button("출고지시서 수정하기", type="primary", use_container_width=True, disabled=(order_status == "취소됨")):
                st.session_state["outbound_cart"] = load_outbound_order(int(order_id))
                st.session_state["editing_order_id"] = int(order_id)
                st.session_state["editing_order_title"] = str(order_row.iloc[0]["title"] or "")
                st.session_state["page"] = "출고지시"
                st.rerun()
        with c_cancel:
            if st.button("출고지시 취소하기", type="primary", use_container_width=True, key=f"cancel_order_{int(order_id)}", disabled=(order_status == "취소됨")):
                st.session_state["confirm_cancel_order_id"] = int(order_id)
                st.rerun()

        if st.session_state.get("confirm_cancel_order_id") == int(order_id):
            _show_cancel_order_confirm(int(order_id))


def page_search():
    st.title("제품 검색")
    term = st.text_input("검색어")
    opts = product_options(term)
    st.dataframe(opts.rename(columns={"standard_name":"표준제품명","warehouse_name":"전산상 명칭","aliases":"별칭"}), hide_index=True, use_container_width=True)

def page_stocktake():
    st.title("재고 실사")
    st.caption("재고 실사용 엑셀을 내려받고, 기준재고를 업로드하거나 제품명/LOT/유통기한/로케이션 단위로 필요한 재고만 조정합니다.")

    file_col, empty_col = st.columns([3, 7], gap="large")
    with file_col:
        st.subheader("실사 파일")
        excel_data = full_inventory_excel_bytes(exclude_zero=st.session_state.get("stocktake_exclude_zero", True))
        st.download_button(
            "재고 실사용 엑셀 내려받기",
            data=excel_data,
            file_name=f"NOHTUS_전체재고실사_{date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        exclude_zero = st.checkbox("재고가 0인 경우는 포함하지 않기", value=st.session_state.get("stocktake_exclude_zero", True), key="stocktake_exclude_zero")

        st.markdown("<div style='text-align:center;font-size:1.02rem;font-weight:900;margin:18px 0 6px;'>기준재고 업로드</div>", unsafe_allow_html=True)
        st.download_button(
            "기준재고 업로드 샘플 양식 다운로드",
            data=baseline_stock_template_excel_bytes(),
            file_name="NOHTUS_기준재고_업로드_샘플양식.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        survey_file = st.file_uploader("기준재고 엑셀 선택", type=["xlsx"], key="stock_survey_upload")
        replace_current = st.checkbox("기존 현재재고를 삭제하고 업로드 파일로 교체", value=True)
        if survey_file is not None:
            st.warning("업로드 실행 시 현재재고가 바뀔 수 있습니다. 운영 DB에서는 파일을 한 번 더 확인하세요.")
            if st.button("기준재고 DB 반영", type="primary", use_container_width=True):
                try:
                    survey_file.seek(0)
                    inserted, skipped, prod_inserted, ambiguous_skipped = import_stock_survey_excel(survey_file, replace_current=replace_current)
                    st.success(f"반영 완료: 재고 {inserted}건 / 제품매칭표 신규 {prod_inserted}건 / 제외 {skipped}건")
                    st.rerun()
                except Exception as e:
                    st.error(f"반영 실패: {e}")
    with empty_col:
        st.empty()

    st.markdown("---")
    st.subheader("재고조정")
    adj_df = q("""
        SELECT id, location, company, product_name, warehouse_name, lot, exp_date, qty
        FROM inventory
        WHERE qty > 0
        ORDER BY product_name, lot, exp_date, location
    """)
    if adj_df.empty:
        st.info("조정할 현재 재고가 없습니다.")
        return

    adjust_left, adjust_right = st.columns([4, 6], gap="large")
    with adjust_left:
        search = st.text_input("조정 대상 제품 검색", placeholder="제품명/전산상 명칭/LOT/로케이션 일부를 입력하세요", key="stock_adjust_search")
        filtered = adj_df.copy()
        if search.strip():
            term = search.strip().lower()
            filtered = filtered[
                filtered["product_name"].fillna("").str.lower().str.contains(term, regex=False)
                | filtered["warehouse_name"].fillna("").str.lower().str.contains(term, regex=False)
                | filtered["lot"].fillna("").str.lower().str.contains(term, regex=False)
                | filtered["location"].fillna("").str.lower().str.contains(term, regex=False)
            ]
        if filtered.empty:
            st.warning("검색어와 일치하는 재고가 없습니다.")
            return

        products = filtered["product_name"].dropna().astype(str).drop_duplicates().tolist()
        product = st.selectbox("제품명", products, key="stock_adjust_product")
        lot_df = filtered[filtered["product_name"] == product].copy()
        lots = lot_df["lot"].fillna("-").astype(str).drop_duplicates().tolist()
        lot = st.selectbox("LOT/제조번호", lots, key=f"stock_adjust_lot_{product}")
        exp_df = lot_df[lot_df["lot"].fillna("-").astype(str) == lot].copy()
        exps = exp_df["exp_date"].fillna("-").astype(str).drop_duplicates().tolist()
        exp = st.selectbox("유통기한", exps, key=f"stock_adjust_exp_{product}_{lot}", format_func=display_date_only)
        target_df = exp_df[exp_df["exp_date"].fillna("-").astype(str) == exp].copy()

    with adjust_right:
        st.markdown("#### 선택 재고")
        show = target_df[["id", "location", "company", "product_name", "warehouse_name", "lot", "exp_date", "qty"]].copy()
        show = show.rename(columns={
            "id": "ID", "location": "로케이션", "company": "사업장", "product_name": "표준제품명",
            "warehouse_name": "전산상명칭", "lot": "제조번호", "exp_date": "유통기한", "qty": "수량"
        })
        show["유통기한"] = show["유통기한"].apply(display_date_only)
        st.dataframe(show, hide_index=True, use_container_width=True)

        labels = []
        id_by_label = {}
        for r in target_df.itertuples():
            label = f"{r.location} / {r.company} / 현재 {int(r.qty)}EA"
            labels.append(label)
            id_by_label[label] = int(r.id)
        selected = st.selectbox("조정 대상 로케이션", labels, key=f"stock_adjust_inv_{product}_{lot}_{exp}")
        inv_id = id_by_label[selected]
        row = target_df[target_df["id"] == inv_id].iloc[0]
        actual = st.number_input("실물수량", min_value=0, value=int(row["qty"]), step=1, key=f"stock_adjust_actual_{inv_id}")
        reason = st.selectbox("사유", ["실사차이", "파손", "유통기한만료", "오출고", "기타"], key=f"stock_adjust_reason_{inv_id}")
        memo = st.text_input("메모", placeholder="필요 시 입력", key=f"stock_adjust_memo_{inv_id}")
        if st.button("재고조정 저장", type="primary", use_container_width=True, key=f"stock_adjust_submit_{inv_id}"):
            try:
                before, after, diff = adjust_inventory(int(inv_id), int(actual), reason, memo)
                st.session_state["_stock_adjust_success_msg"] = f"재고조정 완료: {before}EA → {after}EA ({diff:+d}EA)"
                st.rerun()
            except Exception as e:
                st.error(str(e))
        stock_adjust_msg = st.session_state.pop("_stock_adjust_success_msg", None)
        if stock_adjust_msg:
            st.success(stock_adjust_msg)

def page_history():
    st.title("이력 조회")

    today = date.today()
    default_start = today.replace(day=1)

    filter_col1, filter_col2, filter_col3, filter_col4 = st.columns(4)
    with filter_col1:
        company = st.selectbox("사업장", ["전체"] + COMPANIES, index=0, key="history_company")
    with filter_col2:
        tx_label = st.selectbox("이력유형", ["전체", "입고", "출고지시", "출고지시취소", "이동", "재고조정", "재고정보수정", "전산재고"], index=0, key="history_tx_label")
    with filter_col3:
        start_date = st.date_input("시작일", value=default_start, key="history_start_date")
    with filter_col4:
        end_date = st.date_input("종료일", value=today, key="history_end_date")

    if start_date and end_date and start_date > end_date:
        st.error("시작일은 종료일보다 늦을 수 없습니다.")
        return

    search_col, blank_col = st.columns(2)
    with search_col:
        term = st.text_input("제품명/로케이션 검색", placeholder="제품명, LOT, 로케이션 일부 입력", key="history_search_term")
    with blank_col:
        st.markdown("&nbsp;", unsafe_allow_html=True)

    filter_key = f"{company}|{tx_label}|{start_date}|{end_date}|{term.strip()}"
    if st.session_state.get("history_filter_key") != filter_key:
        st.session_state["history_filter_key"] = filter_key
        st.session_state["history_visible_limit"] = 500
    visible_limit = int(st.session_state.get("history_visible_limit", 500) or 500)
    visible_limit = max(500, visible_limit)

    conditions = []
    params = []
    if start_date:
        conditions.append("date(created_at) >= ?")
        params.append(str(start_date))
    if end_date:
        conditions.append("date(created_at) <= ?")
        params.append(str(end_date))
    if company != "전체":
        conditions.append("(from_company=? OR to_company=?)")
        params.extend([company, company])
    if tx_label != "전체":
        if tx_label == "이동":
            conditions.append("tx_type IN ('위치이동','사업장이동','사업장+위치이동','비자료전환','이동')")
        elif tx_label == "전산재고":
            conditions.append("tx_type IN ('기준재고','전산재고')")
        else:
            conditions.append("tx_type=?")
            params.append(tx_label)
    else:
        conditions.append("tx_type NOT IN ('재고조사불러오기','ERP비교','출고','출고확정')")
    if term.strip():
        like = f"%{term.strip()}%"
        conditions.append("(product_name LIKE ? OR lot LIKE ? OR from_location LIKE ? OR to_location LIKE ? OR memo LIKE ?)")
        params.extend([like, like, like, like, like])
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    total_df = q(f"SELECT COUNT(*) AS cnt FROM transactions {where}", tuple(params))
    total_count = int(total_df.iloc[0]["cnt"] or 0) if not total_df.empty else 0
    if total_count == 0:
        st.info("조회된 이력이 없습니다.")
        return

    df = q(f"""
        SELECT * FROM transactions
        {where}
        ORDER BY id DESC
        LIMIT ? OFFSET 0
    """, tuple(params + [visible_limit]))

    st.caption(f"총 {total_count:,}건 중 {min(len(df), total_count):,}건 표시 중 (한 번에 500건씩 추가 로딩)")

    # warehouse_name은 화면에서 제외하고, 재고조정/재고실사는 +/-와 현재 최종재고를 표시한다.
    final_values = []
    qty_values = []
    for r in df.itertuples():
        typ = str(getattr(r, "tx_type", "") or "")
        qty = int(getattr(r, "qty", 0) or 0)
        if typ in ["재고조정", "재고실사"]:
            qty_values.append(f"{qty:+d}")
        else:
            qty_values.append(str(qty))
        final_stock_value = getattr(r, "final_stock", None)
        if final_stock_value is not None and not pd.isna(final_stock_value):
            final_values.append(int(final_stock_value))
        else:
            # 구버전 이력처럼 final_stock 스냅샷이 없는 기록은 현재 재고로 다시 계산하지 않는다.
            # 이력의 최종재고는 작업 당시 값이어야 하므로, 저장값이 없으면 빈칸으로 둔다.
            final_values.append("")
    show = df.copy()
    show["수량"] = qty_values
    show["최종재고"] = final_values
    if "exp_date" in show.columns:
        show["exp_date"] = show["exp_date"].apply(display_date_only)
    rename_cols = {
        "created_at":"일시", "tx_type":"이력유형", "product_name":"제품명",
        "lot":"LOT", "exp_date":"유통기한", "from_company":"출발사업장", "from_location":"출발위치",
        "to_company":"도착사업장", "to_location":"도착위치", "memo":"메모"
    }
    show = show.rename(columns=rename_cols)
    wanted = ["일시","이력유형","제품명","LOT","유통기한","출발사업장","출발위치","도착사업장","도착위치","수량","최종재고","메모"]
    show = show[[c for c in wanted if c in show.columns]]
    st.dataframe(show, use_container_width=True, hide_index=True)

    if visible_limit < total_count:
        if st.button("500건 더 보기", use_container_width=True):
            st.session_state["history_visible_limit"] = visible_limit + 500
            st.rerun()

def page_master():
    st.title("제품 마스터")
    st.caption("제품 자체의 기준명은 표준제품명으로 관리하고, ERP별 이름은 별도 매핑으로 관리합니다. 노투스팜/NOH ERP 제품코드는 각 ERP명 바로 뒤에서 관리하고, 노투스 ERP명 오른쪽에는 비자료명을 관리합니다.")
    df = q("SELECT id, product_code, standard_name, aliases, erp_nohtuspharm_name, erp_noh_name, erp_noh_code, erp_nohtus_name, bidata_name FROM products ORDER BY standard_name")

    top1, top2 = st.columns(2, gap="large")
    with top1:
        st.download_button(
            "제품 마스터 엑셀 양식 다운로드",
            data=product_master_excel_bytes(),
            file_name=f"NOHTUS_제품마스터_{date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with top2:
        uploaded = st.file_uploader("수정한 제품 마스터 엑셀 업로드", type=["xlsx"], key="product_master_upload")
        if uploaded is not None:
            if st.button("업로드 파일로 제품 마스터 업데이트", type="primary", use_container_width=True):
                try:
                    updated, inserted, skipped = import_product_master_excel(uploaded)
                    st.success(f"업데이트 완료: 수정 {updated}건 / 추가 {inserted}건 / 건너뜀 {skipped}건")
                    st.rerun()
                except Exception as e:
                    st.error(f"업로드 실패: {e}")

    st.markdown("### 제품 목록")
    view = df[["standard_name", "erp_nohtuspharm_name", "product_code", "erp_noh_name", "erp_noh_code", "erp_nohtus_name", "bidata_name", "aliases"]].rename(columns={
        "standard_name":"표준제품명",
        "erp_nohtuspharm_name":"노투스팜 ERP명",
        "product_code":"노투스팜 ERP 제품코드",
        "erp_noh_name":"NOH ERP명",
        "erp_noh_code":"NOH ERP 제품코드",
        "erp_nohtus_name":"노투스 ERP명",
        "bidata_name":"비자료명",
        "aliases":"별칭",
    })
    st.dataframe(view, use_container_width=True, hide_index=True)

    with st.expander("기존 제품 수정", expanded=False):
        if df.empty:
            st.info("수정할 제품이 없습니다.")
        else:
            edit_term = st.text_input("수정할 제품 검색", placeholder="표준제품명/ERP명/별칭 일부 입력")
            edit_df = df.copy()
            if edit_term.strip():
                term = edit_term.strip().lower()
                edit_df = edit_df[
                    edit_df["standard_name"].fillna("").str.lower().str.contains(term, regex=False)
                    | edit_df["aliases"].fillna("").str.lower().str.contains(term, regex=False)
                    | edit_df["product_code"].fillna("").str.lower().str.contains(term, regex=False)
                    | edit_df["erp_nohtuspharm_name"].fillna("").str.lower().str.contains(term, regex=False)
                    | edit_df["erp_noh_name"].fillna("").str.lower().str.contains(term, regex=False)
                    | edit_df["erp_noh_code"].fillna("").str.lower().str.contains(term, regex=False)
                    | edit_df["erp_nohtus_name"].fillna("").str.lower().str.contains(term, regex=False)
                    | edit_df["bidata_name"].fillna("").str.lower().str.contains(term, regex=False)
                ]
            if edit_df.empty:
                st.warning("일치하는 제품이 없습니다.")
            else:
                options = [f"{r.standard_name} / 노투스팜:{r.erp_nohtuspharm_name or '-'} / NOH:{r.erp_noh_name or '-'} / 노투스:{r.erp_nohtus_name or '-'}" for r in edit_df.itertuples()]
                selected = st.selectbox("수정할 제품 선택", options)
                row = edit_df.iloc[options.index(selected)]
                with st.form("edit_product"):
                    name = st.text_input("표준제품명", value=str(row["standard_name"] or ""))
                    erp_np = st.text_input("노투스팜 ERP명", value=str(row.get("erp_nohtuspharm_name", "") or ""))
                    code = st.text_input("노투스팜 ERP 제품코드", value=str(row["product_code"] or ""))
                    erp_noh = st.text_input("NOH ERP명", value=str(row.get("erp_noh_name", "") or ""))
                    erp_noh_code = st.text_input("NOH ERP 제품코드", value=str(row.get("erp_noh_code", "") or ""))
                    erp_nt = st.text_input("노투스 ERP명", value=str(row.get("erp_nohtus_name", "") or ""))
                    bidata_name = st.text_input("비자료명", value=str(row.get("bidata_name", "") or ""))
                    aliases = st.text_input("별칭", value=str(row["aliases"] or ""))
                    if st.form_submit_button("수정 저장", use_container_width=True):
                        if not name.strip():
                            st.error("표준제품명은 필수입니다.")
                        else:
                            exec_sql("UPDATE products SET product_code=?, standard_name=?, warehouse_name=?, aliases=?, erp_nohtuspharm_name=?, erp_noh_name=?, erp_noh_code=?, erp_nohtus_name=?, bidata_name=? WHERE id=?", (code.strip(), name.strip(), name.strip(), aliases.strip(), erp_np.strip(), erp_noh.strip(), erp_noh_code.strip(), erp_nt.strip(), bidata_name.strip(), int(row["id"])))
                            st.success("제품 수정 완료")
                            st.rerun()

    with st.expander("제품 추가"):
        with st.form("add_product"):
            name=st.text_input("표준제품명")
            erp_np=st.text_input("노투스팜 ERP명")
            code=st.text_input("노투스팜 ERP 제품코드")
            erp_noh=st.text_input("NOH ERP명")
            erp_noh_code=st.text_input("NOH ERP 제품코드")
            erp_nt=st.text_input("노투스 ERP명")
            bidata_name=st.text_input("비자료명")
            aliases=st.text_input("별칭")
            if st.form_submit_button("추가", use_container_width=True) and name:
                exec_sql("INSERT INTO products(product_code,standard_name,warehouse_name,aliases,erp_nohtuspharm_name,erp_noh_name,erp_noh_code,erp_nohtus_name,bidata_name) VALUES(?,?,?,?,?,?,?,?,?)", (code.strip(),name.strip(),name.strip(),aliases.strip(),erp_np.strip(),erp_noh.strip(),erp_noh_code.strip(),erp_nt.strip(),bidata_name.strip()))
                st.success("제품 추가 완료"); st.rerun()

    with st.expander("ERP 확인 필요 후보 관리", expanded=True):
        st.caption("같은 ERP명칭이 실제 여러 제품일 수 있는 경우 후보를 등록합니다. ERP 업로드 시 후보가 2개 이상이면 사람이 선택해야 합니다.")
        c1, c2, c3 = st.columns([1,1.4,1.6], gap="medium")
        with c1:
            erp_company = st.selectbox("ERP구분", ["노투스팜", "NOH", "노투스"], key="amb_erp_company")
        with c2:
            erp_name = st.text_input("ERP명칭", placeholder="예: JS TOX", key="amb_erp_name")
        with c3:
            products = q("SELECT standard_name FROM products ORDER BY standard_name")
            cand = st.selectbox("후보 표준제품", products["standard_name"].tolist() if not products.empty else [], key="amb_candidate")
        memo = st.text_input("메모", placeholder="예: ERP명만으로 실제 출고제품 판단 불가", key="amb_memo")
        if st.button("확인 필요 후보 추가", type="primary", use_container_width=True):
            if not erp_name.strip() or not cand:
                st.error("ERP명칭과 후보 제품을 입력하세요.")
            else:
                exec_sql("INSERT INTO erp_ambiguous_candidates(erp_company, erp_name, candidate_product, memo) VALUES(?,?,?,?)", (erp_company, erp_name.strip(), cand, memo.strip()))
                st.success("후보 추가 완료")
                st.rerun()
        amb = q("SELECT id, erp_company AS ERP구분, erp_name AS ERP명칭, candidate_product AS 후보제품, memo AS 메모 FROM erp_ambiguous_candidates ORDER BY ERP구분, ERP명칭, 후보제품")
        if amb.empty:
            st.info("등록된 확인 필요 후보가 없습니다.")
        else:
            st.caption("삭제할 후보만 체크한 뒤 삭제 버튼을 누르세요.")
            amb_edit = amb.copy()
            amb_edit.insert(0, "삭제", False)
            edited_amb = st.data_editor(
                amb_edit,
                hide_index=True,
                use_container_width=True,
                disabled=["id", "ERP구분", "ERP명칭", "후보제품", "메모"],
                column_config={"id": None, "삭제": st.column_config.CheckboxColumn("삭제")},
                key="amb_delete_editor",
            )
            delete_ids = edited_amb.loc[edited_amb["삭제"] == True, "id"].astype(int).tolist()
            if st.button("체크한 후보 삭제", disabled=(len(delete_ids) == 0), use_container_width=True):
                with connect() as con:
                    con.executemany("DELETE FROM erp_ambiguous_candidates WHERE id=?", [(i,) for i in delete_ids])
                    con.commit()
                st.success(f"{len(delete_ids)}건 삭제 완료")
                st.rerun()

    with st.expander("ERP명칭 매칭 테스트"):
        t1, t2 = st.columns([1,2], gap="medium")
        with t1:
            test_company = st.selectbox("테스트 ERP구분", ["노투스팜", "NOH", "노투스"], key="test_erp_company")
        with t2:
            test_name = st.text_input("테스트 ERP명칭", placeholder="ERP 매입/매출 엑셀에 나온 제품명", key="test_erp_name")
        if st.button("매칭 확인", use_container_width=True):
            res = match_erp_name(test_company, test_name)
            if res["status"] == "auto":
                st.success(res["message"])
            elif res["status"] == "ambiguous":
                st.warning(res["message"])
                st.write("후보제품")
                st.write(res["candidates"])
            else:
                st.error(res["message"])


# ---------------- ERP / customer master ----------------
def customer_master_template_bytes():
    df = pd.DataFrame(columns=["거래처코드", "거래처명", "사업장", "유형", "담당자", "연락처", "주소", "메모"])
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="거래처관리")
        ws = writer.book["거래처관리"]
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill = PatternFill("solid", fgColor="E5E7EB")
        widths = {"A":18,"B":30,"C":16,"D":16,"E":18,"F":18,"G":44,"H":30}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=8):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = fill
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:H2"
    bio.seek(0)
    return bio.getvalue()


def import_customer_master_excel(uploaded_file):
    df = pd.read_excel(uploaded_file, dtype=str).fillna("")
    rename = {
        "거래처코드":"customer_code", "코드":"customer_code",
        "거래처명":"customer_name", "거래처":"customer_name", "상호":"customer_name",
        "사업장":"company", "법인":"company",
        "유형":"customer_type", "거래처유형":"customer_type",
        "담당자":"manager", "담당":"manager",
        "연락처":"phone", "전화번호":"phone", "핸드폰":"phone",
        "주소":"address", "납품처":"address", "배송지":"address",
        "메모":"memo", "비고":"memo",
    }
    df = df.rename(columns={c: rename.get(str(c).strip(), c) for c in df.columns})
    if "customer_name" not in df.columns:
        raise ValueError("엑셀에 '거래처명' 컬럼이 필요합니다.")
    for c in ["customer_code", "company", "customer_type", "manager", "phone", "address", "memo"]:
        if c not in df.columns:
            df[c] = ""
    updated = inserted = skipped = 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with connect() as con:
        cur = con.cursor()
        for _, r in df.iterrows():
            name = "" if pd.isna(r.get("customer_name")) else str(r.get("customer_name")).strip()
            code = "" if pd.isna(r.get("customer_code")) else str(r.get("customer_code")).strip()
            if not name:
                skipped += 1
                continue
            vals = []
            for c in ["company", "customer_type", "manager", "phone", "address", "memo"]:
                vals.append("" if pd.isna(r.get(c)) else str(r.get(c)).strip())
            existing = None
            if code:
                existing = cur.execute("SELECT id FROM customers WHERE customer_code=?", (code,)).fetchone()
            if not existing:
                existing = cur.execute("SELECT id FROM customers WHERE customer_name=?", (name,)).fetchone()
            if existing:
                cur.execute("UPDATE customers SET customer_code=?, customer_name=?, company=?, customer_type=?, manager=?, phone=?, address=?, memo=?, updated_at=? WHERE id=?", (code, name, vals[0], vals[1], vals[2], vals[3], vals[4], vals[5], now, int(existing[0])))
                updated += 1
            else:
                cur.execute("INSERT INTO customers(customer_code, customer_name, company, customer_type, manager, phone, address, memo, updated_at) VALUES(?,?,?,?,?,?,?,?,?)", (code, name, vals[0], vals[1], vals[2], vals[3], vals[4], vals[5], now))
                inserted += 1
        con.commit()
    return updated, inserted, skipped




def customer_export_excel_bytes():
    df = q("SELECT customer_code AS 거래처코드, customer_name AS 거래처명, company AS 사업장, customer_type AS 유형, manager AS 담당자, phone AS 연락처, address AS 주소, memo AS 메모 FROM customers ORDER BY customer_name")
    return dataframe_to_excel_bytes(df, "거래처관리")

def page_inventory_metadata_edit():
    st.title("재고정보 수정")
    st.caption("기존 재고의 제조번호/유통기한이 잘못 입력된 경우에만 사용합니다. 수량은 변경하지 않습니다.")
    render_inventory_metadata_editor()


def page_customer_master():
    st.title("거래처 관리")
    st.caption("거래처 엑셀을 업로드하면 출고지시와 업무일지에서 매출처/담당자 정보를 재사용할 수 있습니다.")
    c1, c2 = st.columns(2, gap="large")
    with c1:
        st.download_button("등록된 거래처 내려받기", data=customer_export_excel_bytes(), file_name=f"NOHTUS_등록거래처_{date.today().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with c2:
        up = st.file_uploader("거래처 관리 엑셀 업로드", type=["xlsx"], key="customer_master_upload")
        if up is not None and st.button("거래처 관리 업데이트", type="primary", use_container_width=True):
            try:
                updated, inserted, skipped = import_customer_master_excel(up)
                st.success(f"업데이트 완료: 수정 {updated}건 / 추가 {inserted}건 / 건너뜀 {skipped}건")
                st.rerun()
            except Exception as e:
                st.error(f"업로드 실패: {e}")
    df = q("SELECT customer_code AS 거래처코드, customer_name AS 거래처명, company AS 사업장, customer_type AS 유형, manager AS 담당자, phone AS 연락처, address AS 주소, memo AS 메모 FROM customers ORDER BY customer_name")
    st.markdown("### 등록된 거래처")
    if df.empty:
        st.info("등록된 거래처가 없습니다.")
    else:
        term = st.text_input("거래처 검색", placeholder="거래처명/담당자/주소 일부 입력")
        if term.strip():
            low = term.strip().lower()
            mask = False
            for col in df.columns:
                mask = mask | df[col].fillna("").astype(str).str.lower().str.contains(low, regex=False)
            df = df[mask]
        st.dataframe(df, use_container_width=True, hide_index=True)


def detect_column(columns, candidates):
    cols = [str(c).strip() for c in columns]
    for cand in candidates:
        for c in cols:
            if cand.lower() == c.lower():
                return c
    for cand in candidates:
        for c in cols:
            if cand.lower() in c.lower():
                return c
    return cols[0] if cols else None



def clean_excel_text(value):
    """openpyxl이 저장하지 못하는 제어문자와 특수 공백을 제거한다."""
    if value is None:
        return ""
    text = str(value)
    try:
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        text = ILLEGAL_CHARACTERS_RE.sub("", text)
    except Exception:
        text = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", text)
    text = re.sub(r"[\x00-\x1F\x7F-\x9F\u200b\u200c\u200d\ufeff]", "", text)
    text = text.replace(" ", " ")
    return text.strip()


def _erp_name_key(value):
    text = clean_excel_text(value)
    return re.sub(r"\s+", "", text).replace("[", "").replace("]", "")


def is_ignored_erp_product_name(value):
    key = _erp_name_key(value)
    return (not key) or key in {"합계", "배송비"} or "합계" in key or "배송비" in key


def _find_required_column(columns, candidates):
    cols = [str(c).strip() for c in columns]
    for cand in candidates:
        for col in cols:
            if col == cand:
                return col
    for cand in candidates:
        for col in cols:
            if cand in col:
                return col
    return None

def _read_erp_current_file(uploaded, company, key_prefix):
    """ERP 현재고 파일을 고정 규칙으로 읽는다.
    - 노투스팜/NOH: 제품명, 현재고수량
    - 노투스: 8행 header, 품목명/규격, 현재재고
    제조번호/유통기한은 ERP 비교에서 사용하지 않는다.
    """
    if uploaded is None:
        return None
    try:
        if company == "노투스":
            raw = pd.read_excel(uploaded, header=7)
            name_col = _find_required_column(raw.columns, ["품목명/규격", "품목명", "제품명"])
            qty_col = _find_required_column(raw.columns, ["현재재고", "현재고", "현재고수량"])
        else:
            raw = pd.read_excel(uploaded)
            name_col = _find_required_column(raw.columns, ["제품명"])
            qty_col = _find_required_column(raw.columns, ["현재고수량", "현재고"])
    except Exception as e:
        st.error(f"{company} 엑셀 읽기 실패: {e}")
        return None
    if raw.empty:
        st.warning(f"{company} 엑셀에 데이터가 없습니다.")
        return None
    raw.columns = [str(c).strip() for c in raw.columns]
    if not name_col or not qty_col or name_col not in raw.columns or qty_col not in raw.columns:
        st.error(f"{company} ERP 파일에서 필요한 컬럼을 찾을 수 없습니다.")
        st.caption("노투스팜/NOH: 제품명, 현재고수량 · 노투스: 8행 헤더의 품목명/규격, 현재재고")
        return None
    return {"company": company, "raw": raw, "name_col": name_col, "qty_col": qty_col}

def _normalize_erp_current_rows(parsed_files):
    """ERP 파일을 사업장 + ERP 제품명 기준으로 합산한다.
    ERP 파일은 이미 ERP명으로 되어 있으므로 제품매칭표 역변환은 WMS 쪽에만 적용한다.
    """
    rows = []
    fail = []
    amb = []
    auto_count = 0
    for info in parsed_files:
        if not info:
            continue
        company = info["company"]
        raw = info["raw"].copy()
        name_col = info["name_col"]
        qty_col = info["qty_col"]
        raw[name_col] = raw[name_col].apply(clean_excel_text)
        raw[qty_col] = pd.to_numeric(raw[qty_col], errors="coerce").fillna(0).astype(int)
        raw = raw[~raw[name_col].apply(is_ignored_erp_product_name)]
        raw = raw[raw[qty_col] != 0]
        if raw.empty:
            continue
        for _, rr in raw.iterrows():
            erp_name = clean_excel_text(rr[name_col])
            erp_qty = int(rr[qty_col])
            if is_ignored_erp_product_name(erp_name):
                continue
            rows.append({"사업장": company, "제품명": erp_name, "ERP수량": erp_qty})
            auto_count += 1
    if rows:
        erp = pd.DataFrame(rows)
        erp = erp.groupby(["사업장", "제품명"], as_index=False)["ERP수량"].sum()
    else:
        erp = pd.DataFrame(columns=["사업장", "제품명", "ERP수량"])
    return erp, pd.DataFrame(amb), pd.DataFrame(fail), auto_count

def dataframe_to_excel_bytes(df, sheet_name="Sheet1"):
    """DataFrame을 엑셀 bytes로 변환한다.
    openpyxl이 허용하지 않는 제어문자/특수 공백은 저장 전에 전부 제거한다.
    """
    bio = BytesIO()
    safe_df = df.copy() if df is not None else pd.DataFrame()
    safe_sheet = clean_excel_text(sheet_name)[:31] or "Sheet1"
    safe_df.columns = [clean_excel_text(c) for c in safe_df.columns]
    for col in safe_df.columns:
        if safe_df[col].dtype == object:
            safe_df[col] = safe_df[col].apply(lambda v: clean_excel_text(v) if v is not None else "")
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        safe_df.to_excel(writer, index=False, sheet_name=safe_sheet)
        ws = writer.book[safe_sheet]
        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        for col in ws.columns:
            max_len = 10
            letter = col[0].column_letter
            for cell in col:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                max_len = max(max_len, len(str(cell.value or "")) + 2)
            ws.column_dimensions[letter].width = min(max_len, 42)
        if safe_sheet == "마감체크":
            # 마감 체크리스트 전용 서식: C열은 좁게, D열은 제품명 확인용으로 넓게.
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 50
            # 현재수량 컬럼은 연한 하늘색으로 표시한다.
            from openpyxl.styles import PatternFill
            current_fill = PatternFill("solid", fgColor="DDEBF7")
            header_map = {str(ws.cell(row=1, column=i).value or "").strip(): i for i in range(1, ws.max_column + 1)}
            cur_col = header_map.get("현재수량")
            if cur_col:
                for rr in range(1, ws.max_row + 1):
                    ws.cell(row=rr, column=cur_col).fill = current_fill
        ws.freeze_panes = "A2"
        if safe_df.shape[1] > 0:
            ws.auto_filter.ref = ws.dimensions
    bio.seek(0)
    return bio.getvalue()



def page_erp_stock_compare():
    st.title("ERP 재고 비교")
    st.caption("ERP/WMS 모두 해당 사업장의 ERP명 기준으로 제조번호·유통기한 없이 총수량만 비교합니다. 비자료는 비교 대상에서 제외합니다.")

    parsed = []
    cols = st.columns(3, gap="large")
    info = [("노투스팜", "SIMS"), ("NOH", "SIMS"), ("노투스", "IBK우리은행 전산")]
    for col, (company, system_name) in zip(cols, info):
        with col:
            st.markdown(f"### {company}")
            st.caption(system_name)
            up = st.file_uploader(f"{company} 현재고 엑셀", type=["xlsx", "xls"], key=f"erp_compare_upload_{company}")
            if up is not None:
                parsed.append(_read_erp_current_file(up, company, "compare"))
                st.success("파일 선택됨")
            else:
                parsed.append(None)

    if not any(parsed):
        st.info("노투스팜 / NOH / 노투스 ERP 현재고 엑셀을 업로드하세요.")
        return

    _, run_col, _ = st.columns([3, 2, 3], gap="large")
    with run_col:
        run_compare = st.button("ERP 재고 비교 실행", type="primary", use_container_width=True)

    if run_compare:
        erp, amb, fail, auto_count = _normalize_erp_current_rows(parsed)

        wms_raw = q("""SELECT company AS 사업장, product_name AS 표준제품명, SUM(qty) AS 수량
                       FROM inventory
                       WHERE qty<>0 AND company IN ('노투스팜','NOH','노투스')
                       GROUP BY company, product_name""")
        if wms_raw.empty:
            wms_sum = pd.DataFrame(columns=["사업장", "제품명", "WMS수량"])
        else:
            wms_raw["제품명"] = wms_raw.apply(lambda r: clean_excel_text(product_compare_name_for(r["사업장"], r["표준제품명"])), axis=1)
            wms_raw = wms_raw[~wms_raw["제품명"].apply(is_ignored_erp_product_name)]
            wms_sum = (
                wms_raw.groupby(["사업장", "제품명"], as_index=False)["수량"]
                .sum()
                .rename(columns={"수량": "WMS수량"})
            )

        comp = wms_sum.merge(erp, how="outer", on=["사업장", "제품명"])
        if comp.empty:
            comp = pd.DataFrame(columns=["사업장", "제품명", "ERP수량", "WMS수량", "차이"])
        else:
            comp["ERP수량"] = comp["ERP수량"].fillna(0).astype(int)
            comp["WMS수량"] = comp["WMS수량"].fillna(0).astype(int)
            comp["차이"] = comp["WMS수량"] - comp["ERP수량"]
            comp = comp[["사업장", "제품명", "ERP수량", "WMS수량", "차이"]].sort_values(["사업장", "제품명"])

        st.session_state["erp_compare_rows"] = comp.to_dict("records")
        st.session_state["erp_compare_amb"] = amb.to_dict("records") if not amb.empty else []
        st.session_state["erp_compare_fail"] = fail.to_dict("records") if not fail.empty else []
        st.session_state["erp_compare_summary"] = {"auto": auto_count, "amb": len(amb), "fail": len(fail)}

    if "erp_compare_summary" in st.session_state:
        sm = st.session_state["erp_compare_summary"]
        m1, m2, m3 = st.columns(3)
        m1.metric("ERP 행 반영", f"{sm['auto']}건")
        m2.metric("확인 필요", f"{sm['amb']}건")
        m3.metric("매칭 실패", f"{sm['fail']}건")
        comp = pd.DataFrame(st.session_state.get("erp_compare_rows", []))
        amb = pd.DataFrame(st.session_state.get("erp_compare_amb", []))
        fail = pd.DataFrame(st.session_state.get("erp_compare_fail", []))
        if not comp.empty:
            st.markdown("### 재고 비교 결과")
            only_diff = st.checkbox("차이 있는 항목만 보기", value=True, key="erp_compare_only_diff")
            shown = comp[comp["차이"] != 0] if only_diff else comp
            st.dataframe(shown, use_container_width=True, hide_index=True)
            st.download_button(
                "비교 결과 엑셀 다운로드",
                data=dataframe_to_excel_bytes(comp, "ERP_WMS_비교"),
                file_name=f"NOHTUS_ERP_WMS_비교_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        if not amb.empty:
            st.warning("확인 필요 항목")
            st.dataframe(amb, use_container_width=True, hide_index=True)
        if not fail.empty:
            st.error("매칭 실패 항목")
            st.dataframe(fail, use_container_width=True, hide_index=True)



def mapping_source_column_for_company(company):
    """사업장별 원본명 컬럼을 반환한다. 비자료는 ERP명이 아니라 비자료명 컬럼만 사용한다."""
    if company == "노투스팜":
        return "erp_nohtuspharm_name", "노투스팜 ERP명"
    if company == "NOH":
        return "erp_noh_name", "NOH ERP명"
    if company == "노투스":
        return "erp_nohtus_name", "노투스 ERP명"
    if company == "비자료":
        return "bidata_name", "비자료명"
    return None, "원본명"


def is_mapping_conflict_approved(company, source_name):
    company = (company or "").strip()
    source_name = (source_name or "").strip()
    if not company or not source_name:
        return False
    df = q("""SELECT 1 FROM product_match_conflict_approvals
             WHERE company=? AND source_name=? LIMIT 1""", (company, source_name))
    return not df.empty


def approve_mapping_conflict(company, source_name):
    company = (company or "").strip()
    source_name = (source_name or "").strip()
    if not company or not source_name:
        return
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    exec_sql("""INSERT OR REPLACE INTO product_match_conflict_approvals(company, source_name, approved_at)
                VALUES(?,?,?)""", (company, source_name, now))


def find_mapping_conflicts_from_inventory():
    """현재 재고의 ERP명/비자료명이 제품매칭표에서 다른 표준제품명에도 연결된 경우를 찾는다.
    사용자가 이미 OK한 원본명은 다시 묻지 않는다.
    """
    inv = q("""
        SELECT company, product_name, COALESCE(warehouse_name, '') AS source_name,
               SUM(qty) AS qty, GROUP_CONCAT(DISTINCT location) AS locations
        FROM inventory
        WHERE TRIM(COALESCE(company,''))<>''
          AND TRIM(COALESCE(product_name,''))<>''
          AND qty > 0
        GROUP BY company, product_name, COALESCE(warehouse_name, '')
        ORDER BY company, source_name, product_name
    """)
    if inv.empty:
        return pd.DataFrame()
    rows = []
    for r in inv.itertuples():
        company = str(r.company or "").strip()
        source_name = str(r.source_name or "").strip()
        current_standard = str(r.product_name or "").strip()
        if not source_name or not current_standard or source_name == "-":
            continue
        col, label = mapping_source_column_for_company(company)
        if not col:
            continue
        if is_mapping_conflict_approved(company, source_name):
            continue
        mapped = q(f"""SELECT standard_name FROM products
                       WHERE TRIM(COALESCE({col}, '')) = ?
                       ORDER BY standard_name""", (source_name,))
        mapped_names = sorted({str(x).strip() for x in mapped.get("standard_name", []) if str(x).strip()}) if not mapped.empty else []
        other_names = [x for x in mapped_names if x != current_standard]
        if other_names:
            rows.append({
                "사업장": company,
                "원본명 구분": label,
                "ERP명/비자료명": source_name,
                "재고 DB 표준제품명": current_standard,
                "제품매칭표 기존 표준제품명": " / ".join(mapped_names),
                "충돌 가능 내용": f"재고 DB는 '{current_standard}'로 쓰고 있지만 제품매칭표에는 '{source_name}'이(가) 다른 표준제품명에도 등록되어 있습니다.",
                "현재재고수량": int(r.qty or 0),
                "재고위치": str(r.locations or ""),
            })
    return pd.DataFrame(rows)

def page_product_matching():
    st.title("제품 매칭 관리")
    st.caption("표준제품명과 사업장별 ERP명/비자료명을 관리합니다.")

    action_col, _spacer = st.columns([4, 6], gap="large")
    with action_col:
        st.download_button(
            "제품 매칭표 엑셀 파일 내려받기",
            data=product_master_excel_bytes(),
            file_name=f"NOHTUS_제품매칭표_{date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        uploaded = st.file_uploader("수정한 제품 매칭표 엑셀 업로드", type=["xlsx"], key="product_matching_upload")
        if uploaded is not None and st.button("업로드 파일로 제품 매칭표 업데이트", type="primary", use_container_width=True):
            try:
                u, i, sk = import_product_master_excel(uploaded)
                total = u + i + sk
                st.success("✅ 제품 매칭표 업데이트 완료")
                st.markdown(f"""
                - 총 처리 : **{total}건**
                - 업로드 파일 기준 반영 : **{i}건**
                - 건너뜀 : **{sk}건**
                - 반영 방식 : **기존 제품매칭표 완전 교체**
                """)
                st.rerun()
            except Exception as e:
                st.error(f"업로드 실패: {e}")

    st.markdown("### 🔍 매칭 이상 검사")
    conflict_df = find_mapping_conflicts_from_inventory()
    if conflict_df.empty:
        st.success("확인이 필요한 ERP명/비자료명 공유 충돌이 없습니다.")
    else:
        st.warning("같은 ERP명/비자료명이 기존 다른 표준제품명에도 등록되어 있습니다. 실제 충돌인지, 여러 표준제품명이 함께 쓰는 것이 맞는지 확인하세요.")
        edit_df = conflict_df.copy()
        edit_df.insert(0, "문제없음", False)
        edited_conflict = st.data_editor(
            edit_df,
            hide_index=True,
            use_container_width=True,
            disabled=[c for c in edit_df.columns if c != "문제없음"],
            column_config={"문제없음": st.column_config.CheckboxColumn("문제없음", help="이 공유 매칭이 의도된 것이면 체크하세요.")},
            key="pm_conflict_editor",
        )
        st.caption("문제없음으로 확인한 ERP명/비자료명은 시스템이 기억하며, 이후 같은 원본명으로는 다시 묻지 않습니다.")
        if st.button("체크한 공유 매칭 확인 완료", type="primary", use_container_width=True):
            checked = edited_conflict[edited_conflict["문제없음"] == True]
            if checked.empty:
                st.info("체크된 행이 없습니다.")
            else:
                for _, rr in checked.iterrows():
                    approve_mapping_conflict(str(rr.get("사업장", "")), str(rr.get("ERP명/비자료명", "")))
                st.success(f"{len(checked)}건을 문제없음으로 확인했습니다.")
                st.rerun()

    st.markdown("### 제품 매칭표 보완용 파일")
    st.caption("제품매칭표 전체를 내려받아 누락된 ERP명/비자료명 정보를 보완한 뒤 다시 업로드할 수 있습니다.")
    st.download_button(
        "제품 매칭표 보완용 파일 내려받기",
        data=product_master_excel_bytes(highlight_missing=True),
        file_name=f"NOHTUS_제품매칭표_보완용_{date.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.markdown("### 제품 매칭표 수정")
    df = q("""SELECT id, standard_name AS 표준제품명, erp_nohtuspharm_name AS '노투스팜 ERP명', product_code AS '노투스팜 ERP 제품코드', erp_noh_name AS 'NOH ERP명', erp_noh_code AS 'NOH ERP 제품코드', erp_nohtus_name AS '노투스 ERP명', bidata_name AS '비자료명', aliases AS 별칭
              FROM products ORDER BY standard_name""")
    if df.empty:
        st.info("등록된 제품이 없습니다.")
        return

    term = st.text_input("수정할 제품 검색", placeholder="표준제품명/ERP명/별칭 일부 입력", key="pm_edit_term")
    shown = df.copy()
    if term.strip():
        mask = False
        for col in ["표준제품명","노투스팜 ERP명","노투스팜 ERP 제품코드","NOH ERP명","NOH ERP 제품코드","노투스 ERP명","비자료명","별칭"]:
            mask = mask | shown[col].fillna("").astype(str).str.contains(term.strip(), case=False, regex=False)
        shown = shown[mask]
    if shown.empty:
        st.info("검색 결과가 없습니다.")
        return

    choice_options = [""] + [f"{int(r.id)} | {r.표준제품명}" for r in shown.itertuples()]
    choice_label = st.selectbox("수정할 제품 선택", choice_options, index=0, key="pm_edit_choice", format_func=lambda x: "제품명을 입력하거나 선택하세요" if x == "" else x)
    if not choice_label:
        return

    pid = int(choice_label.split(" | ")[0])
    row = df[df["id"] == pid].iloc[0]
    edit_key = f"pm_edit_{pid}"
    ec1, ec2, ec3, ec4 = st.columns(4)
    with ec1:
        e_std = st.text_input("표준제품명", value=str(row["표준제품명"] or ""), key=f"{edit_key}_std")
        e_alias = st.text_input("별칭", value=str(row["별칭"] or ""), key=f"{edit_key}_alias")
    with ec2:
        e_np = st.text_input("노투스팜 ERP명", value=str(row["노투스팜 ERP명"] or ""), key=f"{edit_key}_np")
        e_code = st.text_input("노투스팜 ERP 제품코드", value=str(row["노투스팜 ERP 제품코드"] or ""), key=f"{edit_key}_code")
    with ec3:
        e_noh = st.text_input("NOH ERP명", value=str(row["NOH ERP명"] or ""), key=f"{edit_key}_noh")
        e_noh_code = st.text_input("NOH ERP 제품코드", value=str(row["NOH ERP 제품코드"] or ""), key=f"{edit_key}_noh_code")
    with ec4:
        e_nt = st.text_input("노투스 ERP명", value=str(row["노투스 ERP명"] or ""), key=f"{edit_key}_nt")
        e_bidata = st.text_input("비자료명", value=str(row["비자료명"] or ""), key=f"{edit_key}_bidata")

    save_col, delete_col = st.columns(2)
    with save_col:
        if st.button("제품명 수정", type="primary", use_container_width=True, key=f"{edit_key}_save"):
            old_std = str(row["표준제품명"] or "").strip()
            new_std = e_std.strip()
            if not new_std:
                st.error("표준제품명은 비워둘 수 없습니다.")
            else:
                exec_sql("""UPDATE products SET standard_name=?, warehouse_name=?, aliases=?, product_code=?, erp_nohtuspharm_name=?, erp_noh_name=?, erp_noh_code=?, erp_nohtus_name=?, bidata_name=? WHERE id=?""",
                         (new_std, new_std, e_alias.strip(), str(e_code).strip(), e_np.strip(), e_noh.strip(), str(e_noh_code).strip(), e_nt.strip(), e_bidata.strip(), pid))
                apply_standard_name_change(old_std, new_std)
                st.success("수정했습니다. 이미 등록된 재고/이력 화면에도 변경된 표준제품명을 반영했습니다.")
                st.rerun()
    with delete_col:
        if st.button("제품명 삭제", type="secondary", use_container_width=True, key=f"{edit_key}_delete"):
            st.session_state["confirm_delete_product_id"] = pid
            st.rerun()

    if st.session_state.get("confirm_delete_product_id") == pid:
        st.warning("정말로 삭제하시겠습니까?")
        dc1, dc2 = st.columns(2)
        with dc1:
            if st.button("취소", use_container_width=True):
                st.session_state.pop("confirm_delete_product_id", None)
                st.rerun()
        with dc2:
            if st.button("삭제", type="primary", use_container_width=True):
                try:
                    delete_product(pid)
                    st.session_state.pop("confirm_delete_product_id", None)
                    st.success("제품을 삭제했습니다.")
                    st.rerun()
                except Exception as e:
                    st.error(str(e))

    render_inventory_metadata_editor()


def render_inventory_metadata_editor():
    st.markdown("---")
    st.markdown("### 재고정보 수정")
    st.caption("기존 재고 정보 자체가 잘못된 경우 제조번호/유통기한만 정정합니다. 수량은 변경하지 않습니다.")

    term = st.text_input("재고 검색", placeholder="제품명, ERP명, 제조번호, 유통기한, 로케이션 일부 입력", key="inv_meta_edit_term")
    where = "WHERE qty <> 0"
    params = []
    if term.strip():
        like = f"%{term.strip()}%"
        where += " AND (product_name LIKE ? OR IFNULL(warehouse_name,'') LIKE ? OR IFNULL(lot,'') LIKE ? OR IFNULL(exp_date,'') LIKE ? OR location LIKE ?)"
        params.extend([like, like, like, like, like])
    inv = q(f"""
        SELECT id, company, product_name, warehouse_name, lot, exp_date, location, qty
        FROM inventory
        {where}
        ORDER BY product_name, company, location, lot, exp_date
        LIMIT 300
    """, tuple(params))
    if inv.empty:
        st.info("수정할 재고가 없습니다.")
        return

    labels = []
    for r in inv.itertuples(index=False):
        wh = getattr(r, "warehouse_name") or "-"
        labels.append(
            f"#{int(getattr(r, 'id'))} / {getattr(r, 'company')} / {getattr(r, 'location')} / "
            f"{getattr(r, 'product_name')} / {wh} / LOT:{getattr(r, 'lot') or '-'} / "
            f"EXP:{display_date_only(getattr(r, 'exp_date') or '-')} / {int(getattr(r, 'qty') or 0)}EA"
        )
    selected = st.selectbox("수정할 재고 선택", labels, key="inv_meta_edit_select")
    row = inv.iloc[labels.index(selected)]

    with st.form("inv_meta_edit_form"):
        c1, c2 = st.columns(2)
        with c1:
            new_lot = st.text_input("제조번호/LOT", value=str(row.get("lot") or "-"))
        with c2:
            new_exp = st.text_input("유통기한", value=str(row.get("exp_date") or "-"), placeholder="예: 28/3/2, 2028-03-02")
        edit_memo = st.text_input("수정 사유/메모", value="")
        submitted = st.form_submit_button("재고정보 수정 저장", type="primary", use_container_width=True)

    if submitted:
        try:
            update_inventory_metadata(int(row.get("id")), new_lot, new_exp, edit_memo)
            st.success("재고 제조번호/유통기한 수정 완료")
            st.rerun()
        except Exception as e:
            st.error(str(e))

def page_erp_data_upload():
    st.title("ERP 데이터 업로드")
    st.caption("매일 아침 노투스팜/NOH/노투스 ERP 데이터를 각각 업로드합니다.")
    upload_dir = Path(__file__).parent / "data" / "erp_uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    upload_area = st.container()
    with upload_area:
        cols = st.columns(3, gap="large")
        info = [("노투스팜", "SIMS"), ("NOH", "SIMS"), ("노투스", "IBK우리은행 전산")]
        for col, (company, system_name) in zip(cols, info):
            with col:
                st.markdown(f"### {company}")
                st.caption(system_name)
                up = st.file_uploader(f"{company} ERP 엑셀", type=["xlsx","xls"], key=f"erp_upload_{company}")
                if up is not None:
                    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    safe = f"{company}_{stamp}_{up.name}"
                    path = upload_dir / safe
                    path.write_bytes(up.getvalue())
                    st.success("업로드 파일 저장 완료")
                    try:
                        df = pd.read_excel(BytesIO(up.getvalue()))
                        st.dataframe(df.head(10), hide_index=True, use_container_width=True)
                    except Exception as e:
                        st.warning(f"미리보기 실패: {e}")
    st.markdown("---")
    files = sorted(upload_dir.glob("*"), reverse=True)
    st.markdown("### 최근 업로드 파일")
    if not files:
        st.info("아직 업로드된 ERP 파일이 없습니다.")
    else:
        recent = pd.DataFrame([{"파일명": f.name, "크기KB": round(f.stat().st_size/1024,1)} for f in files[:30]])
        st.dataframe(recent, hide_index=True, use_container_width=True)



def _infer_customer_from_title(title, customers_df=None):
    """출고지시서 제목에서 거래처명을 추정한다.
    제목 규칙: [출고처] [첫 제품명] 외 x품목.
    거래처 관리에 등록된 이름 중 title 시작과 일치하는 가장 긴 이름을 우선 사용한다.
    """
    title = str(title or "").strip()
    if not title:
        return "", ""
    if customers_df is None:
        customers_df = q("SELECT customer_name, manager FROM customers ORDER BY LENGTH(customer_name) DESC")
    if not customers_df.empty:
        for r in customers_df.itertuples():
            name = str(getattr(r, "customer_name", "") or "").strip()
            if name and title.startswith(name):
                return name, str(getattr(r, "manager", "") or "")
    # 거래처 관리에 없으면 제목 첫 토큰을 임시 출고처로 사용한다.
    return title.split()[0] if title.split() else title, ""


def _extract_inbound_source_from_memo(memo):
    """입고 이력 memo에서 입고처만 추출한다.
    저장 형식 예: '매입처: 거래처명 / 기타메모'
    """
    text = str(memo or "").strip()
    if not text or text == "입고 등록":
        return ""
    prefixes = ["매입처:", "입고처:"]
    for prefix in prefixes:
        if text.startswith(prefix):
            text = text[len(prefix):].strip()
            if " / " in text:
                text = text.split(" / ", 1)[0].strip()
            break
    return text

def page_closing():
    st.title("마감")
    st.caption("출고의 마지막 단계입니다. 오늘 출고 체크, ERP 재고 비교, 업무일지 작성 기능을 한 화면에서 전환합니다.")
    tab = st.radio("마감", ["오늘 출고 체크", "ERP 재고 비교", "업무일지 작성"], horizontal=True, key="closing_sub")

    if tab == "ERP 재고 비교":
        page_erp_stock_compare()
        return

    target_date = st.date_input("기준일", value=date.today(), key="closing_date")
    ds = str(target_date)

    if tab == "오늘 출고 체크":
        items = q("""SELECT o.id AS 지시서번호,
                            COALESCE(o.title, '') AS 출고지시서제목,
                            i.inventory_id AS 재고ID,
                            i.company AS 사업장,
                            i.location AS 로케이션,
                            i.product_name AS 표준제품명,
                            COALESCE(i.lot, '-') AS 제조번호,
                            COALESCE(i.exp_date, '-') AS 유통기한,
                            i.qty AS 출고수량,
                            COALESCE(inv.qty, 0) AS 현재수량
                     FROM outbound_orders o
                     JOIN outbound_order_items i ON o.id=i.order_id
                     LEFT JOIN inventory inv ON i.inventory_id=inv.id
                     WHERE o.order_date=? AND IFNULL(o.status,'')<>'취소됨'
                     ORDER BY o.id, i.company, i.location, i.product_name, i.lot, i.exp_date""", (ds,))
        if items.empty:
            st.info("해당 날짜의 출고지시가 없습니다.")
        else:
            items["유통기한"] = items["유통기한"].apply(display_date_only)
            items["출고수량"] = pd.to_numeric(items["출고수량"], errors="coerce").fillna(0).astype(int)
            items["현재수량"] = pd.to_numeric(items["현재수량"], errors="coerce").fillna(0).astype(int)
            valid_inv = items["재고ID"].notna()
            items["동일재고출고합계"] = items.groupby("재고ID")["출고수량"].transform("sum").where(valid_inv, items["출고수량"])
            # 오늘 출고 체크는 실제 지시내역 확인용이다.
            # 현재수량은 출고지시 저장으로 이미 차감된 뒤의 inventory 수량이며,
            # 기존수량은 오늘 해당 재고ID의 출고수량을 되돌려 계산한 출고 전 수량이다.
            items["기존수량"] = items["현재수량"] + items["동일재고출고합계"]
            items["실물수량"] = ""
            show_cols = ["지시서번호", "사업장", "로케이션", "표준제품명", "제조번호", "유통기한", "기존수량", "출고수량", "현재수량", "실물수량"]
            st.dataframe(items[show_cols], hide_index=True, use_container_width=True, column_config={"지시서번호": st.column_config.NumberColumn(width="small"), "사업장": st.column_config.TextColumn(width="small"), "로케이션": st.column_config.TextColumn(width="small"), "표준제품명": st.column_config.TextColumn(width="large"), "제조번호": st.column_config.TextColumn(width="medium"), "유통기한": st.column_config.TextColumn(width="small"), "기존수량": st.column_config.NumberColumn(width="small"), "출고수량": st.column_config.NumberColumn(width="small"), "현재수량": st.column_config.NumberColumn(width="small"), "실물수량": st.column_config.TextColumn(width="small")})
            st.download_button("마감 체크리스트 엑셀 다운로드", data=dataframe_to_excel_bytes(items[show_cols], "마감체크"), file_name=f"NOHTUS_마감체크_{ds}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    else:
        st.markdown("### 출고")
        out_raw = q("""SELECT o.id AS 지시서번호, COALESCE(o.title, '') AS 출고지시서제목,
                              i.product_name AS 표준제품명, i.qty AS 수량
                       FROM outbound_orders o
                       JOIN outbound_order_items i ON o.id=i.order_id
                       WHERE o.order_date=? AND IFNULL(o.status,'')<>'취소됨'
                       ORDER BY o.id, i.id""", (ds,))
        if out_raw.empty: st.info("출고 업무일지 데이터가 없습니다.")
        else:
            customers_for_worklog = q("SELECT customer_name, manager FROM customers ORDER BY LENGTH(customer_name) DESC")
            out_raw[["출고처", "담당자"]] = out_raw["출고지시서제목"].apply(
                lambda x: pd.Series(_infer_customer_from_title(x, customers_for_worklog))
            )
            out = (out_raw.assign(내역=out_raw["표준제품명"].astype(str) + " * " + out_raw["수량"].astype(int).astype(str))
                         .groupby(["출고처", "담당자"], as_index=False)["내역"]
                         .agg(lambda x: ", ".join(x)))
            tsv = out.to_csv(sep='	', index=False, header=False)
            st.text_area("드래그해서 복사", value=tsv, height=140, key="worklog_out_tsv")

        st.markdown("### 입고")
        inbound_raw = q("""SELECT COALESCE(t.memo,'') AS 메모,
                                 COALESCE(t.to_location, '') AS 적치위치,
                                 t.product_name AS 표준제품명,
                                 COALESCE(t.lot, '-') AS 제조번호,
                                 COALESCE(t.exp_date, '-') AS 유통기한,
                                 t.qty AS 수량
                          FROM transactions t
                          WHERE t.tx_type='입고' AND substr(t.created_at,1,10)=? ORDER BY t.id""", (ds,))
        if inbound_raw.empty: st.info("입고 업무일지 데이터가 없습니다.")
        else:
            inbound_raw["매입처"] = inbound_raw["메모"].apply(_extract_inbound_source_from_memo)
            inbound_raw["적치위치"] = inbound_raw["적치위치"].astype(str).replace("", "-")
            inbound_raw["제조번호"] = inbound_raw["제조번호"].astype(str).replace("", "-")
            inbound_raw["유통기한"] = inbound_raw["유통기한"].apply(display_date_only)
            inbound_raw["수량"] = pd.to_numeric(inbound_raw["수량"], errors="coerce").fillna(0).astype(int)
            inbound_raw["내역"] = inbound_raw["표준제품명"].astype(str) + " * " + inbound_raw["수량"].astype(str)
            inbound = inbound_raw[["매입처", "적치위치", "내역", "제조번호", "유통기한"]]
            tsv = inbound.to_csv(sep='\t', index=False, header=False)
            st.text_area("드래그해서 복사", value=tsv, height=140, key="worklog_in_tsv")

        st.markdown("### 이동")
        moves = q("""SELECT product_name || '*' || qty AS 이동내역, COALESCE(from_location,'') || ' → ' || COALESCE(to_location,'') AS 이동위치
                     FROM transactions WHERE tx_type='이동' AND substr(created_at,1,10)=? ORDER BY id""", (ds,))
        if moves.empty: st.info("이동 업무일지 데이터가 없습니다.")
        else:
            tsv = moves.to_csv(sep='\t', index=False, header=False)
            st.text_area("드래그해서 복사", value=tsv, height=120, key="worklog_move_tsv")


# ---------------- style/nav ----------------
def apply_style():
    st.markdown("""
    <style>
    .stApp {background:#f8fafc;}
    h1,h2,h3 {color:#111827;}
    section[data-testid="stSidebar"] {background:linear-gradient(180deg,#08213d,#103e69); color:white; font-size:123%; width:255px!important; min-width:255px!important;}
    section[data-testid="stSidebar"] > div:first-child {width:255px!important; min-width:255px!important;}
    .block-container {padding-left:2.3rem!important; padding-right:1.75rem!important;}
    section[data-testid="stSidebar"] * {color:white;}
    section[data-testid="stSidebar"] .stButton > button {background:transparent!important;border:0!important;color:white!important;text-align:left!important;justify-content:flex-start!important;border-radius:10px!important;padding:8px 10px!important;font-weight:800!important;font-size:123%!important;}
    section[data-testid="stSidebar"] .stButton > button p {text-align:left!important;width:100%!important;}
    section[data-testid="stSidebar"] div[data-testid="stButton"] > button[kind="primary"], section[data-testid="stSidebar"] div[data-testid="stButton"] > button {text-align:left!important;justify-content:flex-start!important;}
    section[data-testid="stSidebar"] .stButton > button:hover {background:rgba(96,165,250,.35);color:white;border:0;}
    .nav-active button {background:#2563eb!important;}

    .legend-wrap {display:flex;gap:12px;flex-wrap:wrap;margin:4px 0 14px 0;}
    .legend-chip {display:flex;align-items:center;gap:8px;border:1px solid #dbe4f0;background:#fff;border-radius:12px;padding:9px 14px;font-weight:800;color:#111827;}
    .swatch {width:18px;height:18px;border-radius:5px;border:1px solid rgba(15,23,42,.12);display:inline-block;}
    .swatch.y {background:#fff39b}.swatch.b{background:#68d2e7}.swatch.p{background:#f0a7e6}.swatch.g{background:#f3f4f6}
    .map-card {background:#fff;border:1px solid #dbe4f0;border-radius:18px;padding:16px;box-shadow:0 8px 24px rgba(15,23,42,.06);}
    .rack-title {text-align:center;font-size:11px;color:#64748b;font-weight:800;margin-top:2px;height:0;overflow:visible;}
    .mapbtn-wrap {position:relative;margin:0;}
    .mapbtn-wrap .stButton > button {height:42px;padding:0!important;border-radius:0!important;border:1px solid rgba(51,65,85,.34)!important;color:#0f172a!important;font-weight:900!important;font-size:13px!important;box-shadow:none!important;}
    .mapbtn-wrap.yellow .stButton > button{background:#fff39b!important}.mapbtn-wrap.blue .stButton > button{background:#68d2e7!important}.mapbtn-wrap.pink .stButton > button{background:#f0a7e6!important}.mapbtn-wrap.gray .stButton > button{background:#f7f8fa!important}.mapbtn-wrap.bidata .stButton > button{background:#d1d5db!important}.mapbtn-wrap.white .stButton > button{background:#fff!important}
    .mapbtn-wrap .stButton > button:hover{outline:3px solid rgba(37,99,235,.22)!important;z-index:2;position:relative;}
    .mapbtn-wrap.selected .stButton > button{outline:3px solid #2563eb!important;}
    .mapbtn-wrap.has-stock:after{content:'';position:absolute;right:8px;top:8px;width:9px;height:9px;border-radius:999px;background:#65d84f;border:1.5px solid #166534;z-index:4;pointer-events:none;}
    .gbox {border:1px solid #dbe4f0;border-radius:12px;overflow:hidden;background:#fff;}
    .blank-g{height:78px;border-bottom:1px solid #dbe4f0}.box-label{text-align:center;padding:10px;font-weight:900;}
    .center-label{text-align:center;font-weight:900;font-size:13px;margin-top:8px;color:#111827}.memo{padding:46px 8px 8px 8px;line-height:2.8;color:#334155;font-size:14px;white-space:nowrap}.qptext{height:42px;display:flex;align-items:center;font-weight:900;border:1px solid #e2e8f0;border-left:0;padding-left:8px;background:#fff;}

    .map-detail-title-wrap div[data-testid="stButton"] > button {font-size:12pt!important;font-weight:400!important;text-align:center!important;justify-content:center!important;border:0!important;background:transparent!important;color:#111827!important;box-shadow:none!important;padding:0.15rem 0!important;}
    .detail-total-text {display:flex;gap:8px;align-items:baseline;justify-content:center;color:#334155;font-size:13px;margin:2px auto 10px;}
    .detail-total-text strong {font-weight:600;color:#111827;}
    .popup-box{background:#ffffff;border:1px solid #bfdbfe;border-radius:16px;padding:14px;margin:8px 0 18px 0;box-shadow:0 10px 28px rgba(37,99,235,.08)}
    .zone-pill {display:inline-block;background:#e8f5ee;color:#15803d;font-weight:800;border-radius:10px;padding:6px 10px;margin-bottom:10px;}
    .detail-card {background:white;border:1px solid #dbe4f0;border-radius:14px;padding:12px;margin:8px 0;box-shadow:0 5px 16px rgba(15,23,42,.05);}
    .card-top {display:flex;justify-content:space-between;align-items:center;gap:8px;}.company-badge {display:inline-block;background:#eff6ff;color:#1d4ed8;font-weight:800;border-radius:999px;padding:3px 8px;font-size:12px;}.product-title {font-weight:400;font-size:14px;margin-top:8px;color:#111827;}.muted {color:#64748b;font-size:12px;line-height:1.6;}.qty-text {font-weight:900;color:#111827;white-space:nowrap;}.photo-box{width:250px;height:250px;margin-left:auto;margin-right:auto;border:1px dashed #cbd5e1;border-radius:16px;background:#f8fafc;display:flex;align-items:center;justify-content:center;color:#94a3b8;font-weight:800;margin-bottom:10px;}
    div[data-testid="stMetric"] {background:white;border:1px solid #dbe4f0;border-radius:16px;padding:16px;box-shadow:0 8px 20px rgba(15,23,42,.05)}
    .mini-cal {background:#fff;border:1px solid #dbe4f0;border-radius:16px;padding:14px;margin:8px 0 16px 0;box-shadow:0 8px 20px rgba(15,23,42,.05)}
    .mini-cal-head {font-weight:900;margin-bottom:10px;color:#111827}.mini-grid{display:grid;grid-template-columns:repeat(7,1fr);gap:6px;margin-bottom:6px}.mini-week span{text-align:center;color:#64748b;font-size:12px;font-weight:900}.cal-day{height:34px;display:flex;align-items:center;justify-content:center;border-radius:10px;background:#f8fafc;color:#334155;font-weight:800;position:relative}.cal-day.on{background:#2563eb;color:white;box-shadow:0 0 0 3px rgba(37,99,235,.15)}.cal-day small{position:absolute;right:4px;top:3px;font-size:10px;background:white;color:#2563eb;border-radius:999px;min-width:15px;height:15px;line-height:15px;text-align:center}.cal-day.empty{background:transparent}
    section[data-testid="stSidebar"] div[data-testid="stButton"] > button, section[data-testid="stSidebar"] div[data-testid="stButton"] > button[kind="secondary"], section[data-testid="stSidebar"] div[data-testid="stButton"] > button[kind="primary"]{background:transparent!important;color:white!important;border:0!important;}
    section[data-testid="stSidebar"] div[data-testid="stButton"] > button p{color:white!important;}
    </style>
    """, unsafe_allow_html=True)

def nav_button(name):
    if st.sidebar.button(name, key=f"nav_{name}", use_container_width=True):
        st.session_state["page"] = name
        st.rerun()



# ---------------- v4.1 reorganized master / ERP / closing ----------------
def page_erp_stock_compare():
    st.title("ERP 재고 비교")
    st.caption("ERP/WMS 모두 해당 사업장의 ERP명 기준으로 제조번호·유통기한 없이 총수량만 비교합니다. 비자료는 비교 대상에서 제외합니다.")

    parsed = []
    cols = st.columns(3, gap="large")
    info = [("노투스팜", "SIMS"), ("NOH", "SIMS"), ("노투스", "IBK우리은행 전산")]
    for col, (company, system_name) in zip(cols, info):
        with col:
            st.markdown(f"### {company}")
            st.caption(system_name)
            up = st.file_uploader(f"{company} 현재고 엑셀", type=["xlsx", "xls"], key=f"erp_compare_upload_{company}")
            if up is not None:
                parsed.append(_read_erp_current_file(up, company, "compare"))
                st.success("파일 선택됨")
            else:
                parsed.append(None)

    if not any(parsed):
        st.info("노투스팜 / NOH / 노투스 ERP 현재고 엑셀을 업로드하세요.")
        return

    _, run_col, _ = st.columns([3, 2, 3], gap="large")
    with run_col:
        run_compare = st.button("ERP 재고 비교 실행", type="primary", use_container_width=True)

    if run_compare:
        erp, amb, fail, auto_count = _normalize_erp_current_rows(parsed)

        wms_raw = q("""SELECT company AS 사업장, product_name AS 표준제품명, SUM(qty) AS 수량
                       FROM inventory
                       WHERE qty<>0 AND company IN ('노투스팜','NOH','노투스')
                       GROUP BY company, product_name""")
        if wms_raw.empty:
            wms_sum = pd.DataFrame(columns=["사업장", "제품명", "WMS수량"])
        else:
            wms_raw["제품명"] = wms_raw.apply(lambda r: clean_excel_text(product_compare_name_for(r["사업장"], r["표준제품명"])), axis=1)
            wms_raw = wms_raw[~wms_raw["제품명"].apply(is_ignored_erp_product_name)]
            wms_sum = (
                wms_raw.groupby(["사업장", "제품명"], as_index=False)["수량"]
                .sum()
                .rename(columns={"수량": "WMS수량"})
            )

        comp = wms_sum.merge(erp, how="outer", on=["사업장", "제품명"])
        if comp.empty:
            comp = pd.DataFrame(columns=["사업장", "제품명", "ERP수량", "WMS수량", "차이"])
        else:
            comp["ERP수량"] = comp["ERP수량"].fillna(0).astype(int)
            comp["WMS수량"] = comp["WMS수량"].fillna(0).astype(int)
            comp["차이"] = comp["WMS수량"] - comp["ERP수량"]
            comp = comp[["사업장", "제품명", "ERP수량", "WMS수량", "차이"]].sort_values(["사업장", "제품명"])

        st.session_state["erp_compare_rows"] = comp.to_dict("records")
        st.session_state["erp_compare_amb"] = amb.to_dict("records") if not amb.empty else []
        st.session_state["erp_compare_fail"] = fail.to_dict("records") if not fail.empty else []
        st.session_state["erp_compare_summary"] = {"auto": auto_count, "amb": len(amb), "fail": len(fail)}

    if "erp_compare_summary" in st.session_state:
        sm = st.session_state["erp_compare_summary"]
        m1, m2, m3 = st.columns(3)
        m1.metric("ERP 행 반영", f"{sm['auto']}건")
        m2.metric("확인 필요", f"{sm['amb']}건")
        m3.metric("매칭 실패", f"{sm['fail']}건")
        comp = pd.DataFrame(st.session_state.get("erp_compare_rows", []))
        amb = pd.DataFrame(st.session_state.get("erp_compare_amb", []))
        fail = pd.DataFrame(st.session_state.get("erp_compare_fail", []))
        if not comp.empty:
            st.markdown("### 재고 비교 결과")
            only_diff = st.checkbox("차이 있는 항목만 보기", value=True, key="erp_compare_only_diff")
            shown = comp[comp["차이"] != 0] if only_diff else comp
            st.dataframe(shown, use_container_width=True, hide_index=True)
            st.download_button(
                "비교 결과 엑셀 다운로드",
                data=dataframe_to_excel_bytes(comp, "ERP_WMS_비교"),
                file_name=f"NOHTUS_ERP_WMS_비교_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        if not amb.empty:
            st.warning("확인 필요 항목")
            st.dataframe(amb, use_container_width=True, hide_index=True)
        if not fail.empty:
            st.error("매칭 실패 항목")
            st.dataframe(fail, use_container_width=True, hide_index=True)


def page_product_matching():
    st.title("제품 매칭 관리")
    st.caption("표준제품명과 사업장별 ERP명/비자료명을 관리합니다.")

    action_col, _spacer = st.columns([4, 6], gap="large")
    with action_col:
        st.download_button(
            "제품 매칭표 엑셀 파일 내려받기",
            data=product_master_excel_bytes(),
            file_name=f"NOHTUS_제품매칭표_{date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        uploaded = st.file_uploader("수정한 제품 매칭표 엑셀 업로드", type=["xlsx"], key="product_matching_upload")
        if uploaded is not None and st.button("업로드 파일로 제품 매칭표 업데이트", type="primary", use_container_width=True):
            try:
                u, i, sk = import_product_master_excel(uploaded)
                total = u + i + sk
                st.success("✅ 제품 매칭표 업데이트 완료")
                st.markdown(f"""
                - 총 처리 : **{total}건**
                - 업로드 파일 기준 반영 : **{i}건**
                - 건너뜀 : **{sk}건**
                - 반영 방식 : **기존 제품매칭표 완전 교체**
                """)
                st.rerun()
            except Exception as e:
                st.error(f"업로드 실패: {e}")

    st.markdown("### 🔍 매칭 이상 검사")
    conflict_df = find_mapping_conflicts_from_inventory()
    if conflict_df.empty:
        st.success("확인이 필요한 ERP명/비자료명 공유 충돌이 없습니다.")
    else:
        st.warning("같은 ERP명/비자료명이 기존 다른 표준제품명에도 등록되어 있습니다. 실제 충돌인지, 여러 표준제품명이 함께 쓰는 것이 맞는지 확인하세요.")
        edit_df = conflict_df.copy()
        edit_df.insert(0, "문제없음", False)
        edited_conflict = st.data_editor(
            edit_df,
            hide_index=True,
            use_container_width=True,
            disabled=[c for c in edit_df.columns if c != "문제없음"],
            column_config={"문제없음": st.column_config.CheckboxColumn("문제없음", help="이 공유 매칭이 의도된 것이면 체크하세요.")},
            key="pm_conflict_editor",
        )
        st.caption("문제없음으로 확인한 ERP명/비자료명은 시스템이 기억하며, 이후 같은 원본명으로는 다시 묻지 않습니다.")
        if st.button("체크한 공유 매칭 확인 완료", type="primary", use_container_width=True):
            checked = edited_conflict[edited_conflict["문제없음"] == True]
            if checked.empty:
                st.info("체크된 행이 없습니다.")
            else:
                for _, rr in checked.iterrows():
                    approve_mapping_conflict(str(rr.get("사업장", "")), str(rr.get("ERP명/비자료명", "")))
                st.success(f"{len(checked)}건을 문제없음으로 확인했습니다.")
                st.rerun()

    st.markdown("### 제품 매칭표 보완용 파일")
    st.caption("제품매칭표 전체를 내려받아 누락된 ERP명/비자료명 정보를 보완한 뒤 다시 업로드할 수 있습니다.")
    st.download_button(
        "제품 매칭표 보완용 파일 내려받기",
        data=product_master_excel_bytes(highlight_missing=True),
        file_name=f"NOHTUS_제품매칭표_보완용_{date.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.markdown("### 제품 매칭표 수정")
    df = q("""SELECT id, standard_name AS 표준제품명, erp_nohtuspharm_name AS '노투스팜 ERP명', product_code AS '노투스팜 ERP 제품코드', erp_noh_name AS 'NOH ERP명', erp_noh_code AS 'NOH ERP 제품코드', erp_nohtus_name AS '노투스 ERP명', bidata_name AS '비자료명', aliases AS 별칭
              FROM products ORDER BY standard_name""")
    if df.empty:
        st.info("등록된 제품이 없습니다.")
        return

    term = st.text_input("수정할 제품 검색", placeholder="표준제품명/ERP명/별칭 일부 입력", key="pm_edit_term")
    shown = df.copy()
    if term.strip():
        mask = False
        for col in ["표준제품명","노투스팜 ERP명","노투스팜 ERP 제품코드","NOH ERP명","NOH ERP 제품코드","노투스 ERP명","비자료명","별칭"]:
            mask = mask | shown[col].fillna("").astype(str).str.contains(term.strip(), case=False, regex=False)
        shown = shown[mask]
    if shown.empty:
        st.info("검색 결과가 없습니다.")
        return

    choice_options = [""] + [f"{int(r.id)} | {r.표준제품명}" for r in shown.itertuples()]
    choice_label = st.selectbox("수정할 제품 선택", choice_options, index=0, key="pm_edit_choice", format_func=lambda x: "제품명을 입력하거나 선택하세요" if x == "" else x)
    if not choice_label:
        return

    pid = int(choice_label.split(" | ")[0])
    row = df[df["id"] == pid].iloc[0]
    edit_key = f"pm_edit_{pid}"
    ec1, ec2, ec3, ec4 = st.columns(4)
    with ec1:
        e_std = st.text_input("표준제품명", value=str(row["표준제품명"] or ""), key=f"{edit_key}_std")
        e_alias = st.text_input("별칭", value=str(row["별칭"] or ""), key=f"{edit_key}_alias")
    with ec2:
        e_np = st.text_input("노투스팜 ERP명", value=str(row["노투스팜 ERP명"] or ""), key=f"{edit_key}_np")
        e_code = st.text_input("노투스팜 ERP 제품코드", value=str(row["노투스팜 ERP 제품코드"] or ""), key=f"{edit_key}_code")
    with ec3:
        e_noh = st.text_input("NOH ERP명", value=str(row["NOH ERP명"] or ""), key=f"{edit_key}_noh")
        e_noh_code = st.text_input("NOH ERP 제품코드", value=str(row["NOH ERP 제품코드"] or ""), key=f"{edit_key}_noh_code")
    with ec4:
        e_nt = st.text_input("노투스 ERP명", value=str(row["노투스 ERP명"] or ""), key=f"{edit_key}_nt")
        e_bidata = st.text_input("비자료명", value=str(row["비자료명"] or ""), key=f"{edit_key}_bidata")

    save_col, delete_col = st.columns(2)
    with save_col:
        if st.button("제품명 수정", type="primary", use_container_width=True, key=f"{edit_key}_save"):
            old_std = str(row["표준제품명"] or "").strip()
            new_std = e_std.strip()
            if not new_std:
                st.error("표준제품명은 비워둘 수 없습니다.")
            else:
                exec_sql("""UPDATE products SET standard_name=?, warehouse_name=?, aliases=?, product_code=?, erp_nohtuspharm_name=?, erp_noh_name=?, erp_noh_code=?, erp_nohtus_name=?, bidata_name=? WHERE id=?""",
                         (new_std, new_std, e_alias.strip(), str(e_code).strip(), e_np.strip(), e_noh.strip(), str(e_noh_code).strip(), e_nt.strip(), e_bidata.strip(), pid))
                apply_standard_name_change(old_std, new_std)
                st.success("수정했습니다. 이미 등록된 재고/이력 화면에도 변경된 표준제품명을 반영했습니다.")
                st.rerun()
    with delete_col:
        if st.button("제품명 삭제", type="secondary", use_container_width=True, key=f"{edit_key}_delete"):
            st.session_state["confirm_delete_product_id"] = pid
            st.rerun()

    if st.session_state.get("confirm_delete_product_id") == pid:
        st.warning("정말로 삭제하시겠습니까?")
        dc1, dc2 = st.columns(2)
        with dc1:
            if st.button("취소", use_container_width=True):
                st.session_state.pop("confirm_delete_product_id", None)
                st.rerun()
        with dc2:
            if st.button("삭제", type="primary", use_container_width=True):
                try:
                    delete_product(pid)
                    st.session_state.pop("confirm_delete_product_id", None)
                    st.success("제품을 삭제했습니다.")
                    st.rerun()
                except Exception as e:
                    st.error(str(e))

def page_erp_data_upload():
    st.title("ERP 데이터 업로드")
    st.caption("매일 아침 노투스팜/NOH/노투스 ERP 데이터를 각각 업로드합니다.")
    upload_dir = Path(__file__).parent / "data" / "erp_uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    upload_area = st.container()
    with upload_area:
        cols = st.columns(3, gap="large")
        info = [("노투스팜", "SIMS"), ("NOH", "SIMS"), ("노투스", "IBK우리은행 전산")]
        for col, (company, system_name) in zip(cols, info):
            with col:
                st.markdown(f"### {company}")
                st.caption(system_name)
                up = st.file_uploader(f"{company} ERP 엑셀", type=["xlsx","xls"], key=f"erp_upload_{company}")
                if up is not None:
                    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    safe = f"{company}_{stamp}_{up.name}"
                    path = upload_dir / safe
                    path.write_bytes(up.getvalue())
                    st.success("업로드 파일 저장 완료")
                    try:
                        df = pd.read_excel(BytesIO(up.getvalue()))
                        st.dataframe(df.head(10), hide_index=True, use_container_width=True)
                    except Exception as e:
                        st.warning(f"미리보기 실패: {e}")
    st.markdown("---")
    files = sorted(upload_dir.glob("*"), reverse=True)
    st.markdown("### 최근 업로드 파일")
    if not files:
        st.info("아직 업로드된 ERP 파일이 없습니다.")
    else:
        recent = pd.DataFrame([{"파일명": f.name, "크기KB": round(f.stat().st_size/1024,1)} for f in files[:30]])
        st.dataframe(recent, hide_index=True, use_container_width=True)



def _extract_inbound_source_from_memo(memo):
    """입고 이력 memo에서 입고처만 추출한다.
    저장 형식 예: '매입처: 거래처명 / 기타메모'
    """
    text = str(memo or "").strip()
    if not text or text == "입고 등록":
        return ""
    prefixes = ["매입처:", "입고처:"]
    for prefix in prefixes:
        if text.startswith(prefix):
            text = text[len(prefix):].strip()
            if " / " in text:
                text = text.split(" / ", 1)[0].strip()
            break
    return text

def page_closing():
    st.title("마감")
    st.caption("출고의 마지막 단계입니다. 오늘 출고 체크, ERP 재고 비교, 업무일지 작성 기능을 한 화면에서 전환합니다.")
    tab = st.radio("마감", ["오늘 출고 체크", "ERP 재고 비교", "업무일지 작성"], horizontal=True, key="closing_sub")

    if tab == "ERP 재고 비교":
        page_erp_stock_compare()
        return

    target_date = st.date_input("기준일", value=date.today(), key="closing_date")
    ds = str(target_date)

    if tab == "오늘 출고 체크":
        items = q("""SELECT o.id AS 지시서번호,
                            COALESCE(o.title, '') AS 출고지시서제목,
                            i.inventory_id AS 재고ID,
                            i.company AS 사업장,
                            i.location AS 로케이션,
                            i.product_name AS 표준제품명,
                            COALESCE(i.lot, '-') AS 제조번호,
                            COALESCE(i.exp_date, '-') AS 유통기한,
                            i.qty AS 출고수량,
                            COALESCE(inv.qty, 0) AS 현재수량
                     FROM outbound_orders o
                     JOIN outbound_order_items i ON o.id=i.order_id
                     LEFT JOIN inventory inv ON i.inventory_id=inv.id
                     WHERE o.order_date=? AND IFNULL(o.status,'')<>'취소됨'
                     ORDER BY o.id, i.company, i.location, i.product_name, i.lot, i.exp_date""", (ds,))
        if items.empty:
            st.info("해당 날짜의 출고지시가 없습니다.")
        else:
            items["유통기한"] = items["유통기한"].apply(display_date_only)
            items["출고수량"] = pd.to_numeric(items["출고수량"], errors="coerce").fillna(0).astype(int)
            items["현재수량"] = pd.to_numeric(items["현재수량"], errors="coerce").fillna(0).astype(int)
            valid_inv = items["재고ID"].notna()
            items["동일재고출고합계"] = items.groupby("재고ID")["출고수량"].transform("sum").where(valid_inv, items["출고수량"])
            # 오늘 출고 체크는 실제 지시내역 확인용이다.
            # 현재수량은 출고지시 저장으로 이미 차감된 뒤의 inventory 수량이며,
            # 기존수량은 오늘 해당 재고ID의 출고수량을 되돌려 계산한 출고 전 수량이다.
            items["기존수량"] = items["현재수량"] + items["동일재고출고합계"]
            items["실물수량"] = ""
            show_cols = ["지시서번호", "사업장", "로케이션", "표준제품명", "제조번호", "유통기한", "기존수량", "출고수량", "현재수량", "실물수량"]
            st.dataframe(items[show_cols], hide_index=True, use_container_width=True, column_config={"지시서번호": st.column_config.NumberColumn(width="small"), "사업장": st.column_config.TextColumn(width="small"), "로케이션": st.column_config.TextColumn(width="small"), "표준제품명": st.column_config.TextColumn(width="large"), "제조번호": st.column_config.TextColumn(width="medium"), "유통기한": st.column_config.TextColumn(width="small"), "기존수량": st.column_config.NumberColumn(width="small"), "출고수량": st.column_config.NumberColumn(width="small"), "현재수량": st.column_config.NumberColumn(width="small"), "실물수량": st.column_config.TextColumn(width="small")})
            st.download_button("마감 체크리스트 엑셀 다운로드", data=dataframe_to_excel_bytes(items[show_cols], "마감체크"), file_name=f"NOHTUS_마감체크_{ds}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    else:
        st.markdown("### 출고")
        out_raw = q("""SELECT o.id AS 지시서번호, COALESCE(o.title, '') AS 출고지시서제목,
                              i.product_name AS 표준제품명, i.qty AS 수량
                       FROM outbound_orders o
                       JOIN outbound_order_items i ON o.id=i.order_id
                       WHERE o.order_date=? AND IFNULL(o.status,'')<>'취소됨'
                       ORDER BY o.id, i.id""", (ds,))
        if out_raw.empty: st.info("출고 업무일지 데이터가 없습니다.")
        else:
            customers_for_worklog = q("SELECT customer_name, manager FROM customers ORDER BY LENGTH(customer_name) DESC")
            out_raw[["출고처", "담당자"]] = out_raw["출고지시서제목"].apply(
                lambda x: pd.Series(_infer_customer_from_title(x, customers_for_worklog))
            )
            out = (out_raw.assign(내역=out_raw["표준제품명"].astype(str) + " * " + out_raw["수량"].astype(int).astype(str))
                         .groupby(["출고처", "담당자"], as_index=False)["내역"]
                         .agg(lambda x: ", ".join(x)))
            tsv = out.to_csv(sep='	', index=False, header=False)
            st.text_area("드래그해서 복사", value=tsv, height=140, key="worklog_out_tsv")

        st.markdown("### 입고")
        inbound_raw = q("""SELECT COALESCE(t.memo,'') AS 메모,
                                 COALESCE(t.to_location, '') AS 적치위치,
                                 t.product_name AS 표준제품명,
                                 COALESCE(t.lot, '-') AS 제조번호,
                                 COALESCE(t.exp_date, '-') AS 유통기한,
                                 t.qty AS 수량
                          FROM transactions t
                          WHERE t.tx_type='입고' AND substr(t.created_at,1,10)=? ORDER BY t.id""", (ds,))
        if inbound_raw.empty: st.info("입고 업무일지 데이터가 없습니다.")
        else:
            inbound_raw["매입처"] = inbound_raw["메모"].apply(_extract_inbound_source_from_memo)
            inbound_raw["적치위치"] = inbound_raw["적치위치"].astype(str).replace("", "-")
            inbound_raw["제조번호"] = inbound_raw["제조번호"].astype(str).replace("", "-")
            inbound_raw["유통기한"] = inbound_raw["유통기한"].apply(display_date_only)
            inbound_raw["수량"] = pd.to_numeric(inbound_raw["수량"], errors="coerce").fillna(0).astype(int)
            inbound_raw["내역"] = inbound_raw["표준제품명"].astype(str) + " * " + inbound_raw["수량"].astype(str)
            inbound = inbound_raw[["매입처", "적치위치", "내역", "제조번호", "유통기한"]]
            tsv = inbound.to_csv(sep='\t', index=False, header=False)
            st.text_area("드래그해서 복사", value=tsv, height=140, key="worklog_in_tsv")

        st.markdown("### 이동")
        moves = q("""SELECT product_name || '*' || qty AS 이동내역, COALESCE(from_location,'') || ' → ' || COALESCE(to_location,'') AS 이동위치
                     FROM transactions WHERE tx_type='이동' AND substr(created_at,1,10)=? ORDER BY id""", (ds,))
        if moves.empty: st.info("이동 업무일지 데이터가 없습니다.")
        else:
            tsv = moves.to_csv(sep='\t', index=False, header=False)
            st.text_area("드래그해서 복사", value=tsv, height=120, key="worklog_move_tsv")



def main():
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    init_db()
    apply_style()
    st.sidebar.markdown(f"# {APP_TITLE}")
    st.sidebar.caption(VERSION)
    if "page" not in st.session_state:
        st.session_state["page"] = "로케이션 맵"

    try:
        if st.query_params.get("map_search_product", ""):
            st.session_state["page"] = "로케이션 맵"
        elif st.query_params.get("inbound_loc", ""):
            st.session_state["page"] = "입고 등록"
    except Exception:
        pass

    def nav_button(label):
        active = st.session_state.get("page") == label
        if st.sidebar.button(label, use_container_width=True, type="primary" if active else "secondary"):
            st.session_state["page"] = label
            if label == "로케이션 맵":
                st.session_state["_scroll_map_top"] = True
            st.rerun()

    nav_button("로케이션 맵")
    st.sidebar.markdown("### 출고")
    nav_button("출고지시")
    nav_button("저장된 출고지시")
    nav_button("마감")

    st.sidebar.markdown("### 재고")
    nav_button("입고 등록")
    nav_button("이동 등록")
    nav_button("이력 조회")
    nav_button("재고 실사")

    st.sidebar.markdown("### 기초")
    nav_button("제품 매칭 관리")
    nav_button("거래처 관리")

    menu = st.session_state["page"]
    if menu == "로케이션 맵": page_map()
    elif menu == "출고지시": page_outbound()
    elif menu == "저장된 출고지시": page_saved_outbound()
    elif menu == "마감": page_closing()
    elif menu == "입고 등록": page_inbound()
    elif menu == "이동 등록": page_move()
    elif menu == "재고 실사": page_stocktake()
    elif menu == "제품 매칭 관리": page_product_matching()
    elif menu == "거래처 관리": page_customer_master()
    elif menu == "이력 조회": page_history()

if __name__ == "__main__":
    main()
